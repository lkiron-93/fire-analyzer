"""
Enhanced FIRE Scraper Module
Financial Institution Regulatory Extractor - Core extraction logic
Supports multiple companies, filing types, and Call Reports with MDRM integration
"""

import requests
import pandas as pd
from bs4 import BeautifulSoup
import re
import os
import json
import time
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import numpy as np
from urllib.parse import urljoin, urlparse
import xml.etree.ElementTree as ET
import logging
from logging.handlers import RotatingFileHandler

# Configure logging
def setup_logging(company_name=None, log_level=logging.DEBUG):
    """Setup comprehensive logging for FIRE scraper"""
    # Create logs directory if it doesn't exist
    log_dir = os.path.join(os.path.dirname(__file__), "logs")
    os.makedirs(log_dir, exist_ok=True)
    
    # Generate log filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    company_prefix = f"{company_name}_" if company_name else ""
    log_filename = f"{company_prefix}fire_extraction_{timestamp}.log"
    log_filepath = os.path.join(log_dir, log_filename)
    
    # Create formatter
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # Setup file handler with rotation
    file_handler = RotatingFileHandler(
        log_filepath,
        maxBytes=10*1024*1024,  # 10MB
        backupCount=5,
        encoding='utf-8'
    )

    file_handler.setLevel(log_level)
    file_handler.setFormatter(formatter)
    
    # Setup console handler (optional - can remove if you only want file logging)
    console_handler = logging.StreamHandler()
    console_handler.setLevel(log_level)
    console_handler.setFormatter(formatter)
    
    # Setup logger
    logger = logging.getLogger('FIRE')
    logger.setLevel(log_level)
    logger.handlers.clear()  # Clear any existing handlers
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    # Log the startup
    logger.info("="*80)
    logger.info(f"🔥 FIRE Analyzer Started - Log File: {log_filename}")
    logger.info(f"Company: {company_name if company_name else 'Unknown'}")
    logger.info("="*80)
    
    return logger, log_filepath


class CallReportNumberFormatter:
    """Utility class for consistent number formatting in Call Reports"""
    
    @staticmethod
    def format_amount(amount_str, in_thousands=True):
        """Format amounts according to Call Report conventions"""
        if not amount_str or amount_str.strip() == '':
            return ''
        
        # Clean the amount
        amount_str = str(amount_str).strip()
        
        # Handle special cases
        if amount_str.upper() in ['NA', 'N/A', '-']:
            return amount_str
        
        # Remove existing formatting
        cleaned = amount_str.replace(',', '').replace('$', '')
        
        # Handle parentheses for negatives
        if cleaned.startswith('(') and cleaned.endswith(')'):
            cleaned = '-' + cleaned[1:-1]
        
        try:
            # Convert to number
            num = float(cleaned)
            
            # Format based on value
            if num == 0:
                return '0'
            elif abs(num) < 1000:
                return f'{int(num):,}'
            else:
                # Standard thousands format
                return f'{int(num):,}'
        except:
            # Return original if can't parse
            return amount_str


class RCONDictionary:
    """
    Manages MDRM dictionary for Call Report code lookups
    """
    
    def __init__(self, dictionary_path=None, logger=None):
        """
        Initialize the RCON Dictionary
        
        Args:
            dictionary_path: Path to call_report_codes.json file
            logger: Logger instance
        """
        self.dictionary = {}
        self.loaded = False
        self.logger = logger or logging.getLogger('FIRE')
        
        # Default path
        if dictionary_path is None:
            dictionary_path = os.path.join(os.path.dirname(__file__), "dictionaries", "call_report_codes.json")
        
        self.load_dictionary(dictionary_path)
    
    def load_dictionary(self, dictionary_path):
        """Load the MDRM dictionary from JSON file"""
        try:
            if os.path.exists(dictionary_path):
                with open(dictionary_path, 'r', encoding='utf-8') as f:
                    self.dictionary = json.load(f)
                self.loaded = True
                self.logger.info(f"✓ Loaded MDRM dictionary with {len(self.dictionary)} codes from: {dictionary_path}")
            else:
                self.logger.warning(f"⚠️ MDRM dictionary not found at: {dictionary_path}")
        except Exception as e:
            self.logger.error(f"✗ Error loading MDRM dictionary: {str(e)}")
    
    def lookup_code(self, code):
        """
        Look up a code in the dictionary
        
        Args:
            code: The RCON/RCFD/etc. code to look up
            
        Returns:
            str: Description if found, empty string otherwise
        """
        if not self.loaded or not code:
            return ""
        
        # Clean the code (remove any whitespace)
        code = str(code).strip().upper()
        
        # Direct lookup
        if code in self.dictionary:
            return self.dictionary[code]
        
        return ""
    
    def get_description_or_default(self, description, code):
        """
        Return existing description or lookup from dictionary
        
        Args:
            description: Current description (may be blank)
            code: The code to lookup if description is blank
            
        Returns:
            str: Either the original description or dictionary lookup
        """
        # Check if description is effectively blank
        if not description or description.strip() == "" or description.strip() == ".":
            # Try to get from dictionary
            dict_description = self.lookup_code(code)
            if dict_description:
                return dict_description
        
        return description


class SECEdgarAPI:
    """Interface to SEC EDGAR API for fetching filings"""
    
    BASE_URL = "https://www.sec.gov"
    ARCHIVES_URL = f"{BASE_URL}/Archives/edgar/data"
    CIK_LOOKUP_URL = f"{BASE_URL}/files/company_tickers.json"
    
    def __init__(self, logger=None):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'FIRE Analyzer lukewkiron@gmail.com',
            'Accept-Encoding': 'gzip, deflate',
            'Accept': 'application/json'
        })
        self.rate_limit_delay = 0.1  # SEC rate limit compliance
        self.logger = logger or logging.getLogger('FIRE')
        
    def search_company(self, query, search_type='ticker'):
        """Search for company by ticker, name, or CIK"""
        try:
            # Get company tickers mapping
            response = self.session.get(self.CIK_LOOKUP_URL)
            response.raise_for_status()
            companies = response.json()
            
            results = []
            query_lower = query.lower()
            
            for cik_str, company_data in companies.items():
                if search_type == 'ticker':
                    if company_data['ticker'].lower() == query_lower:
                        results.append({
                            'cik': str(company_data['cik_str']).zfill(10),
                            'ticker': company_data['ticker'],
                            'name': company_data['title']
                        })
                elif search_type == 'name':
                    if query_lower in company_data['title'].lower():
                        results.append({
                            'cik': str(company_data['cik_str']).zfill(10),
                            'ticker': company_data['ticker'],
                            'name': company_data['title']
                        })
                elif search_type == 'cik':
                    if str(company_data['cik_str']) == query or str(company_data['cik_str']).zfill(10) == query:
                        results.append({
                            'cik': str(company_data['cik_str']).zfill(10),
                            'ticker': company_data['ticker'],
                            'name': company_data['title']
                        })
            
            time.sleep(self.rate_limit_delay)
            return results
            
        except Exception as e:
            self.logger.error(f"Error searching company: {e}")
            return []
    
    def get_filings(self, cik, filing_type='10-K', count=10):
        """Get recent filings for a company"""
        try:
            # Format CIK
            cik = str(cik).zfill(10)
            
            # Remove leading zeros for the API call
            cik_clean = str(int(cik))
            submissions_url = f"https://data.sec.gov/submissions/CIK{cik_clean.zfill(10)}.json"
            response = self.session.get(submissions_url)
            response.raise_for_status()
            
            data = response.json()
            filings = data['filings']['recent']
            
            results = []
            for i in range(len(filings['form'])):
                if filing_type == 'All Types' or filings['form'][i] == filing_type:
                    results.append({
                        'form': filings['form'][i],
                        'filing_date': filings['filingDate'][i],
                        'accession_number': filings['accessionNumber'][i].replace('-', ''),
                        'primary_document': filings['primaryDocument'][i],
                        'url': f"{self.ARCHIVES_URL}/{cik}/{filings['accessionNumber'][i].replace('-', '')}/{filings['primaryDocument'][i]}"
                    })
                
                if len(results) >= count:
                    break
            
            time.sleep(self.rate_limit_delay)
            return results
            
        except Exception as e:
            self.logger.error(f"Error getting filings: {e}")
            self.logger.error(f"Attempted URL: {submissions_url}")
            self.logger.error(f"CIK used: {cik}")
            self.logger.error(f"CIK clean used: {cik_clean}")
    
            # Try with the raw CIK number (without padding)
            try:
                alt_url = f"https://data.sec.gov/submissions/CIK{cik_clean}.json"
                self.logger.info(f"Trying alternative URL: {alt_url}")
                response = self.session.get(alt_url)
                response.raise_for_status()
                data = response.json()
                
                # Process the data (copy the same logic from above)
                filings = data['filings']['recent']
                results = []
                for i in range(len(filings['form'])):
                    if filing_type == 'All Types' or filings['form'][i] == filing_type:
                        results.append({
                            'form': filings['form'][i],
                            'filing_date': filings['filingDate'][i],
                            'accession_number': filings['accessionNumber'][i].replace('-', ''),
                            'primary_document': filings['primaryDocument'][i],
                            'url': f"{self.ARCHIVES_URL}/{cik}/{filings['accessionNumber'][i].replace('-', '')}/{filings['primaryDocument'][i]}"
                        })
                        if len(results) >= count:
                            break
                
                time.sleep(self.rate_limit_delay)
                return results
        
            except Exception as e2:
                self.logger.error(f"Alternative URL also failed: {e2}")
                return []


class CallReportAPI:
    """Interface to FFIEC Call Report API"""
    
    BASE_URL = "https://cdr.ffiec.gov/public/PWS/DownloadBulkData"
    
    def __init__(self, logger=None):
        self.session = requests.Session()
        self.logger = logger or logging.getLogger('FIRE')
        
    def get_call_report(self, rssd_id, report_date):
        """Fetch Call Report data for a bank"""
        try:
            # Format date as YYYYMMDD
            date_str = report_date.strftime('%Y%m%d')
            
            # Construct API URL
            url = f"{self.BASE_URL}/{rssd_id}/{date_str}"
            
            response = self.session.get(url)
            response.raise_for_status()
            
            # Parse Call Report data (would need specific implementation)
            return self._parse_call_report(response.content)
            
        except Exception as e:
            self.logger.error(f"Error fetching Call Report: {e}")
            return None
    
    def _parse_call_report(self, content):
        """Parse Call Report XML/CSV data"""
        # Implementation would depend on actual Call Report format
        # This is a placeholder
        return {
            'tables': [],
            'metadata': {}
        }


class EnhancedFIREScraper:
    """Enhanced FIRE scraper with multi-company and filing type support"""
    
    def __init__(self, company_info=None, filing_url=None, local_file_path=None):
        """
        Initialize the enhanced FIRE scraper
        
        Args:
            company_info (dict): Company information (ticker, name, cik)
            filing_url (str): URL to the regulatory filing
            local_file_path (str): Path to local file
        """
        self.company_info = company_info or {}
        self.filing_url = filing_url
        self.local_file_path = local_file_path
        self.soup = None
        self.tables = []
        self.metadata = {}
        
        # Setup logging
        company_name = self.company_info.get('ticker', 'Unknown')
        self.logger, self.log_filepath = setup_logging(company_name)
        self.logger.info(f"Initializing FIRE scraper for {company_name}")
        
        # Initialize APIs with logger
        self.sec_api = SECEdgarAPI(self.logger)
        self.call_api = CallReportAPI(self.logger)
        
        # Initialize MDRM dictionary for Call Reports with logger
        self.rcon_dictionary = RCONDictionary(logger=self.logger)
        
        # Initialize collapsed cell counter
        self.collapsed_cell_count = 0
        
        # Add RC Balance Sheet tracking (ADD THESE TWO LINES)
        self.rc_balance_sheet_active = False
        self.rc_balance_sheet_name = ""


        
    def set_company(self, ticker=None, name=None, cik=None):
        """Set company information"""
        if ticker:
            results = self.sec_api.search_company(ticker, 'ticker')
        elif name:
            results = self.sec_api.search_company(name, 'name')
        elif cik:
            results = self.sec_api.search_company(cik, 'cik')
        else:
            return False
            
        if results:
            self.company_info = results[0]
            return True
        return False
    
    def get_filing_url(self, filing_type='10-K', date=None):
        """Get URL for specific filing"""
        if not self.company_info.get('cik'):
            return None
            
        filings = self.sec_api.get_filings(
            self.company_info['cik'],
            filing_type,
            count=1 if date else 10
        )
        
        if filings:
            if date:
                # Find filing closest to requested date
                for filing in filings:
                    filing_date = datetime.strptime(filing['filing_date'], '%Y-%m-%d')
                    if filing_date <= date:
                        return filing['url']
            else:
                # Return most recent
                return filings[0]['url']
        
        return None
    
    def load_filing(self):
        """Load the filing from URL or local file"""
        try:
            if self.local_file_path:
                # Check if it's a PDF file
                if self.local_file_path.lower().endswith('.pdf'):
                    # For PDF files, we'll handle them differently
                    self.logger.info(f"✓ Detected PDF file: {self.local_file_path}")
                    return True  # Skip BeautifulSoup parsing for PDFs
                else:
                    # Load from local file (for HTML/XBRL files)
                    with open(self.local_file_path, 'r', encoding='utf-8') as file:
                        content = file.read()
                    self.logger.info(f"✓ Loaded filing from local file: {self.local_file_path}")
            
            elif self.filing_url:
                    # Download from URL
                    response = self.sec_api.session.get(self.filing_url)
                    response.raise_for_status()
                    content = response.text
                    self.logger.info(f"✓ Downloaded filing from: {self.filing_url}")
                
            else:
                # Try to get filing URL
                url = self.get_filing_url()
                if url:
                    response = self.sec_api.session.get(url)
                    response.raise_for_status()
                    content = response.text
                    self.logger.info(f"✓ Downloaded filing from SEC EDGAR")
                else:
                    self.logger.error("✗ No filing URL available")
                    return False
            
            # Detect file type and use appropriate parser
            if self.local_file_path and self.local_file_path.lower().endswith('.xbrl'):
                # Parse as XML for XBRL files
                self.soup = BeautifulSoup(content, 'xml')
                self.is_xbrl = True
                self.logger.info("✓ Detected XBRL file, using XML parser")
            else:
                # Parse as HTML for SEC filings
                self.soup = BeautifulSoup(content, 'lxml')
                self.is_xbrl = False
                self.logger.info("✓ Detected HTML file, using HTML parser")
            
            # Extract metadata
            self._extract_metadata()
            
            return True
            
        except Exception as e:
            self.logger.error(f"✗ Error loading filing: {str(e)}")
            return False
    
    def _extract_metadata(self):
        """Extract filing metadata"""
        self.metadata = {
            'company': self.company_info.get('name', 'Unknown'),
            'ticker': self.company_info.get('ticker', 'N/A'),
            'cik': self.company_info.get('cik', 'N/A'),
            'filing_date': None,
            'period_end': None,
            'form_type': None
        }
        
        # Try to extract from HTML
        # Look for XBRL context information
        contexts = self.soup.find_all(['xbrli:context', 'context'])
        for context in contexts:
            period = context.find(['xbrli:period', 'period'])
            if period:
                instant = period.find(['xbrli:instant', 'instant'])
                if instant:
                    self.metadata['period_end'] = instant.text
                    break
    
    def check_mdrm_dictionary_status(self):
        """Check if MDRM dictionary is loaded and working"""
        if hasattr(self, 'rcon_dictionary') and self.rcon_dictionary.loaded:
            self.logger.info(f"✓ MDRM dictionary loaded with {len(self.rcon_dictionary.dictionary)} codes")
            # Test a known code
            test_code = "RCON2170"
            test_result = self.rcon_dictionary.lookup_code(test_code)
            if test_result:
                self.logger.info(f"✓ Test lookup successful: {test_code} = {test_result}")
            else:
                self.logger.warning(f"⚠️ Test lookup failed for {test_code}")
        else:
            self.logger.error("✗ MDRM dictionary not loaded")
    
    def identify_financial_sections(self):
        """Identify sections containing financial tables with improved detection"""
        financial_keywords = [
            # Primary financial statements
            'consolidated statements of income',
            'consolidated statements of operations',
            'consolidated balance sheets',
            'consolidated statements of cash flows',
            'consolidated statements of stockholders',
            'consolidated statements of shareholders',
            'consolidated statements of equity',
            'consolidated statements of comprehensive income',
            
            # Condensed versions
            'condensed consolidated',
            
            # Other common sections
            'statements of income',
            'balance sheets',
            'cash flows',
            'stockholders equity',
            'shareholders equity',
            
            # Specific financial data
            'earnings per share',
            'segment information',
            'quarterly financial data',
            'selected financial data',
            'supplemental financial information',
            
            # Banking specific
            'net interest income',
            'credit card',
            'allowance for loan',
            'allowance for credit losses',
            'regulatory capital',
            'tier 1 capital',
            'risk-weighted assets',
            
            # Fair value and other disclosures
            'fair value measurements',
            'derivative instruments',
            'investment securities',
            'loans and leases',
            'deposits',
            
            # Notes sections
            'notes to consolidated',
            'note 1',
            'note 2',
            'summary of significant accounting'
        ]
        
        sections = []
        processed_elements = set()
        
        # Search through various HTML elements
        for element in self.soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'p', 'div', 'span']):
            # Skip if already processed
            if element in processed_elements:
                continue
                
            text = element.get_text().lower().strip()
            
            # Skip very long text blocks
            if len(text) > 300:
                continue
            
            # Check for financial keywords
            for keyword in financial_keywords:
                if keyword in text:
                    sections.append({
                        'element': element,
                        'text': text[:100],  # Truncate for display
                        'keyword': keyword,
                        'priority': self._get_section_priority(keyword)
                    })
                    processed_elements.add(element)
                    break
        
        # Sort by priority
        sections.sort(key=lambda x: x['priority'])
        
        self.logger.info(f"✓ Found {len(sections)} potential financial sections")
        return sections
    
    def _get_section_priority(self, keyword):
        """Assign priority to sections for better organization"""
        priority_map = {
            'consolidated statements of income': 1,
            'consolidated statements of operations': 1,
            'consolidated balance sheets': 2,
            'consolidated statements of cash flows': 3,
            'consolidated statements of stockholders': 4,
            'consolidated statements of shareholders': 4,
            'consolidated statements of equity': 4,
        }
        
        for key, priority in priority_map.items():
            if key in keyword:
                return priority
        
        return 10  # Default priority for other sections
    
    def extract_tables_from_section(self, section_element, section_name):
        """Extract tables with improved detection and parsing"""
        tables = []
        tables_found = 0
        
        # Define search radius
        search_radius = 100  # Number of elements to search forward
        
        current_element = section_element
        elements_checked = 0
        
        while current_element and elements_checked < search_radius:
            elements_checked += 1
            
            # Look for tables in current element
            if current_element.name == 'table':
                table_elements = [current_element]
            else:
                table_elements = current_element.find_all('table', limit=5)
            
            for table in table_elements:
                if self.is_financial_table(table):
                    parsed_data = self.parse_table_with_formatting(table)
                    if parsed_data and len(parsed_data['data']) > 2:  # Minimum viable table
                        table_name = self._generate_table_name(section_name, tables_found)
                        
                        tables.append({
                            'name': table_name,
                            'data': parsed_data,
                            'section': section_name,
                            'metadata': {
                                'rows': len(parsed_data['data']),
                                'columns': len(parsed_data['data'][0]) if parsed_data['data'] else 0,
                                'has_numbers': self._table_has_numbers(parsed_data)
                            }
                        })
                        tables_found += 1
            
            # Move to next sibling
            current_element = current_element.find_next_sibling()
            
            # Stop if we hit another major section
            if current_element and current_element.name in ['h1', 'h2', 'h3']:
                text = current_element.get_text().lower()
                if any(marker in text for marker in ['item ', 'part ', 'signatures', 'exhibit']):
                    break
        
        return tables
    
    def _generate_table_name(self, section_name, table_index):
        """Generate meaningful table names"""
        # Clean section name
        clean_section = re.sub(r'[^\w\s]', '', section_name)
        clean_section = clean_section.replace(' ', '_')[:40]
        
        # Add company ticker if available
        prefix = self.company_info.get('ticker', 'TABLE')
        
        return f"{prefix}_{clean_section}_{table_index + 1}"
    
    def _table_has_numbers(self, parsed_data):
        """Check if table contains numerical data"""
        for row in parsed_data['data']:
            for cell in row:
                if re.search(r'\d+[,.]?\d*', str(cell)):
                    return True
        return False
    
    def is_financial_table(self, table):
        """Enhanced financial table detection"""
        # Get table text
        table_text = table.get_text().lower()
        
        # Skip empty tables
        if len(table_text.strip()) < 20:
            return False
        
        # Check table structure
        rows = table.find_all('tr')
        if len(rows) < 3:
            return False
        
        # Count cells
        total_cells = len(table.find_all(['td', 'th']))
        if total_cells < 6:  # At least 2x3 table
            return False
        
        # Financial indicators
        financial_indicators = [
            '$', 'million', 'thousand', 'billion',
            'revenue', 'income', 'expense', 'cost',
            'asset', 'liability', 'equity', 'capital',
            'cash', 'debt', 'loan', 'deposit',
            'shares', 'earnings', 'loss', 'profit',
            'balance', 'total', 'interest', 'tax',
            'gross', 'net', 'operating', 'investing',
            'financing', 'continuing', 'discontinued',
            '%', 'percent', 'rate', 'ratio',
            'allowance', 'provision', 'reserve'
        ]
        
        # Count indicators
        indicator_count = sum(1 for indicator in financial_indicators if indicator in table_text)
        
        # Count numeric patterns
        numeric_patterns = re.findall(r'[\$\(]?[\d,]+\.?\d*[%\)]?', table_text)
        numeric_count = len(numeric_patterns)
        
        # Count year patterns (like 2023, 2024)
        year_patterns = re.findall(r'\b20[0-9]{2}\b', table_text)
        has_years = len(year_patterns) >= 2
        
        # Decision logic
        if indicator_count >= 3 and numeric_count >= 5:
            return True
        
        if has_years and numeric_count >= 10:
            return True
        
        if indicator_count >= 5:
            return True
        
        return False
    
    def clean_cell_text(self, text):
        """Enhanced cell text cleaning"""
        if not text:
            return ''
        
        # Convert to string
        text = str(text)
        
        # Remove extra whitespace
        text = ' '.join(text.split())
        
        # Handle special characters
        text = text.replace('\xa0', ' ')  # Non-breaking space
        text = text.replace('\n', ' ')
        
        # Handle footnote references
        text = re.sub(r'\s*\([a-z]\)', '', text)  # Remove (a), (b), etc.
        text = re.sub(r'\s*\(\d+\)', '', text)  # Remove (1), (2), etc.
        text = re.sub(r'\s*\*+', '', text)  # Remove asterisks
        
        # Handle parentheses for negative numbers
        if re.match(r'^\s*\(\s*[\d,]+\.?\d*\s*\)\s*$', text):
            # Convert (123) to -123
            text = '-' + re.sub(r'[^\d,.]', '', text)
        
        # Clean currency symbols
        text = re.sub(r'\$\s+', '$', text)
        
        # Handle percentage signs
        text = re.sub(r'\s+%', '%', text)
        
        # Final trim
        text = text.strip()
        
        return text
    
    def parse_table_with_formatting(self, table):
        """Enhanced table parsing with better structure preservation"""
        try:
            rows = table.find_all('tr')
            if not rows:
                return None
            
            parsed_table = {
                'data': [],
                'formatting': [],
                'merged_cells': [],
                'styles': [],
                'column_widths': [],
                'row_heights': []
            }
            
            # First pass: determine table structure
            max_cols = 0
            for row in rows:
                cells = row.find_all(['td', 'th'])
                col_count = sum(int(cell.get('colspan', 1)) for cell in cells)
                max_cols = max(max_cols, col_count)
            
            # Second pass: parse data
            for row_idx, row in enumerate(rows):
                cells = row.find_all(['td', 'th'])
                row_data = []
                row_formatting = []
                row_styles = []
                
                col_idx = 0
                for cell in cells:
                    # Extract and clean text
                    cell_text = self.clean_cell_text(cell.get_text())
                    
                    # Get formatting
                    formatting_info = self.extract_cell_formatting(cell)
                    
                    # Get dimensions
                    colspan = int(cell.get('colspan', 1))
                    rowspan = int(cell.get('rowspan', 1))
                    
                    # Add cell data
                    row_data.append(cell_text)
                    row_formatting.append(formatting_info)
                    row_styles.append(self.extract_cell_styles(cell))
                    
                    # Track merged cells
                    if colspan > 1 or rowspan > 1:
                        parsed_table['merged_cells'].append({
                            'row': row_idx,
                            'col': col_idx,
                            'colspan': colspan,
                            'rowspan': rowspan,
                            'value': cell_text
                        })
                    
                    # Fill colspan
                    for i in range(1, colspan):
                        row_data.append('')
                        row_formatting.append({'merged': True})
                        row_styles.append({})
                    
                    col_idx += colspan
                
                # Pad row to max columns
                while len(row_data) < max_cols:
                    row_data.append('')
                    row_formatting.append({})
                    row_styles.append({})
                
                parsed_table['data'].append(row_data)
                parsed_table['formatting'].append(row_formatting)
                parsed_table['styles'].append(row_styles)
            
            # Post-process: detect header rows
            self._detect_header_rows(parsed_table)
            
            return parsed_table
            
        except Exception as e:
            self.logger.warning(f"Warning: Error parsing table - {str(e)}")
            return None
    
    def _detect_header_rows(self, parsed_table):
        """Detect which rows are headers"""
        if not parsed_table['data']:
            return
        
        # Check first few rows
        for i in range(min(3, len(parsed_table['data']))):
            row = parsed_table['data'][i]
            formatting = parsed_table['formatting'][i]
            
            # Count non-numeric cells
            non_numeric = 0
            for cell in row:
                if cell and not re.match(r'^[\$\-\+]?[\d,]+\.?\d*[%]?$', cell):
                    non_numeric += 1
            
            # If mostly non-numeric, likely a header
            if non_numeric > len(row) * 0.7:
                for j, fmt in enumerate(formatting):
                    fmt['is_header'] = True
    
    def extract_cell_formatting(self, cell):
        """Extract comprehensive formatting information"""
        formatting = {
            'is_bold': False,
            'is_italic': False,
            'is_underline': False,
            'is_header': False,
            'text_align': 'left',
            'is_number': False,
            'is_currency': False,
            'is_percentage': False,
            'is_date': False,
            'indent_level': 0,
            'font_size': 'normal'
        }
        
        # Check cell type
        formatting['is_header'] = cell.name == 'th'
        
        # Check for styling
        style = cell.get('style', '')
        classes = cell.get('class', [])
        
        # Bold detection
        if any(tag in str(cell) for tag in ['<b>', '<strong>']):
            formatting['is_bold'] = True
        if 'font-weight: bold' in style or 'font-weight:bold' in style:
            formatting['is_bold'] = True
        
        # Alignment
        align = cell.get('align', '')
        if align:
            formatting['text_align'] = align
        elif 'text-align: center' in style:
            formatting['text_align'] = 'center'
        elif 'text-align: right' in style:
            formatting['text_align'] = 'right'
        
        # Analyze content
        text = cell.get_text().strip()
        
        # Currency detection
        if '$' in text or 'usd' in text.lower():
            formatting['is_currency'] = True
            formatting['is_number'] = True
        
        # Number detection
        elif re.match(r'^[\(\-]?[\d,]+\.?\d*\)?$', text.replace(' ', '')):
            formatting['is_number'] = True
        
        # Percentage detection
        if '%' in text:
            formatting['is_percentage'] = True
            formatting['is_number'] = True
        
        # Date detection
        if re.match(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', text):
            formatting['is_date'] = True
        
        # Indentation detection
        # Check for leading spaces or CSS padding
        if 'padding-left' in style:
            padding = re.search(r'padding-left:\s*(\d+)', style)
            if padding:
                formatting['indent_level'] = int(padding.group(1)) // 20
        
        return formatting
    
    def extract_cell_styles(self, cell):
        """Extract visual styles from cell"""
        styles = {
            'background_color': None,
            'border': None,
            'font_color': None
        }
        
        style_attr = cell.get('style', '')
        
        # Background color
        bg_match = re.search(r'background-color:\s*([^;]+)', style_attr)
        if bg_match:
            styles['background_color'] = bg_match.group(1).strip()
        
        # Border
        if 'border' in style_attr:
            styles['border'] = True
        
        # Font color
        color_match = re.search(r'color:\s*([^;]+)', style_attr)
        if color_match:
            styles['font_color'] = color_match.group(1).strip()
        
        return styles
    
    def scrape_all_tables(self):
        """Main method to scrape all financial tables"""
        if not self.load_filing():
            return False

        if hasattr(self, 'is_xbrl') and self.is_xbrl:
            # Handle XBRL files (Call Reports)
            self.logger.info("🔍 Extracting XBRL data from Call Report...")
            return self.extract_xbrl_tables()
        elif self.local_file_path and self.local_file_path.lower().endswith('.pdf'):
            # Handle PDF files (Call Reports)
            self.logger.info("📄 Detected PDF Call Report file...")
            return self.extract_pdf_tables()
        elif self.local_file_path and (self.local_file_path.lower().endswith('.sdf') or 
                                       self.local_file_path.lower().endswith('.txt')):
            # Handle SDF files (Call Reports)
            self.logger.info("📊 Detected SDF Call Report file...")
            return self.extract_sdf_tables()
        else:
            # Handle HTML files (SEC filings)
            self.logger.info("🔍 Searching for financial sections...")
            sections = self.identify_financial_sections()
            
            all_tables = []
            
            for i, section in enumerate(sections):
                section_name = f"section_{i + 1}_{section['keyword'].replace(' ', '_')}"
                self.logger.info(f"📊 Processing: {section['text'][:60]}...")
                
                section_tables = self.extract_tables_from_section(section['element'], section_name)
                all_tables.extend(section_tables)
            
            self.tables = all_tables
            self.logger.info(f"✓ Found {len(all_tables)} financial tables total")
            
            return True
        
    def extract_xbrl_tables(self):
        """Extract financial data from XBRL Call Report files"""
        try:
            # Find all numeric data elements in XBRL
            self.logger.info("🔍 Parsing XBRL structure...")
            
            # Look for XBRL facts (numeric data points)
            facts = self.soup.find_all(['us-gaap:*', 'call:*', 'xbrli:*'], recursive=True)
            if not facts:
                # Fallback: find all elements with numeric content
                facts = self.soup.find_all(string=lambda text: text and self._is_numeric(text.strip()))
            
            self.logger.info(f"✓ Found {len(facts)} XBRL data points")
            
            if len(facts) == 0:
                return False
            
            # Group data by context/period
            grouped_data = self._group_xbrl_data(facts)
            
            # Create tables from grouped data
            tables = []
            for group_name, data_points in grouped_data.items():
                if len(data_points) > 3:  # Only create tables with sufficient data
                    table = self._create_table_from_xbrl_data(group_name, data_points)
                    if table:
                        tables.append(table)
            
            self.tables = tables
            self.logger.info(f"✓ Created {len(tables)} tables from XBRL data")
            
            return len(tables) > 0
            
        except Exception as e:
            self.logger.error(f"✗ Error extracting XBRL data: {str(e)}")
            return False
        
    #STARTTTTT
    def extract_pdf_tables(self):
        """Extract financial data from Call Report PDF files with schedule preservation"""
        try:
            import pdfplumber
            self.logger.info("🔍 Parsing PDF Call Report structure...")
            
            all_schedules = {}  # Dictionary to store schedules
            current_schedule = None
            current_schedule_data = []
            current_schedule_name = ""
            
            with pdfplumber.open(self.local_file_path) as pdf:
                self.logger.info(f"✓ Found {len(pdf.pages)} pages in PDF")
                
                for page_num, page in enumerate(pdf.pages):
                    self.logger.info(f"📄 Processing page {page_num + 1}...")
                    
                    # Extract text to identify schedules
                    page_text = page.extract_text()
                    
                    # Enhanced pattern to capture all schedule types including Parts
                    # First try a specific pattern for RC Balance Sheet
                    rc_balance_match = re.search(
                        r'^Schedule\s+RC\s*[-–]?\s*(?:Consolidated\s+)?Balance\s+Sheet',
                        page_text,
                        re.MULTILINE | re.IGNORECASE
                    )

                    if rc_balance_match:
                        # Save previous schedule if exists
                        if current_schedule and current_schedule_data:
                            # Don't save empty RC schedule if we're in RC Balance Sheet mode
                            if not (current_schedule == "RC" and self.rc_balance_sheet_active and len(current_schedule_data) == 0):
                                all_schedules[current_schedule] = {
                                    'name': f"Schedule {current_schedule}",
                                    'title': current_schedule_name,
                                    'data': current_schedule_data
                                }
                                self.logger.info(f"  ✓ Saved Schedule {current_schedule} with {len(current_schedule_data)} rows")
                        
                        # Handle RC Balance Sheet specifically
                        base_schedule = 'RC'
                        sub_letter = ''
                        part_info = ''
                        schedule_code = 'RC'
                        schedule_title = 'Consolidated Balance Sheet'

                        # Set the RC Balance Sheet flag
                        self.rc_balance_sheet_active = True
                        self.rc_balance_sheet_name = schedule_title
                        
                        self.logger.info(f"✓ Detected Schedule: {schedule_code} - {schedule_title}")
                        self.logger.info("🎯 RC BALANCE SHEET MODE ACTIVATED - Will apply special processing to tables")
                        self.logger.info(f"📍 Activated on page {page_num + 1}")
                        
                        # Start new schedule
                        current_schedule = schedule_code
                        current_schedule_name = schedule_title
                        current_schedule_data = []
                        
                    else:
                        # Try general pattern for other schedules
                        schedule_match = re.search(
                            r'^Schedule\s+(RC|RI)(?:-([A-Z]))?(?:\s+(Part\s+[IVX]+))?\s*[-–]?\s*(.+?)(?:\s*\(Form Type[^)]+\))?$',
                            page_text, 
                            re.MULTILINE | re.IGNORECASE
                        )
                        
                        if schedule_match:
                            # Parse the new schedule FIRST
                            base_schedule = schedule_match.group(1).upper()
                            sub_letter = schedule_match.group(2).upper() if schedule_match.group(2) else ""
                            part_info = schedule_match.group(3) if schedule_match.group(3) else ""
                            title_text = schedule_match.group(4).strip() if schedule_match.group(4) else ""

                            # Construct schedule code
                            if sub_letter and part_info:
                                schedule_code = f"{base_schedule}-{sub_letter} {part_info}"
                            elif sub_letter:
                                schedule_code = f"{base_schedule}-{sub_letter}"
                            else:
                                schedule_code = base_schedule
                            
                            schedule_title = ' '.join(title_text.split())
                            
                            # Check if we need to deactivate RC Balance Sheet mode
                            if self.rc_balance_sheet_active and schedule_code != "RC":
                                self.logger.info(f"🔚 RC BALANCE SHEET MODE DEACTIVATED on page {page_num + 1}")
                                self.logger.info(f"   Reason: New schedule detected: {schedule_code}")
                                self.logger.info(f"   Processed {len(current_schedule_data)} rows while in RC mode")
                                self.rc_balance_sheet_active = False
                                self.rc_balance_sheet_name = ""
                            
                            # Save previous schedule if exists
                            if current_schedule and current_schedule_data:
                                all_schedules[current_schedule] = {
                                    'name': f"Schedule {current_schedule}",
                                    'title': current_schedule_name,
                                    'data': current_schedule_data
                                }
                                self.logger.info(f"  ✓ Saved Schedule {current_schedule} with {len(current_schedule_data)} rows")
                            
                            self.logger.info(f"✓ Detected Schedule: {schedule_code} - {schedule_title} on page {page_num + 1}")
                            
                            # Debug info
                            self.logger.debug(f"DEBUG: Schedule transition - From: {current_schedule} To: {schedule_code}")
                            self.logger.debug(f"DEBUG: RC Balance Sheet active: {self.rc_balance_sheet_active}")
                            
                            # Start new schedule
                            current_schedule = schedule_code
                            current_schedule_name = schedule_title
                            current_schedule_data = []
                            
                    # Extract tables from the page
                    page_tables = page.extract_tables()

                    #DEBUG LINE ADDED 7-7-2025
                    if self.rc_balance_sheet_active and page_tables:
                        self.logger.info("=== RAW PDF TABLE DEBUG ===")
                        for table_idx, table in enumerate(page_tables[:1]):  # Just first table
                            self.logger.info(f"Table {table_idx} raw structure:")
                            for row_idx, row in enumerate(table[:5]):  # First 5 rows
                                self.logger.info(f"  Row {row_idx}: {len(row)} cells")
                                for cell_idx, cell in enumerate(row):
                                    if cell:
                                        self.logger.info(f"    Cell [{cell_idx}]: {repr(cell[:100])}")

                    for table_idx, table_data in enumerate(page_tables):
                        if not table_data or len(table_data) < 2:
                            continue

                        # NEW: Enhanced debugging - show what's in the table
                        self.logger.debug(f"📊 Table {table_idx} on page {page_num + 1}:")
                        self.logger.debug(f"   Dimensions: {len(table_data)} rows x {len(table_data[0]) if table_data else 0} cols")
                        
                        # NEW: Check for potential collapsed cells BEFORE processing
                        potential_collapsed = False
                        for row_idx, row in enumerate(table_data[:20]):  # Check first 20 rows
                            for cell in row:
                                if cell and isinstance(cell, str):
                                    # Look for multiple MDRM codes in one cell
                                    if len(re.findall(r'(RCFD|RCON|RCFN|RCOA|RCOB|RCOC|RCOD)\d+', str(cell))) > 1:
                                        potential_collapsed = True
                                        self.logger.warning(f"🎯 POTENTIAL COLLAPSED CELL on page {page_num + 1}, row {row_idx}: {str(cell)[:100]}...")
                        
                        if potential_collapsed:
                            self.logger.warning(f"⚠️ PAGE {page_num + 1} CONTAINS POTENTIAL COLLAPSED CELLS!")

                        # Check if we're in RC Balance Sheet mode
                        if self.rc_balance_sheet_active:
                            self.logger.info(f"!!! SPECIAL RC PROCESSING ACTIVATED on PAGE {page_num + 1} !!!")
                            self.logger.info(f"!!! Current Schedule: {current_schedule} !!!")
                            
                            # Try to extract tables with better detection
                            page_tables = page.extract_tables()
                            
                            # Check if we got proper 4-column extraction
                            got_full_table = False
                            if page_tables:
                                for table in page_tables:
                                    if table and len(table) > 0 and len(table[0]) >= 4:
                                        got_full_table = True
                                        break
                            
                            if not got_full_table:
                                self.logger.warning(f"⚠️ Standard extraction only got {len(page_tables[0][0]) if page_tables and page_tables[0] else 0} columns, using enhanced method...")
                                
                                # Extract words with positioning for enhanced method
                                words = page.extract_words(
                                    x_tolerance=3,
                                    y_tolerance=3,
                                    keep_blank_chars=True,
                                    use_text_flow=False,
                                    horizontal_ltr=True,
                                    vertical_ttb=True
                                )
                                
                                # Use enhanced extraction
                                processed_rows = self._extract_rc_balance_sheet_from_words(words, page_num)
                                
                                if processed_rows:
                                    current_schedule_data.extend(processed_rows)
                                    self.logger.info(f"  ✅ Extracted {len(processed_rows)} rows using enhanced word analysis")
                                continue
                            
                            # If we got a proper 4-column table, process it
                            for table_idx, table_data in enumerate(page_tables):
                                if not table_data or len(table_data) < 2:
                                    continue
                                
                                self.logger.debug(f"📊 Table {table_idx} on page {page_num + 1}:")
                                self.logger.debug(f"   Dimensions: {len(table_data)} rows x {len(table_data[0]) if table_data else 0} cols")
                                
                                if len(table_data[0]) >= 4:
                                    # Process complete 4-column table
                                    processed_table = self._process_complete_rc_balance_sheet_table(table_data)
                                    self.logger.info(f"  ✅ Processing with 4-column handler")
                                else:
                                    # Fall back to enhanced extraction
                                    words = page.extract_words()
                                    processed_table = self._extract_rc_balance_sheet_from_words(words, page_num)
                                    self.logger.info(f"  ⚠️ Falling back to word extraction due to incomplete columns")
                                
                                if processed_table:
                                    current_schedule_data.extend(processed_table)
                                    self.logger.info(f"    Added {len(processed_table)} rows to {current_schedule}")

                                #END
                          
                        elif current_schedule and current_schedule != "RC":  # We've moved past RC
                            # Turn off RC Balance Sheet mode when we hit a different schedule
                            if self.rc_balance_sheet_active:
                                self.logger.info("📊 Exiting RC Balance Sheet mode")
                                self.rc_balance_sheet_active = False
                                self.rc_balance_sheet_name = ""
                            
                            # Standard processing for other schedules
                            processed_table = self._process_call_report_table(table_data)
                        else:
                            # Standard processing
                            processed_table = self._process_call_report_table(table_data)

                        if processed_table:
                            if current_schedule:
                                current_schedule_data.extend(processed_table)
                                self.logger.info(f"    Added {len(processed_table)} rows to {current_schedule}")

                               
            # Don't forget the last schedule
            if current_schedule and current_schedule_data:
                all_schedules[current_schedule] = {
                    'name': f"Schedule {current_schedule}",
                    'title': current_schedule_name,
                    'data': current_schedule_data
                }
                self.logger.info(f"  ✓ Saved final schedule {current_schedule} with {len(current_schedule_data)} rows")
                # Add PDF processing summary
                self.logger.info("="*80)
                self.logger.info("PDF PROCESSING SUMMARY")
                self.logger.info(f"Total pages processed: {len(pdf.pages)}")
                self.logger.info(f"Total schedules found: {len(all_schedules)}")
                self.logger.info(f"Collapsed cells detected: {self.collapsed_cell_count}")

                # List all schedules found
                self.logger.info("\nSchedules extracted:")
                for sched_code, sched_data in all_schedules.items():
                    self.logger.info(f"  - {sched_code}: {sched_data['title']} ({len(sched_data['data'])} rows)")

                # Check if RC Balance Sheet mode is still active (shouldn't be)
                if self.rc_balance_sheet_active:
                    self.logger.warning("⚠️ RC Balance Sheet mode still active at end of document!")
                    self.logger.warning(f"   Last schedule processed: {current_schedule}")
                else:
                    self.logger.info("✓ RC Balance Sheet mode properly deactivated")

                # Summary of potential issues
                if self.collapsed_cell_count == 0:
                    self.logger.warning("\n⚠️ NO COLLAPSED CELLS DETECTED!")
                    self.logger.warning("   Expected: Collapsed cells in RC Balance Sheet")
                    self.logger.warning("   Check the log for 'POTENTIAL COLLAPSED CELL' warnings")
                else:
                    self.logger.info(f"\n✓ Successfully detected and processed {self.collapsed_cell_count} collapsed cells")

                self.logger.info("="*80)

            # Convert schedules to our table format
            self.tables = []

            # Enhanced schedule name mapping with proper FFIEC titles
            schedule_name_map = {
                'RC': 'Consolidated Balance Sheet',
                'RC-A': 'Cash and Balances Due',
                'RC-B': 'Securities',
                'RC-C Part I': 'Loans and Leases',
                'RC-C Part II': 'Small Business and Farm Loans',
                'RC-D': 'Trading Assets and Liabilities',
                'RC-E Part I': 'Deposits',
                'RC-E Part II': 'Deposits in Foreign Offices',
                'RC-F': 'Other Assets',
                'RC-G': 'Other Liabilities',
                'RC-H': 'Selected Balance Sheet Items',
                'RC-K': 'Quarterly Averages',
                'RC-L': 'Derivatives and Off-Balance Sheet',
                'RC-M': 'Memoranda',
                'RC-N': 'Past Due and Nonaccrual',
                'RC-O': 'Other Data for Deposit Insurance',
                'RC-P': '1-4 Family Residential Mortgage',
                'RC-Q': 'Fair Value Measurements',
                'RC-R Part I': 'Regulatory Capital Components',
                'RC-R Part II': 'Risk-Weighted Assets',
                'RC-S': 'Servicing Securitization Asset Sales',
                'RC-T': 'Fiduciary and Related Services',
                'RC-V': 'Variable Interest Entities',
                'RI': 'Income Statement',
                'RI-A': 'Changes in Bank Equity Capital',
                'RI-B Part I': 'Charge-offs and Recoveries',
                'RI-B Part II': 'Allowance for Credit Losses',
                'RI-C': 'Disaggregated Data on Income',
                'RI-E': 'Explanations'
            }

            for schedule_code, schedule_info in all_schedules.items():
                # Create descriptive tab name
                base_schedule = schedule_code.split(' Part ')[0]
                
                # Get the proper name from our mapping
                if schedule_code in schedule_name_map:
                    proper_name = schedule_name_map[schedule_code]
                elif base_schedule in schedule_name_map:
                    proper_name = schedule_name_map[base_schedule]
                else:
                    # Use the title from the PDF if not in our map
                    proper_name = schedule_info.get('title', schedule_code)
                
                # Create the tab name (Excel limit is 31 chars)
                if " Part " in schedule_code:
                    # Include part info in name
                    tab_name = f"{schedule_code}"
                else:
                    tab_name = schedule_code
                
                # Ensure it fits Excel's limit
                if len(tab_name) > 31:
                    tab_name = tab_name[:31]
                
                # Create formatted table structure
                formatted_table = self._format_call_report_schedule_enhanced(
                    schedule_info['data'], 
                    schedule_code,
                    proper_name
                )
                
                self.tables.append({
                    'name': tab_name,
                    'full_name': f"{schedule_code} - {proper_name}",
                    'data': formatted_table,
                    'section': schedule_info['name'],
                    'metadata': {
                        'rows': len(schedule_info['data']),
                        'columns': 3,  # Description, Code, Amount
                        'has_numbers': True,
                        'schedule_code': schedule_code,
                        'schedule_title': proper_name
                    }
                })
                
                self.logger.info(f"  ✓ Created table: {tab_name} - {proper_name} ({len(schedule_info['data'])} rows)")
            
            self.logger.info(f"✓ Extracted {len(self.tables)} schedules from PDF")
            return len(self.tables) > 0
            
        except Exception as e:
            self.logger.error(f"✗ Error extracting PDF data: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
            #ALIGN end
        
   
    def process_rc_balance_sheet_table(self, df, schedule_name):
        """Special processing for RC Balance Sheet to handle collapsed cells using DataFrame."""
        logger = self.logger  # Use the instance logger
        
        logger.info(f">>> ENTERING process_rc_balance_sheet_table for {schedule_name}")
        logger.debug(f"DataFrame shape: {df.shape}")
        logger.debug(f"DataFrame columns: {df.columns.tolist()}")
        
        # Log first few rows to see structure
        logger.debug("First 3 rows of dataframe:")
        for idx in range(min(3, len(df))):
            logger.debug(f"Row {idx}: {df.iloc[idx].tolist()}")
        
        # Pattern to detect multiple MDRM codes in a single cell
        pattern = re.compile(r'(RCFD|RCON|RCFN|RCOA|RCOB|RCOC|RCOD)([A-Z0-9]+)\s+([-]?[\d,]+)')
        
        new_rows = []
        collapsed_cell_count = 0
        rows_to_remove = []
        
        # Check each cell for multiple MDRM codes
        for row_idx in range(len(df)):
            row_has_collapsed_cell = False
            
            for col_idx in range(len(df.columns)):
                cell = df.iloc[row_idx, col_idx]
                
                # Debug log every cell in first 5 rows
                if row_idx < 5:
                    logger.debug(f"Cell at ({row_idx}, {col_idx}): {repr(cell)}")
                
                if pd.isna(cell):
                    continue
                    
                # Look for cells with multiple MDRM codes
                cell_str = str(cell)
                matches = list(pattern.findall(cell_str))
                
                # Debug log for cells that might have codes
                if any(code in cell_str for code in ['RCFD', 'RCON', 'RCFN']):
                    logger.debug(f"Potential MDRM cell at ({row_idx}, {col_idx}): {cell_str[:100]}...")
                    logger.debug(f"Regex matches found: {len(matches)}")
                
                if len(matches) > 1:
                    logger.info(f"🔍 COLLAPSED CELL FOUND at row {row_idx}, col {col_idx}")
                    logger.info(f"   Original content: {cell_str[:200]}...")
                    logger.info(f"   Found {len(matches)} MDRM codes")
                    
                    collapsed_cell_count += 1
                    row_has_collapsed_cell = True
                    
                    # Extract line item and description from other columns
                    line_item = ""
                    description = ""
                    
                    # Look for description in cells before the collapsed cell
                    for desc_idx in range(col_idx):
                        cell_val = df.iloc[row_idx, desc_idx]
                        if pd.notna(cell_val) and str(cell_val).strip():
                            desc_text = str(cell_val).strip()
                            # Check if this looks like a line item number
                            if re.match(r'^\d+\.?[a-z]?\.?$', desc_text):
                                line_item = desc_text
                            else:
                                description = desc_text
                                break
                    
                    # Create a new row for each code/amount pair
                    for i, (prefix, code_num, amount) in enumerate(matches):
                        mdrm_code = f"{prefix}{code_num}"
                        logger.debug(f"   Splitting out: {mdrm_code} = {amount}")
                        
                        # Create new row data
                        new_row = df.iloc[row_idx].copy()
                        
                        # For RC Balance Sheet, we expect 4 columns:
                        # [Line Item, Description, MDRM Code, Amount]
                        if len(df.columns) >= 4:
                            new_row.iloc[0] = line_item if i == 0 else ""  # Only first split gets line item
                            new_row.iloc[1] = description if i == 0 else self.rcon_dictionary.lookup_code(mdrm_code)
                            new_row.iloc[2] = mdrm_code
                            new_row.iloc[3] = amount
                        else:
                            # Fallback for different column structure
                            new_row.iloc[col_idx] = mdrm_code
                            if col_idx + 1 < len(df.columns):
                                new_row.iloc[col_idx + 1] = amount
                        
                        new_rows.append(new_row)
                    
                    # Mark this row for removal
                    rows_to_remove.append(row_idx)
                    break  # Don't check other cells in this row
        
        # Remove rows that had collapsed cells
        if rows_to_remove:
            df = df.drop(rows_to_remove).reset_index(drop=True)
        
        # Add new rows
        if new_rows:
            new_df = pd.DataFrame(new_rows)
            df = pd.concat([df, new_df], ignore_index=True)
            logger.info(f"✓ Split {collapsed_cell_count} collapsed cells into {len(new_rows)} new rows")
        
        self.collapsed_cell_count += collapsed_cell_count
        
        logger.info(f"📊 Final DataFrame shape: {df.shape}")
        logger.info(f"📊 Total collapsed cells found in this table: {collapsed_cell_count}")
        
        return df

    def _process_rc_balance_sheet_table(self, table_data, page_text):
        """Special processing for RC Balance Sheet tables with 4-column output"""
        processed_rows = []

        # ADD THESE DEBUG LINES
        self.logger.debug("=" * 80)
        self.logger.debug("ENTERING RC BALANCE SHEET SPECIAL PROCESSING")
        self.logger.debug(f"Table has {len(table_data)} rows")
        self.logger.debug("=" * 80)
        
        self.logger.info("Starting RC Balance Sheet processing")
        
        for row_idx, row in enumerate(table_data):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            # Skip header rows and footnotes
            skip_row = False
            for cell in row:
                if cell and any(skip_text in str(cell) for skip_text in [
                    'Dollar amounts in thousands',
                    'Schedule RC',
                    'All schedules',
                    'Form Type',
                    'Last Updated',
                    'Report Date'
                ]):
                    skip_row = True
                    break
            
            if skip_row:
                continue
            
            self.logger.debug(f"Processing row {row_idx} with {len(row)} cells")
            
            # Check EACH CELL for multiple codes
            row_has_multi_codes = False
            
            for cell_idx, cell in enumerate(row):
                if cell is None:
                    continue
                    
                cell_text = str(cell).strip()
                if not cell_text:
                    continue
                
                self.logger.debug(f"  Cell {cell_idx} content: {cell_text[:100]}...")
                
                # Check if THIS CELL contains multiple RCFD/RCON codes
                # Updated regex to handle space-separated codes on same line
                multi_code_matches = re.findall(
                    r'(RCFD|RCON|RCFN|RCOA|RCOB|RCOC|RCOD)([A-Z0-9]+)\s+([-]?[\d,]+)',
                    cell_text
                )
                
                if len(multi_code_matches) > 1:
                    row_has_multi_codes = True
                    self.collapsed_cell_count += 1  # Increment counter
                    
                    # ENHANCED DEBUG LOGGING
                    self.logger.debug(f"🔍 COLLAPSED CELL FOUND at row {row_idx}, col {cell_idx}:")  # Changed from warning to debug
                    self.logger.debug(f"   Original: {cell_text[:100]}...")
                    self.logger.debug(f"   Found {len(multi_code_matches)} MDRM codes")
                    for match in multi_code_matches:
                        code = match[0] + match[1]
                        amount = match[2]
                        self.logger.debug(f"   - {code}: {amount}")
                    
                    # Get description from other cells in the row
                    description = ""
                    line_item = ""
                    
                    # Look for description in cells before this one
                    for desc_idx in range(cell_idx):
                        if row[desc_idx]:
                            desc_text = str(row[desc_idx]).strip()
                            if desc_text and not re.match(r'^[\d,]+$', desc_text):
                                # Extract line item and description
                                line_item_match = re.match(r'^(\d+\.?[a-z]?\.?)\s+(.+)', desc_text)
                                if line_item_match:
                                    line_item = line_item_match.group(1).rstrip('.')
                                    description = line_item_match.group(2)
                                else:
                                    description = desc_text
                                break
                    
                    # Process each code/amount pair
                    for i, (prefix, code_num, amount) in enumerate(multi_code_matches):
                        rcon_code = f"{prefix}{code_num}"
                        
                        # Get description for this specific code
                        item_description = ""
                        
                        if i == 0 and description:
                            item_description = description
                        else:
                            dict_description = self.rcon_dictionary.lookup_code(rcon_code)
                            if dict_description:
                                item_description = dict_description
                                self.logger.info(f"    ✓ MDRM lookup for {rcon_code}: {dict_description}")
                            else:
                                item_description = f"Line item for {rcon_code}"
                        
                        # Handle parentheses for negative amounts
                        if amount.startswith('(') and amount.endswith(')'):
                            amount = '-' + amount[1:-1]
                        
                        processed_rows.append({
                            'line_item': line_item if i == 0 else "",
                            'description': item_description,
                            'code': rcon_code,
                            'amount': amount.strip(),
                            'is_section_header': False,
                            'is_total': any(word in item_description.upper() for word in ['TOTAL', 'SUBTOTAL', 'NET'])
                        })
                    
                    break  # Don't process other cells in this row
            
            # If no multi-code cells found, process normally
            if not row_has_multi_codes:
                # Original single-code processing logic
                # Join all cells for analysis
                full_row_text = ' '.join(str(cell).strip() for cell in row if cell)
                
                self.logger.debug(f"Processing single-code row -> {full_row_text[:100]}...")
                
                # First check if this is a section header (description only, no codes)
                if not re.search(r'(RCFD|RCON|RCFN|RCOA|RCOB|RCOC|RCOD)[A-Z0-9]+', full_row_text):
                    # This might be a section header
                    description = full_row_text.strip()
                    
                    # Extract line item if present
                    line_item = ""
                    line_item_match = re.match(r'^(\d+\.?[a-z]?\.?)\s+(.+)', description)
                    if line_item_match:
                        line_item = line_item_match.group(1).rstrip('.')
                        description = line_item_match.group(2)
                    
                    if len(description) > 3 and not description.isdigit():
                        processed_rows.append({
                            'line_item': line_item,
                            'description': description,
                            'code': '',
                            'amount': '',
                            'is_section_header': True,
                            'is_total': False
                        })
                    continue
                
                # Try standard single-code patterns
                patterns = [
                    # Main numbered items: "1. Description RCFD1234 123,456"
                    r'^(\d+)\.\s+(.+?)\s+(RCFD|RCON|RCFN)([A-Z0-9]+)\s+([-]?[\d,]+)',
                    # Sub-items: "a. Description RCFD1234 123,456" or "1.a. Description..."
                    r'^(\d*\.?[a-z])\.\s+(.+?)\s+(RCFD|RCON|RCFN)([A-Z0-9]+)\s+([-]?[\d,]+)',
                    # Nested items: "(1) Description RCFD1234 123,456"
                    r'^(\(\d+\))\s+(.+?)\s+(RCFD|RCON|RCFN)([A-Z0-9]+)\s+([-]?[\d,]+)',
                    # Roman numerals: "(i) Description RCFD1234 123,456"
                    r'^(\([ivx]+\))\s+(.+?)\s+(RCFD|RCON|RCFN)([A-Z0-9]+)\s+([-]?[\d,]+)',
                    # Generic: "Description RCFD1234 123,456"
                    r'^(.+?)\s+(RCFD|RCON|RCFN)([A-Z0-9]+)\s+([-]?[\d,]+)',
                ]
                
                matched = False
                for pattern_idx, pattern in enumerate(patterns):
                    match = re.search(pattern, full_row_text)
                    if match:
                        # Extract based on pattern type
                        if pattern_idx < 4:  # Patterns with explicit line items
                            line_item = match.group(1).rstrip('.')
                            description = match.group(2).strip()
                            code_prefix = match.group(3)
                            code_num = match.group(4)
                            amount = match.group(5)
                        else:  # Generic pattern
                            line_item = ""
                            description = match.group(1).strip()
                            code_prefix = match.group(2)
                            code_num = match.group(3)
                            amount = match.group(4)
                        
                        # Clean description
                        description = re.sub(r'\s*\([^)]+\)\s*:?\s*$', '', description)
                        description = re.sub(r'\s*\.+\s*$', '', description).strip()
                        
                        rcon_code = f"{code_prefix}{code_num}"
                        
                        # Only use MDRM if description is blank
                        if not description or description in ["", ".", "-"]:
                            dict_description = self.rcon_dictionary.lookup_code(rcon_code)
                            if dict_description:
                                description = dict_description
                                self.logger.info(f"  ✓ Auto-populated description for {rcon_code}: {dict_description}")
                            else:
                                description = f"Line item {rcon_code}"
                        
                        # Clean amount
                        amount = amount.strip()
                        if amount.startswith('(') and amount.endswith(')'):
                            amount = '-' + amount[1:-1]
                        
                        processed_rows.append({
                            'line_item': line_item,
                            'description': description,
                            'code': rcon_code,
                            'amount': amount,
                            'is_section_header': False,
                            'is_total': any(word in description.upper() for word in ['TOTAL', 'SUBTOTAL', 'NET'])
                        })
                        matched = True
                        break
                
                if not matched:
                    # Last resort: extract any code and amount
                    code_match = re.search(r'(RCFD|RCON|RCFN)([A-Z0-9]+)', full_row_text)
                    amount_match = re.search(r'([-]?[\d,]+)', full_row_text)
                    
                    if code_match and amount_match:
                        rcon_code = f"{code_match.group(1)}{code_match.group(2)}"
                        amount = amount_match.group(1)
                        
                        # Extract any text before the code
                        pre_code_text = full_row_text[:code_match.start()].strip()
                        
                        # Extract line item if present
                        line_item = ""
                        description = pre_code_text
                        
                        line_item_match = re.match(r'^(\d+\.?[a-z]?\.?)\s+(.+)', pre_code_text)
                        if line_item_match:
                            line_item = line_item_match.group(1).rstrip('.')
                            description = line_item_match.group(2)
                        
                        # Only use MDRM if description is blank
                        if not description:
                            description = self.rcon_dictionary.lookup_code(rcon_code) or f"Item {rcon_code}"
                        
                        # Clean amount
                        if amount.startswith('(') and amount.endswith(')'):
                            amount = '-' + amount[1:-1]
                        
                        processed_rows.append({
                            'line_item': line_item,
                            'description': description,
                            'code': rcon_code,
                            'amount': amount,
                            'is_section_header': False,
                            'is_total': False
                        })
        
        # Final validation: ensure no empty descriptions
        for row in processed_rows:
            if not row['description'] or row['description'].strip() == "":
                if row['code']:
                    row['description'] = f"Line item {row['code']}"
                else:
                    row['description'] = "Unlabeled item"
        
        self.logger.info(f"  Processed {len(processed_rows)} rows from RC Balance Sheet")
        return processed_rows
    
    #First New Method Added 7-7-2025
    def _extract_rc_balance_sheet_from_words(self, words, page_num):
        """Extract RC Balance Sheet data by analyzing word positions"""
        self.logger.info(f"🔍 Analyzing word positions for RC Balance Sheet on page {page_num + 1}")
        
        # Group words by approximate Y position (rows)
        rows_by_y = {}
        for word in words:
            y_pos = round(word['top'], 1)  # Round to nearest 0.1 point
            if y_pos not in rows_by_y:
                rows_by_y[y_pos] = []
            rows_by_y[y_pos].append(word)
        
        # Sort rows by Y position
        sorted_y_positions = sorted(rows_by_y.keys())
        
        processed_rows = []
        
        for y_pos in sorted_y_positions:
            row_words = sorted(rows_by_y[y_pos], key=lambda w: w['x0'])
            
            # Skip if too few words
            if len(row_words) < 2:
                continue
            
            # Combine words into text segments based on X position
            row_text = ' '.join(w['text'] for w in row_words)
            
            # Skip header/footer rows
            if any(skip in row_text for skip in ['Dollar amounts', 'Schedule RC', 'Form Type', 'Last Updated']):
                continue
            
            # Pattern for RC Balance Sheet rows with all 4 components
            patterns = [
                # Full pattern with line item, description, code, and amount
                r'^(\d+\.?[a-z]?\.?)\s+(.+?)\s+(RCFD|RCON|RCFN)([A-Z0-9]+)\s+([-]?[\d,]+)\s*(\d+\.?[a-z]?\.?)?\s*$',
                # Pattern without trailing line reference
                r'^(\d+\.?[a-z]?\.?)\s+(.+?)\s+(RCFD|RCON|RCFN)([A-Z0-9]+)\s+([-]?[\d,]+)\s*$',
                # Pattern for items without line numbers
                r'^(.+?)\s+(RCFD|RCON|RCFN)([A-Z0-9]+)\s+([-]?[\d,]+)\s*$'
            ]
            
            matched = False
            for pattern in patterns:
                match = re.search(pattern, row_text)
                if match:
                    if len(match.groups()) >= 5:  # Full pattern
                        line_item = match.group(1)
                        description = match.group(2).strip()
                        code = f"{match.group(3)}{match.group(4)}"
                        amount = match.group(5)
                    else:  # Pattern without line number
                        line_item = ""
                        description = match.group(1).strip()
                        code = f"{match.group(2)}{match.group(3)}"
                        amount = match.group(4)
                    
                    # Clean up description
                    description = re.sub(r'\s*\.+\s*$', '', description)
                    description = re.sub(r'\s+', ' ', description)
                    
                    # Only use MDRM lookup if description is empty
                    if not description or description in ["", ".", "-"]:
                        dict_description = self.rcon_dictionary.lookup_code(code)
                        if dict_description:
                            description = dict_description
                            self.logger.debug(f"  ✓ MDRM lookup for {code}: {dict_description}")
                    
                    # Clean amount
                    if amount.startswith('(') and amount.endswith(')'):
                        amount = '-' + amount[1:-1]
                    
                    processed_rows.append({
                        'line_item': line_item,
                        'description': description,
                        'code': code,
                        'amount': amount,
                        'is_section_header': False,
                        'is_total': any(word in description.upper() for word in ['TOTAL', 'SUBTOTAL', 'NET'])
                    })
                    
                    matched = True
                    break
            
            if not matched:
                # Check if this is a section header (no MDRM code)
                if not re.search(r'(RCFD|RCON|RCFN)[A-Z0-9]+', row_text) and len(row_text) > 5:
                    # Extract line item if present
                    line_item_match = re.match(r'^(\d+\.?[a-z]?\.?)\s+(.+)', row_text)
                    if line_item_match:
                        line_item = line_item_match.group(1)
                        description = line_item_match.group(2)
                    else:
                        line_item = ""
                        description = row_text.strip()
                    
                    if description and not description.isdigit():
                        processed_rows.append({
                            'line_item': line_item,
                            'description': description,
                            'code': '',
                            'amount': '',
                            'is_section_header': True,
                            'is_total': False
                        })
        
        self.logger.info(f"  ✅ Extracted {len(processed_rows)} rows from word analysis")
        return processed_rows
    
    #Additional New Method Added 7-7-2025
    def _process_complete_rc_balance_sheet_table(self, table_data):
        """Process RC Balance Sheet table when we have all 4 columns"""
        processed_rows = []
        
        for row_idx, row in enumerate(table_data):
            if not row or len(row) < 4:
                continue
            
            # Skip header rows
            skip_row = False
            for cell in row:
                if cell and any(skip_text in str(cell) for skip_text in [
                    'Dollar amounts', 'Schedule RC', 'Form Type', 'Last Updated'
                ]):
                    skip_row = True
                    break
            
            if skip_row:
                continue
            
            # Extract 4 columns
            line_item = str(row[0]).strip() if row[0] else ""
            description = str(row[1]).strip() if row[1] else ""
            code = str(row[2]).strip() if row[2] else ""
            amount = str(row[3]).strip() if row[3] else ""
            
            # Validate MDRM code
            if code and not re.match(r'^(RCFD|RCON|RCFN)[A-Z0-9]+$', code):
                continue
            
            # Use MDRM lookup only if description is empty
            if code and (not description or description in ["", ".", "-"]):
                dict_description = self.rcon_dictionary.lookup_code(code)
                if dict_description:
                    description = dict_description
            
            # Clean amount
            if amount.startswith('(') and amount.endswith(')'):
                amount = '-' + amount[1:-1]
            
            processed_rows.append({
                'line_item': line_item,
                'description': description,
                'code': code,
                'amount': amount,
                'is_section_header': not code,
                'is_total': any(word in description.upper() for word in ['TOTAL', 'SUBTOTAL', 'NET'])
            })
        
        return processed_rows

    def _process_call_report_table(self, table_data):
        """Process a Call Report table with enhanced structure detection"""
        processed_rows = []
        
        for row_idx, row in enumerate(table_data):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            # Initialize variables for this row
            full_row_text = []
            has_rcon_code = False
            rcon_code = ""
            amount = ""
            
            # Collect all non-empty cells
            for cell in row:
                if cell is not None:
                    cell_text = str(cell).strip()
                    if cell_text:
                        # Check if this is an RCON/RCFD code
                        if re.match(r'^(RCFD|RCON|RIAD|RCFN|RCOA|RCOB|RCOC|RCOD)[A-Z0-9]+$', cell_text):
                            rcon_code = cell_text
                            has_rcon_code = True
                        # Check if this is an amount
                        elif re.match(r'^-?\d{1,3}(,\d{3})*(\.\d+)?$', cell_text):
                            amount = cell_text
                        else:
                           # This is descriptive text
                           full_row_text.append(cell_text)
           
            # Build the description from all text cells
            description = ' '.join(full_row_text)
            
            # Clean up the description
            description = re.sub(r'\s+\d+\.\s*$', '', description)
            description = re.sub(r'\s+\d+\.[a-z]\.\s*\d+$', '', description)
            description = re.sub(r'\s*\.+\s*$', '', description)
            
            # MDRM DICTIONARY LOOKUP
            if rcon_code and (not description or description.strip() == ""):
                dict_description = self.rcon_dictionary.lookup_code(rcon_code)
                if dict_description:
                    description = dict_description
                    self.logger.info(f"  ✓ Auto-populated description for {rcon_code}: {dict_description}")
            
            # Only add rows that have meaningful content
            if description or rcon_code or amount:
                # Enhanced indentation detection
                indent_level = 0
                is_section_header = False
                
                # Check for main numbered items (e.g., "1. Cash and...")
                if re.match(r'^\d+\.\s+[A-Z]', description):
                    indent_level = 0
                    # Mark as section header if it's a main category
                    is_section_header = True
                # Check for lettered sub-items (e.g., "a. Noninterest-bearing...")
                elif re.match(r'^[a-z]\.\s+', description) or re.match(r'^\d+\.[a-z]\.\s+', description):
                    indent_level = 1
                    is_section_header = False
                # Check for numbered sub-sub-items (e.g., "(1) ...")
                elif re.match(r'^\(\d+\)', description):
                    indent_level = 2
                    is_section_header = False
                # Check for Roman numeral items
                elif re.match(r'^\([ivx]+\)', description, re.IGNORECASE):
                    indent_level = 3
                    is_section_header = False
                else:
                    # Check if line starts with lowercase (often indicates continuation)
                    if description and description[0].islower():
                        indent_level = 1
                    is_section_header = False
                
                # Detect if this is a total row
                is_total = any(word in description.upper() for word in 
                                ['TOTAL', 'SUBTOTAL', 'NET INCOME', 'NET LOSS', 
                                'GROSS', 'BALANCE', 'AGGREGATE'])
                
                processed_rows.append({
                    'description': description.strip(),
                    'code': rcon_code,
                    'amount': amount,
                    'indent': indent_level,
                    'is_section_header': is_section_header,
                    'is_total': is_total
                })
        
        return processed_rows
    
    def _format_call_report_schedule(self, schedule_data):
        """Format Call Report schedule data for Excel output with MDRM lookup"""
        # Create headers
        formatted_table = {
            'data': [['Line Item Description', 'Code', 'Dollar Amounts (in thousands)']],
            'formatting': [[
                {'is_header': True, 'is_bold': True, 'text_align': 'center'},
                {'is_header': True, 'is_bold': True, 'text_align': 'center'},
                {'is_header': True, 'is_bold': True, 'text_align': 'center'}
            ]],
            'merged_cells': [],
            'styles': [],
            'column_widths': [60, 15, 25]  # Suggested column widths
        }
        
        # Add data rows
        for row_data in schedule_data:
            # Apply MDRM lookup one more time in case it was missed
            description = row_data['description']
            if row_data['code'] and (not description or description.strip() == ""):
                dict_description = self.rcon_dictionary.lookup_code(row_data['code'])
                if dict_description:
                    description = dict_description
            
            # Create row with proper indentation
            if row_data['indent'] > 0:
                description = "  " * row_data['indent'] + description
            
            formatted_table['data'].append([
                description,
                row_data['code'],
                row_data['amount']
            ])
            
            # Add formatting
            row_formatting = [
                {'is_bold': False, 'text_align': 'left', 'indent_level': row_data['indent']},
                {'is_bold': False, 'text_align': 'center', 'is_code': True, 'font_color': '#0066CC'},
                {'is_bold': False, 'text_align': 'right', 'is_number': True, 'is_currency': True}
            ]
            
            formatted_table['formatting'].append(row_formatting)
        
        return formatted_table
    
    def _format_call_report_schedule_enhanced(self, schedule_data, schedule_code, schedule_title):
        """Enhanced formatting for Call Report schedules with 4-column layout for RC"""
        
        # Determine if this is RC schedule that needs 4 columns
        is_rc_schedule = schedule_code == "RC"
        
        # Create formatted table with enhanced structure
        formatted_table = {
            'data': [],
            'formatting': [],
            'merged_cells': [],
            'styles': [],
            'column_widths': [10, 60, 15, 25] if is_rc_schedule else [65, 12, 25],
            'freeze_row': 3,  # Freeze after header rows
            'schedule_info': {
                'code': schedule_code,
                'title': schedule_title
            }
        }
        
        # Add schedule header (merged across all columns)
        num_cols = 4 if is_rc_schedule else 3
        header_row = [f"Schedule {schedule_code} - {schedule_title}"] + [""] * (num_cols - 1)
        formatted_table['data'].append(header_row)
        formatted_table['formatting'].append([
            {'is_header': True, 'is_bold': True, 'text_align': 'center', 'font_size': 14}
        ] + [{'is_header': True}] * (num_cols - 1))
        formatted_table['merged_cells'].append({
            'row': 0,
            'col': 0,
            'colspan': num_cols,
            'rowspan': 1,
            'value': f"Schedule {schedule_code} - {schedule_title}"
        })
        
        # Add blank row for spacing
        formatted_table['data'].append([""] * num_cols)
        formatted_table['formatting'].append([{}] * num_cols)
        
        # Add column headers
        if is_rc_schedule:
            headers = ['Line Item', 'Description', 'MDRM #', 'Dollar Amounts\n(in thousands)']
        else:
            headers = ['Line Item Description', 'MDRM #', 'Dollar Amounts\n(in thousands)']
        
        formatted_table['data'].append(headers)
        formatted_table['formatting'].append([
            {'is_header': True, 'is_bold': True, 'text_align': 'center', 'background_color': 'E6E6E6'},
            {'is_header': True, 'is_bold': True, 'text_align': 'center', 'background_color': 'E6E6E6'},
            {'is_header': True, 'is_bold': True, 'text_align': 'center', 'background_color': 'E6E6E6'},
            {'is_header': True, 'is_bold': True, 'text_align': 'center', 'background_color': 'E6E6E6', 'wrap_text': True}
        ] if is_rc_schedule else [
            {'is_header': True, 'is_bold': True, 'text_align': 'center', 'background_color': 'E6E6E6'},
            {'is_header': True, 'is_bold': True, 'text_align': 'center', 'background_color': 'E6E6E6'},
            {'is_header': True, 'is_bold': True, 'text_align': 'center', 'background_color': 'E6E6E6', 'wrap_text': True}
        ])
        
        # Process data rows
        for row_idx, row_data in enumerate(schedule_data):
            # Determine if this is a total/subtotal row
            is_total_row = row_data.get('is_total', False)
            
            # Clean amount value
            amount = self._clean_amount_value(row_data['amount'])
            
            if is_rc_schedule:
                # 4-column format for RC schedule
                row = [
                    row_data.get('line_item', ''),
                    row_data['description'],
                    row_data['code'],
                    amount
                ]
                
                # Row formatting for 4 columns
                row_formatting = [
                    {
                        'is_bold': is_total_row,
                        'text_align': 'center',
                        'border_top': is_total_row
                    },
                    {
                        'is_bold': is_total_row,
                        'text_align': 'left',
                        'border_top': is_total_row
                    },
                    {
                        'is_bold': False,
                        'text_align': 'center',
                        'is_code': True,
                        'font_color': '0066CC'
                    },
                    {
                        'is_bold': is_total_row,
                        'text_align': 'right',
                        'is_number': True,
                        'is_currency': True,
                        'border_top': is_total_row,
                        'number_format': '#,##0'
                    }
                ]
            else:
                # 3-column format for other schedules (keep existing logic)
                # Apply MDRM lookup
                description = row_data['description']
                if row_data['code'] and (not description or description.strip() == ""):
                    dict_description = self.rcon_dictionary.lookup_code(row_data['code'])
                    if dict_description:
                        description = dict_description
                
                # Create row with proper indentation
                indent_spaces = "    " * row_data.get('indent', 0)
                formatted_description = indent_spaces + description
                
                row = [
                    formatted_description,
                    row_data['code'],
                    amount
                ]
                
                # Row formatting for 3 columns
                row_formatting = [
                    {
                        'is_bold': is_total_row,
                        'text_align': 'left',
                        'indent_level': row_data.get('indent', 0),
                        'border_top': is_total_row
                    },
                    {
                        'is_bold': False,
                        'text_align': 'center',
                        'is_code': True,
                        'font_color': '0066CC'
                    },
                    {
                        'is_bold': is_total_row,
                        'text_align': 'right',
                        'is_number': True,
                        'is_currency': True,
                        'border_top': is_total_row,
                        'number_format': '#,##0'
                    }
                ]
            
            # Add background color for major sections
            if row_data.get('is_section_header', False) and not is_total_row:
                for fmt in row_formatting:
                    fmt['background_color'] = 'F5F5F5'
            
            formatted_table['data'].append(row)
            formatted_table['formatting'].append(row_formatting)
        
        return formatted_table

    def _clean_amount_value(self, amount_str):
        """Clean and standardize amount values"""
        if not amount_str:
            return ""
        
        # Remove any whitespace
        amount_str = amount_str.strip()
        
        # Handle parentheses for negative numbers
        if amount_str.startswith('(') and amount_str.endswith(')'):
            amount_str = '-' + amount_str[1:-1]
        
        # Ensure thousands separators are present
        if re.match(r'^-?\d+$', amount_str):
            # Convert plain number to formatted number
            try:
                num = int(amount_str)
                return f"{num:,}"
            except:
                return amount_str
        
        return amount_str
    
    def _identify_schedule_sections(self, schedule_data):
        """Identify major sections within a schedule for better formatting"""
        sections = []
        current_section = None
        
        for idx, row in enumerate(schedule_data):
            if row.get('is_section_header'):
                if current_section:
                    sections.append(current_section)
                current_section = {
                    'start_idx': idx,
                    'end_idx': idx,
                    'title': row['description']
                }
            elif current_section:
                current_section['end_idx'] = idx
        
        if current_section:
            sections.append(current_section)
        
        return sections
    
    def _clean_pdf_table(self, table_data):
        """Clean and filter PDF table data"""
        cleaned_rows = []
        
        for row in table_data:
            if not row:
                continue
                
            # Clean each cell
            cleaned_row = []
            for cell in row:
                if cell is None:
                    cleaned_row.append('')
                else:
                    # Clean the cell text
                    cell_text = str(cell).strip()
                    # Remove multiple spaces
                    cell_text = ' '.join(cell_text.split())
                    cleaned_row.append(cell_text)
            
            # Skip rows that are all empty
            if any(cell for cell in cleaned_row):
                cleaned_rows.append(cleaned_row)
        
        return cleaned_rows
    
    def _identify_pdf_schedule(self, page_text, page_num):
        """Identify the schedule name from PDF page text"""
        # Look for schedule identifiers
        schedule_patterns = [
            r'Schedule\s+RC-([A-Z])\s+',
            r'Schedule\s+RC\s+-\s+',
            r'Schedule\s+RI\s+-\s+',
            r'Schedule\s+RI-([A-Z])\s+'
        ]
        
        for pattern in schedule_patterns:
            match = re.search(pattern, page_text)
            if match:
                if match.group(1) if len(match.groups()) > 0 else None:
                    return f"Schedule_RC_{match.group(1)}"
                else:
                    return "Schedule_RC"
        
        # Default naming
        return f"Page_{page_num + 1}"
    
    def _format_pdf_table(self, table_data):
        """Format PDF table data to match our standard structure"""
        formatted_table = {
            'data': table_data,
            'formatting': [],
            'merged_cells': [],
            'styles': []
        }
        
        # Add formatting for each row
        for i, row in enumerate(table_data):
            row_formatting = []
            for j, cell in enumerate(row):
                fmt = {'is_header': i == 0}  # First row is header
                
                # Check if cell contains numbers
                if self._is_numeric(str(cell)):
                    fmt['is_number'] = True
                    fmt['text_align'] = 'right'
                    
                    # Check for currency
                    if '$' in str(cell):
                        fmt['is_currency'] = True
                
                # Check for Call Report codes (RCFD, RCON, etc.)
                if re.match(r'^(RCFD|RCON|RIAD|RCFN)[A-Z0-9]+$', str(cell)):
                    fmt['is_code'] = True
                    fmt['text_align'] = 'center'
                
                row_formatting.append(fmt)
            
            formatted_table['formatting'].append(row_formatting)
        
        return formatted_table
    
    def extract_sdf_tables(self):
        """Extract financial data from Call Report SDF (Standard Data Format) files"""
        try:
            self.logger.info("🔍 Parsing SDF Call Report structure...")
            
            tables = []
            
            # Read the SDF file
            with open(self.local_file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            
            # SDF files are typically delimited text files
            # Try to detect the delimiter
            if '\t' in content[:1000]:
                delimiter = '\t'
                self.logger.info("✓ Detected tab-delimited SDF format")
            elif '|' in content[:1000]:
                delimiter = '|'
                self.logger.info("✓ Detected pipe-delimited SDF format")
            else:
                delimiter = ','
                self.logger.info("✓ Assuming comma-delimited SDF format")
            
            # Parse as CSV-like format
            import csv
            from io import StringIO
            
            csv_reader = csv.reader(StringIO(content), delimiter=delimiter)
            rows = list(csv_reader)
            
            if not rows:
                self.logger.error("✗ No data found in SDF file")
                return False
            
            self.logger.info(f"✓ Found {len(rows)} rows in SDF file")
            
            # Group data by schedule/form
            current_schedule = "Call_Report_Data"
            schedule_data = []
            
            for row in rows:
                if not row or all(cell.strip() == '' for cell in row):
                    # Empty row might indicate new section
                    if schedule_data:
                        # Save current section
                        table = self._create_sdf_table(current_schedule, schedule_data)
                        if table:
                            tables.append(table)
                        schedule_data = []
                    continue
                
                # Check if this row indicates a new schedule
                first_cell = str(row[0]).strip() if row else ''
                if 'Schedule' in first_cell or 'SCHEDULE' in first_cell:
                    # Save previous schedule data
                    if schedule_data:
                        table = self._create_sdf_table(current_schedule, schedule_data)
                        if table:
                            tables.append(table)
                    
                    # Start new schedule
                    current_schedule = first_cell
                    schedule_data = [row]
                else:
                    schedule_data.append(row)
            
            # Don't forget the last section
            if schedule_data:
                table = self._create_sdf_table(current_schedule, schedule_data)
                if table:
                    tables.append(table)
            
            self.tables = tables
            self.logger.info(f"✓ Created {len(tables)} tables from SDF data")
            
            return len(tables) > 0
            
        except Exception as e:
            self.logger.error(f"✗ Error extracting SDF data: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def _create_sdf_table(self, schedule_name, data_rows):
        """Create a table from SDF data rows"""
        if not data_rows or len(data_rows) < 2:
            return None
        
        # Clean the schedule name
        clean_name = re.sub(r'[^\w\s]', '', schedule_name).replace(' ', '_')[:40]
        
        # Format the data
        formatted_table = {
            'data': data_rows,
            'formatting': [],
            'merged_cells': [],
            'styles': []
        }
        
        # Add formatting
        for i, row in enumerate(data_rows):
            row_formatting = []
            for j, cell in enumerate(row):
                fmt = {'is_header': i == 0}
                
                # Check for numeric data
                if self._is_numeric(str(cell)):
                    fmt['is_number'] = True
                    fmt['text_align'] = 'right'
                
                row_formatting.append(fmt)
            
            formatted_table['formatting'].append(row_formatting)
        
        return {
            'name': f"COF_SDF_{clean_name}",
            'data': formatted_table,
            'section': schedule_name,
            'metadata': {
                'rows': len(data_rows),
                'columns': len(data_rows[0]) if data_rows else 0,
                'has_numbers': True
            }
        }

    def _is_numeric(self, text):
        """Check if text contains numeric data"""
        if not text:
            return False
        
        # Convert to string first
        text = str(text)
        
        # Remove common formatting
        clean_text = text.replace(',', '').replace('$', '').replace('%', '')
        clean_text = clean_text.replace('.', '').replace('(', '').replace(')', '')
        
        try:
            float(clean_text)
            return True
        except ValueError:
            return False

    def _group_xbrl_data(self, facts):
        """Group XBRL facts by context or similar categories"""
        grouped = {}
        
        for fact in facts:
            if hasattr(fact, 'name') and fact.name:
                # Use tag name as grouping key
                category = fact.name.split(':')[-1] if ':' in fact.name else fact.name
                if category not in grouped:
                    grouped[category] = []
                grouped[category].append({
                    'name': fact.name,
                    'value': fact.get_text().strip() if fact.get_text() else '',
                    'context': fact.get('contextref', ''),
                    'unit': fact.get('unitref', ''),
                    'decimals': fact.get('decimals', ''),
                })
            elif isinstance(fact, str) and self._is_numeric(fact):
                # Handle string facts
                if 'numeric_data' not in grouped:
                    grouped['numeric_data'] = []
                grouped['numeric_data'].append({
                    'name': 'NumericValue',
                    'value': fact.strip(),
                    'context': '',
                    'unit': '',
                    'decimals': '',
                })
        
        return grouped

    def _create_table_from_xbrl_data(self, group_name, data_points):
        """Create a formatted table from XBRL data points"""
        try:
            # Create table structure
            headers = ['Item', 'Value', 'Context', 'Unit']
            rows = [headers]
            
            for point in data_points:
                row = [
                    point['name'],
                    point['value'],
                    point['context'],
                    point['unit']
                ]
                rows.append(row)
            
            # Create table data structure compatible with existing export logic
            table_data = {
                'data': rows,
                'formatting': [[{'is_header': True} for _ in range(len(headers))]],  # First row is header
                'merged_cells': [],
                'styles': []
            }
            
            # Add formatting for data rows
            for i in range(1, len(rows)):
                row_formatting = []
                for j, cell in enumerate(rows[i]):
                    fmt = {'is_header': False}
                    if j == 1 and self._is_numeric(str(cell)):  # Value column
                        fmt['is_number'] = True
                        fmt['text_align'] = 'right'
                    row_formatting.append(fmt)
                table_data['formatting'].append(row_formatting)
            
            # Create table metadata
            return {
                'name': f"COF_CallReport_{group_name}",
                'data': table_data,
                'section': f"XBRL_{group_name}",
                'metadata': {
                    'rows': len(rows),
                    'columns': len(headers),
                    'has_numbers': True
                }
            }
            
        except Exception as e:
            self.logger.error(f"✗ Error creating table for {group_name}: {str(e)}")
            return None  
    
    def save_to_excel_formatted(self, output_file=None):
        """Save with enhanced formatting specifically for Call Reports"""
        if not self.tables:
            self.logger.error("No tables to save!")
            return False
        
        # Generate filename if not provided
        if not output_file:
            company = self.company_info.get('ticker', 'company')
            
            # For Call Reports, try to extract the report date
            if 'Call Report' in str(self.metadata.get('form_type', '')):
                # Try to extract date from filename or metadata
                date_match = re.search(r'(\d{1,2}[-/]\d{1,2}[-/]\d{4})', str(self.local_file_path))
                if date_match:
                    date_str = date_match.group(1).replace('/', '-').replace('-', '_')
                    output_file = f"FIRE_{company}_Call_Report_{date_str}_formatted.xlsx"  
                else:
                    date = datetime.now().strftime('%Y%m%d')
                    output_file = f"FIRE_{company}_Call_Report_{date}_formatted.xlsx"  
            else:
                date = datetime.now().strftime('%Y%m%d')
                output_file = f"FIRE_{company}_financial_tables_{date}_formatted.xlsx"  
        
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Add metadata sheet with Call Report specific info
            self._add_call_report_metadata_sheet(wb)
            
            # Add table of contents
            self._add_summary_sheet(wb)
            
            # Add each schedule as a separate sheet
            for i, table_data in enumerate(self.tables):
                sheet_name = table_data['name']
                ws = wb.create_sheet(sheet_name)
                
                # Write table with enhanced formatting
                self.write_formatted_table_to_sheet(ws, table_data['data'])
                
                # Add sheet tab color based on schedule type
                if sheet_name.startswith('RC-'):
                    ws.sheet_properties.tabColor = "4472C4"  # Blue for Balance Sheet items
                elif sheet_name.startswith('RI-'):
                    ws.sheet_properties.tabColor = "70AD47"  # Green for Income Statement items
                else:
                    ws.sheet_properties.tabColor = "FFC000"  # Orange for other schedules
                
                self.logger.info(f"💾 Saved formatted schedule: {table_data.get('full_name', table_data['name'])}")
            
            # Add summary to log
            self.logger.info("="*80)
            self.logger.info("EXTRACTION SUMMARY")
            self.logger.info(f"Total schedules extracted: {len(self.tables)}")
            self.logger.info(f"Output file: {output_file}")
            self.logger.info(f"Log file: {self.log_filepath}")
            self.logger.info(f"Collapsed cells detected and split: {self.collapsed_cell_count}")
            self.logger.info("="*80)
            
            # Also print log location to console
            print(f"\n📝 Full log saved to: {self.log_filepath}")
            
            wb.save(output_file)
            self.logger.info(f"✅ Call Report saved to: {output_file}")
            return True
            
        except Exception as e:
            import traceback
            self.logger.error(f"✗ Error saving formatted Excel: {str(e)}")
            self.logger.error("Full error traceback:")
            traceback.print_exc()
            return False
    
    def _add_metadata_sheet(self, workbook):
        """Add metadata sheet with filing information"""
        ws = workbook.create_sheet("Filing_Info", 0)
        
        # Headers
        headers = ['Property', 'Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Metadata
        metadata_rows = [
            ('Company Name', self.metadata.get('company', 'N/A')),
            ('Ticker Symbol', self.metadata.get('ticker', 'N/A')),
            ('CIK', self.metadata.get('cik', 'N/A')),
            ('Filing Date', self.metadata.get('filing_date', 'N/A')),
            ('Period End', self.metadata.get('period_end', 'N/A')),
            ('Form Type', self.metadata.get('form_type', '10-K/10-Q')),
            ('Extraction Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Tables Extracted', len(self.tables))
        ]
        
        for row_idx, (prop, value) in enumerate(metadata_rows, 2):
            ws.cell(row=row_idx, column=1, value=prop).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
    
    def _add_call_report_metadata_sheet(self, workbook):
        """Add Call Report specific metadata sheet"""
        ws = workbook.create_sheet("Report Information", 0)
        
        # Title
        ws['A1'] = "🔥 FIRE - Call Report Analysis"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:C1')
        
        # Report header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Headers
        ws['A3'] = "Report Information"
        ws['A3'].font = Font(bold=True, color="FFFFFF", size=12)
        ws['A3'].fill = header_fill
        ws.merge_cells('A3:C3')
        
        # Metadata rows
        metadata_items = [
            ('Institution Name', self.metadata.get('company', 'N/A')),
            ('RSSD ID', self.metadata.get('rssd_id', 'N/A')),
            ('Report Type', 'Call Report (FFIEC 031)'),
            ('Report Period', self.metadata.get('period_end', 'N/A')),
            ('Extraction Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Total Schedules', len(self.tables)),
            ('', ''),  # Blank row
            ('Data Quality', ''),
            ('MDRM Codes Found', sum(1 for table in self.tables 
                                    for row in table['data']['data'] 
                                    if len(row) > 1 and row[1])),  # Count codes
            ('Descriptions Populated', 'Yes - Using MDRM Dictionary'),
            ('Dictionary Version', f'{len(self.rcon_dictionary.dictionary)} codes'),
            ('Collapsed Cells Split', self.collapsed_cell_count)
        ]
        
        row_num = 4
        for label, value in metadata_items:
            if label:  # Skip blank rows for label
                ws.cell(row=row_num, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=value)
            row_num += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        
        # Add borders to data area
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=4, max_row=row_num-1, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
    
    def _add_summary_sheet(self, workbook):
        """Add enhanced summary sheet with table overview"""
        ws = workbook.create_sheet("Table of Contents", 1)
        
        # Title
        ws['A1'] = "Call Report - Table of Contents"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # Company info
        ws['A3'] = f"Institution: {self.metadata.get('company', 'N/A')}"
        ws['A4'] = f"Report Date: {self.metadata.get('period_end', 'N/A')}"
        ws['A5'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        for row in range(3, 6):
            ws[f'A{row}'].font = Font(size=11)
        
        # Headers
        headers = ['Schedule', 'Description', 'Rows', 'Tab Name']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Table listing
        for row_idx, table in enumerate(self.tables, 8):
            schedule_code = table['metadata'].get('schedule_code', '')
            schedule_title = table['metadata'].get('schedule_title', '')
            
            ws.cell(row=row_idx, column=1, value=schedule_code)
            ws.cell(row=row_idx, column=2, value=schedule_title)
            ws.cell(row=row_idx, column=3, value=table['metadata']['rows'])
            ws.cell(row=row_idx, column=4, value=table['name'])
            
            # Alternate row coloring
            if row_idx % 2 == 0:
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
                    )
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 25
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = thin_border
    
    def _generate_sheet_name(self, table_name, index):
        """Generate valid Excel sheet name"""
        # Excel sheet name limitations: 31 chars max, no special chars
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        sheet_name = table_name
        
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, '_')
        
        # Truncate to 31 characters
        if len(sheet_name) > 31:
            sheet_name = f"{sheet_name[:27]}_{index}"
        
        return sheet_name
    
    def write_formatted_table_to_sheet(self, worksheet, table_data):
        """Write table data to Excel worksheet with enhanced Call Report formatting"""
        data = table_data['data']
        formatting = table_data['formatting']
        merged_cells = table_data.get('merged_cells', [])
        styles = table_data.get('styles', [])
        
        # Define style presets
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Call Report specific styles
        schedule_header_font = Font(bold=True, size=14)
        column_header_font = Font(bold=True, size=11)
        column_header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        code_font = Font(color="0066CC", size=10)
        total_font = Font(bold=True, size=11)
        section_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Number formats
        number_format = '#,##0'  # Thousands separator, no decimals
        currency_format = '$#,##0'
        
        # Borders
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        thick_top_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Write data with formatting
        for row_idx, row in enumerate(data, 1):
            for col_idx, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Apply formatting if available
                if row_idx <= len(formatting) and col_idx <= len(formatting[row_idx - 1]):
                    fmt = formatting[row_idx - 1][col_idx - 1]
                    
                    # Schedule header row (row 1)
                    if row_idx == 1 and fmt.get('is_header'):
                        cell.font = schedule_header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Column headers (row 3)
                    elif row_idx == 3 and fmt.get('is_header'):
                        cell.font = column_header_font
                        cell.fill = column_header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = thin_border
                    
                    # Data rows
                    else:
                        # Font styles
                        if fmt.get('is_bold'):
                            cell.font = total_font
                        elif fmt.get('is_code'):
                            cell.font = code_font
                        
                        # Background color
                        if fmt.get('background_color') == 'F5F5F5':
                            cell.fill = section_fill
                        
                        # Alignment
                        h_align = fmt.get('text_align', 'left')
                        indent = fmt.get('indent_level', 0)
                        cell.alignment = Alignment(horizontal=h_align, indent=indent)
                        
                        # Number formatting
                        if fmt.get('is_currency'):
                            cell.number_format = number_format
                        elif fmt.get('is_number'):
                            cell.number_format = number_format
                        
                        # Borders
                        if fmt.get('border_top'):
                            cell.border = thick_top_border
                        else:
                            cell.border = thin_border
        
        # Handle merged cells
        for merge_info in merged_cells:
            start_row = merge_info['row'] + 1
            start_col = merge_info['col'] + 1
            end_row = start_row + merge_info['rowspan'] - 1
            end_col = start_col + merge_info['colspan'] - 1
            
            if end_row > start_row or end_col > start_col:
                try:
                    worksheet.merge_cells(
                        start_row=start_row, start_column=start_col,
                        end_row=end_row, end_column=end_col
                    )
                except:
                    pass
        
        # Set column widths
        if hasattr(table_data, 'get') and table_data.get('column_widths'):
            for i, width in enumerate(table_data['column_widths'], 1):
                col_letter = get_column_letter(i)
                worksheet.column_dimensions[col_letter].width = width
        
        # Set row heights for better readability
        worksheet.row_dimensions[1].height = 25  # Schedule header
        worksheet.row_dimensions[3].height = 30  # Column headers
        
        # Freeze panes (after header rows)
        freeze_row = table_data.get('freeze_row', 3)
        if len(data) > freeze_row:
            worksheet.freeze_panes = f'A{freeze_row + 1}'
        
        # Add print settings for better output
        worksheet.page_setup.orientation = 'portrait'
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_margins.left = 0.5
        worksheet.page_margins.right = 0.5
        worksheet.print_options.horizontalCentered = True
    
    def save_to_excel_basic(self, output_file=None):
        """Save basic Excel without formatting"""
        if not self.tables:
            self.logger.error("No tables to save!")
            return False
        
        # Generate filename if not provided
        if not output_file:
            company = self.company_info.get('ticker', 'company')
            date = datetime.now().strftime('%Y%m%d')
            output_file = f"{company}_financial_tables_{date}_basic.xlsx"
        
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Add metadata sheet
                metadata_df = pd.DataFrame([
                    ['Company', self.metadata.get('company', 'N/A')],
                    ['Ticker', self.metadata.get('ticker', 'N/A')],
                    ['CIK', self.metadata.get('cik', 'N/A')],
                    ['Tables Extracted', len(self.tables)]
                ], columns=['Property', 'Value'])
                metadata_df.to_excel(writer, sheet_name='Info', index=False)
                
                # Add each table
                for i, table_data in enumerate(self.tables):
                    df = pd.DataFrame(table_data['data']['data'])
                    sheet_name = self._generate_sheet_name(table_data['name'], i)
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                    self.logger.info(f"💾 Saved table: {table_data['name']} ({len(df)} rows)")
            
            self.logger.info(f"✅ All tables saved to: {output_file}")
            return True
            
        except Exception as e:
            self.logger.error(f"✗ Error saving to Excel: {str(e)}")
            return False
    
    def save_to_csv(self, output_dir=None):
        """Save each table as CSV with company-specific naming"""
        if not self.tables:
            self.logger.error("No tables to save!")
            return False
        
        # Generate directory name if not provided
        if not output_dir:
            company = self.company_info.get('ticker', 'company')
            date = datetime.now().strftime('%Y%m%d')
            output_dir = f"FIRE_{company}_financial_tables_{date}_csv"
        
        try:
            os.makedirs(output_dir, exist_ok=True)
            
            # Save metadata
            metadata_file = os.path.join(output_dir, '_metadata.json')
            with open(metadata_file, 'w') as f:
                json.dump({
                    'company_info': self.company_info,
                    'metadata': self.metadata,
                    'extraction_date': datetime.now().isoformat(),
                    'table_count': len(self.tables)
                }, f, indent=2)
            
            # Save each table
            for table_data in self.tables:
                df = pd.DataFrame(table_data['data']['data'])
                
                # Clean filename
                filename = re.sub(r'[^\w\s-]', '', table_data['name'])
                filename = re.sub(r'[-\s]+', '-', filename)
                filename = f"{filename}.csv"
                filepath = os.path.join(output_dir, filename)
                
                df.to_csv(filepath, index=False, header=False)
                self.logger.info(f"💾 Saved: {filename} ({len(df)} rows)")
            
            self.logger.info(f"✅ All tables saved to directory: {output_dir}")
            return True
            
        except Exception as e:
            self.logger.error(f"✗ Error saving CSVs: {str(e)}")
            return False
    
    def save_to_json(self, output_file=None):
        """Save all data as JSON for further processing"""
        if not self.tables:
            self.logger.error("No tables to save!")
            return False
        
        # Generate filename if not provided
        if not output_file:
            company = self.company_info.get('ticker', 'company')
            date = datetime.now().strftime('%Y%m%d')
            output_file = f"FIRE_{company}_financial_data_{date}.json"
        
        try:
            output_data = {
                'company_info': self.company_info,
                'metadata': self.metadata,
                'extraction_date': datetime.now().isoformat(),
                'tables': []
            }
            
            for table in self.tables:
                table_data = {
                    'name': table['name'],
                    'section': table['section'],
                    'metadata': table['metadata'],
                    'data': table['data']['data'],
                    'formatting': table['data']['formatting']
                }
                output_data['tables'].append(table_data)
            
            with open(output_file, 'w') as f:
                json.dump(output_data, f, indent=2)
            
            self.logger.info(f"✅ Data saved to JSON: {output_file}")
            return True
            
        except Exception as e:
            self.logger.error(f"✗ Error saving JSON: {str(e)}")
            return False
    
    def print_table_summary(self):
        """Print enhanced summary of extracted tables"""
        if not self.tables:
            self.logger.warning("No tables extracted yet!")
            return
        
        print(f"\n📋 Financial Tables Summary for {self.company_info.get('name', 'Company')}")
        print(f"   Ticker: {self.company_info.get('ticker', 'N/A')} | CIK: {self.company_info.get('cik', 'N/A')}")
        print("=" * 80)
        print(f"📊 Total Tables Extracted: {len(self.tables)}")
        print("=" * 80)
        
        # Group tables by section
        sections = {}
        for table in self.tables:
            section = table['section']
            if section not in sections:
                sections[section] = []
            sections[section].append(table)
        
        # Print by section
        for section, tables in sections.items():
            print(f"\n📁 {section.upper()}")
            print("-" * 40)
            
            for i, table in enumerate(tables, 1):
                data = table['data']['data']
                rows = len(data)
                cols = len(data[0]) if data else 0
                has_numbers = table['metadata']['has_numbers']
                
                print(f"  {i}. {table['name']}")
                print(f"     📐 Size: {rows} rows × {cols} columns")
                print(f"     🔢 Contains numbers: {'Yes' if has_numbers else 'No'}")
                
                # Show preview of first row
                if data and len(data) > 0:
                    preview_cells = []
                    for cell in data[0][:3]:
                        cell_text = str(cell)[:20]
                        if cell_text:
                            preview_cells.append(cell_text)
                    
                    if preview_cells:
                        preview = " | ".join(preview_cells)
                        print(f"     👀 Preview: {preview}...")
                print()


def main():
    """Example usage of enhanced FIRE scraper"""
    print("🔥 FIRE - Financial Institution Regulatory Extractor")
    print("=" * 50)
    
    # Example 1: Analyze by ticker symbol
    print("\n📊 Example 1: Analyzing Apple Inc. (AAPL) 10-K")
    scraper = EnhancedFIREScraper()
    
    # Set company by ticker
    if scraper.set_company(ticker='AAPL'):
        print(f"✓ Found company: {scraper.company_info['name']}")
        
        # Get latest 10-K
        filing_url = scraper.get_filing_url('10-K')
        if filing_url:
            print(f"✓ Found filing URL: {filing_url[:60]}...")
            scraper.filing_url = filing_url
            
            # Extract tables
            if scraper.scrape_all_tables():
                scraper.print_table_summary()
                
                # Save in multiple formats
                scraper.save_to_excel_formatted()
                scraper.save_to_excel_basic()
                scraper.save_to_csv()
                scraper.save_to_json()
    
    # Example 2: Analyze local file
    print("\n📊 Example 2: Analyzing local filing")
    local_scraper = EnhancedFIREScraper(
        company_info={'ticker': 'LOCAL', 'name': 'Local Company'},
        local_file_path='sample_10k.htm'
    )
    
    if os.path.exists('sample_10k.htm'):
        if local_scraper.scrape_all_tables():
            local_scraper.print_table_summary()
            local_scraper.save_to_excel_formatted('local_filing_formatted.xlsx')


if __name__ == "__main__":
    main()