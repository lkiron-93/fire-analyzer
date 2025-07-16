"""
FIRE Bulk Data Processor Module
Handles FFIEC Bulk Data Downloads and converts to 6-column Excel format
Enhanced with Treasury Risk and ALM Metrics
"""

# Standard library imports
import gc
import glob
import json
import logging
import multiprocessing as mp
import os
import re
import time
from datetime import datetime
from functools import partial
from pathlib import Path
from typing import Dict, List, Tuple, Optional

# Third-party imports
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Local imports
from bulk_file_manager import BulkFileManager, BulkDataOrganizer


# Configure logging with both file and console output
log_dir = os.path.join(os.path.dirname(__file__), 'logs')
os.makedirs(log_dir, exist_ok=True)

# Create log filename with timestamp
log_filename = os.path.join(log_dir, f'bulk_processing_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename),
        logging.StreamHandler()  # This will also print to console
    ]
)

# Create logger for this module
logger = logging.getLogger('FIRE.BulkData')

class EnhancedMDRMDictionary:
    """
    Enhanced MDRM Dictionary with additional lookup capabilities for Bulk Data processing
    """
    
    def __init__(self, dictionary_path=None, logger=None):
        self.logger = logger or logging.getLogger('FIRE.BulkData')
        self.mdrm_codes = {}
        self.schedule_metadata = {}
        self.institution_lookup = {}
        self.line_item_structure = {}
        
        # Load MDRM dictionary
        if dictionary_path:
            self.load_mdrm_dictionary(dictionary_path)
        
        # Initialize schedule metadata
        self._initialize_schedule_metadata()
        
        # Load institution names
        self.institution_lookup = self._load_institution_lookup()
        
    def _load_institution_lookup(self):
        """Load institution name lookup"""
        lookup_path = os.path.join(
            os.path.dirname(__file__), 
            "dictionaries", 
            "institution_lookup.json"
        )
        
        if os.path.exists(lookup_path):
            try:
                with open(lookup_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                self.logger.error(f"Failed to load institution lookup: {e}")
        
        return {}
            
    def load_mdrm_dictionary(self, dictionary_path):
        """Load MDRM dictionary from JSON file"""
        try:
            with open(dictionary_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
            # Handle both full dictionary and excerpt formats
            if '_sample_codes' in data:
                self.mdrm_codes = data['_sample_codes']
            else:
                self.mdrm_codes = data
                
            self.logger.info(f"✓ Loaded {len(self.mdrm_codes)} MDRM codes")
            
        except Exception as e:
            self.logger.error(f"✗ Error loading MDRM dictionary: {str(e)}")
    
    def _initialize_schedule_metadata(self):
        """Initialize metadata for Call Report schedules"""
        self.schedule_metadata = {
            'RC': {
                'name': 'Consolidated Balance Sheet',
                'prefix': 'RC',
                'description': 'Assets, Liabilities, and Equity Capital'
            },
            'RCA': {
                'name': 'Cash and Balances Due',
                'prefix': 'RC-A',
                'description': 'Cash and Balances Due from Depository Institutions'
            },
            'RCB': {
                'name': 'Securities',
                'prefix': 'RC-B',
                'description': 'Securities Holdings'
            },
            'RCC1': {
                'name': 'Loans and Leases - Part I',
                'prefix': 'RC-C I',
                'description': 'Loans and Leases Held for Sale and Held for Investment'
            },
            'RCC2': {
                'name': 'Loans and Leases - Part II',
                'prefix': 'RC-C II',
                'description': 'Loans to Small Businesses and Small Farms'
            },
            'RCE': {
                'name': 'Deposits',
                'prefix': 'RC-E',
                'description': 'Deposit Liabilities'
            },
            'RCF': {
                'name': 'Other Assets',
                'prefix': 'RC-F',
                'description': 'All Other Assets'
            },
            'RCG': {
                'name': 'Other Liabilities',
                'prefix': 'RC-G',
                'description': 'All Other Liabilities'
            },
            'RCL': {
                'name': 'Derivatives and Off-Balance Sheet',
                'prefix': 'RC-L',
                'description': 'Derivatives and Off-Balance Sheet Items'
            },
            'RCN': {
                'name': 'Past Due and Nonaccrual',
                'prefix': 'RC-N',
                'description': 'Past Due and Nonaccrual Loans, Leases, and Other Assets'
            },
            'RCO': {
                'name': 'Other Data',
                'prefix': 'RC-O',
                'description': 'Other Data for Deposit Insurance Assessments'
            },
            'RCR1': {
                'name': 'Regulatory Capital - Part I',
                'prefix': 'RC-R I',
                'description': 'Regulatory Capital Components and Ratios'
            },
            'RCR2': {
                'name': 'Regulatory Capital - Part II',
                'prefix': 'RC-R II',
                'description': 'Risk-Weighted Assets'
            },
            'RCS': {
                'name': 'Servicing Assets',
                'prefix': 'RC-S',
                'description': 'Servicing, Securitization, and Asset Sale Activities'  
            },
            'RCT': {
                'name': 'Fiduciary and Related Services',
                'prefix': 'RC-T',
                'description': 'Fiduciary and Related Services'
            },
            'RI': {
                'name': 'Income Statement',
                'prefix': 'RI',
                'description': 'Income Statement'
            },
            'RIA': {
                'name': 'Changes in Equity Capital',
                'prefix': 'RI-A',
                'description': 'Changes in Bank Equity Capital'
            },
            'RIB': {
                'name': 'Charge-offs and Recoveries',
                'prefix': 'RI-B',
                'description': 'Charge-offs and Recoveries and Changes in Allowance'
            },
            'RIE': {
                'name': 'Explanations',
                'prefix': 'RI-E',
                'description': 'Explanations'
            }
        }
    
    def get_mdrm_description(self, code):
        """Get description for MDRM code"""
        return self.mdrm_codes.get(code, "")
    
    def get_schedule_info(self, schedule_code):
        """Get schedule metadata"""
        return self.schedule_metadata.get(schedule_code, {})
    
    def add_institution(self, rssd_id, name):
        """Add institution to lookup"""
        self.institution_lookup[str(rssd_id)] = name
    
    def get_institution_name(self, rssd_id):
        """Get institution name by RSSD ID"""
        # First check if we have institution names loaded
        if hasattr(self, 'institution_lookup') and self.institution_lookup:
            return self.institution_lookup.get(str(rssd_id), f"Institution {rssd_id}")
        
        # Fallback to default
        return f"Institution {rssd_id}"


class LineItemMapper:
    """
    Enhanced mapper for MDRM codes to hierarchical line items within schedules
    Provides comprehensive coverage for all major Call Report schedules
    """
    
    def __init__(self, logger=None):
        self.logger = logger or logging.getLogger('FIRE.LineMapper')
        self.mappings = {}
        self._initialize_comprehensive_mappings()
    
    def _initialize_comprehensive_mappings(self):
        """Initialize comprehensive line item mappings for all schedules"""
        
        # Schedule RC - Balance Sheet
        self.mappings['RC'] = {
            # Assets
            'RCON0081': 'RC.1.a',   # Noninterest-bearing balances
            'RCON0071': 'RC.1.b',   # Interest-bearing balances  
            'RCON1287': 'RC.1.c',   # Federal funds sold
            'RCON1350': 'RC.1.d',   # Securities purchased under agreements to resell
            'RCON3545': 'RC.2.a',   # Trading assets
            'RCON1773': 'RC.2.b',   # Available-for-sale securities
            'RCON1754': 'RC.2.c',   # Held-to-maturity securities
            'RCON2122': 'RC.4',     # Loans and leases, net
            'RCON3123': 'RC.5',     # Trading assets
            'RCON2145': 'RC.6',     # Premises and fixed assets
            'RCON2150': 'RC.7',     # Other real estate owned
            'RCON2130': 'RC.8',     # Investments in unconsolidated subsidiaries
            'RCON3656': 'RC.9',     # Customer liability on acceptances
            'RCON2160': 'RC.10',    # Intangible assets
            'RCON2170': 'RC.12',    # Total assets
            
            # Liabilities
            'RCON2200': 'RC.13.a.1', # Noninterest-bearing deposits
            'RCON2520': 'RC.13.a.2', # Interest-bearing deposits
            'RCON2800': 'RC.14',     # Federal funds purchased
            'RCON3190': 'RC.16',     # Trading liabilities
            'RCON2930': 'RC.17',     # Other borrowed money
            'RCON3200': 'RC.19',     # Subordinated notes
            'RCON2948': 'RC.21',     # Total liabilities
            
            # Equity
            'RCON3230': 'RC.23',     # Perpetual preferred stock
            'RCON3838': 'RC.24',     # Common stock
            'RCON3839': 'RC.25',     # Surplus
            'RCON3632': 'RC.26.a',   # Retained earnings
            'RCON2134': 'RC.26.b',   # Accumulated other comprehensive income
            'RCON3210': 'RC.28',     # Total equity capital
            
            # Add RCFD versions (consolidated)
            'RCFD2170': 'RC.12',     # Total assets (consolidated)
            'RCFD3210': 'RC.28',     # Total equity capital (consolidated)
        }
        
        # Schedule RI - Income Statement
        self.mappings['RI'] = {
            # Interest Income
            'RIAD4010': 'RI.1.a.1',  # Interest on loans secured by real estate
            'RIAD4065': 'RI.1.a.2',  # Interest on commercial and industrial loans
            'RIAD4115': 'RI.1.a.3',  # Interest on loans to individuals
            'RIAD4107': 'RI.1.a',    # Total interest income on loans
            'RIAD4060': 'RI.1.b',    # Income from lease financing
            'RIADB985': 'RI.1.c',    # Interest on balances due from depository institutions
            'RIAD4020': 'RI.1.d',    # Interest on securities
            'RIAD4065': 'RI.1.e',    # Interest on trading assets
            'RIAD4115': 'RI.1.f',    # Interest on federal funds sold
            
            # Interest Expense
            'RIAD4170': 'RI.2.a',    # Interest on deposits
            'RIAD4180': 'RI.2.b',    # Interest on federal funds purchased
            'RIAD4185': 'RI.2.c',    # Interest on trading liabilities
            'RIAD4200': 'RI.2.d',    # Interest on subordinated notes
            'RIAD4073': 'RI.2',      # Total interest expense
            
            # Net Interest Income
            'RIAD4074': 'RI.3',      # Net interest income
            'RIAD4230': 'RI.4',      # Provision for loan losses
            'RIAD4079': 'RI.5',      # Noninterest income
            'RIAD4093': 'RI.7',      # Noninterest expense
            'RIAD4340': 'RI.11',     # Net income
        }
        
        # Schedule RC-N - Past Due and Nonaccrual
        self.mappings['RCN'] = {
            'RCON1407': 'RC-N.1.a',   # Real estate loans past due 30-89 days
            'RCON1403': 'RC-N.1.b',   # Real estate loans past due 90+ days
            'RCON5389': 'RC-N.1.c',   # Real estate loans nonaccrual
            'RCON1251': 'RC-N.2.a',   # Commercial loans past due 30-89 days
            'RCON1252': 'RC-N.2.b',   # Commercial loans past due 90+ days
            'RCON1253': 'RC-N.2.c',   # Commercial loans nonaccrual
            'RCON5459': 'RC-N.7',     # Total past due 30-89 days
            'RCON5460': 'RC-N.8',     # Total past due 90+ days
            'RCON5461': 'RC-N.9',     # Total nonaccrual
        }
        
        # Schedule RC-R - Regulatory Capital (Part I)
        self.mappings['RCR'] = self.mappings.get('RCR', {})
        self.mappings['RCR'].update({
            'RCOA8274': 'RC-R.1',     # Common equity tier 1 capital
            'RCOA8260': 'RC-R.2',     # Additional tier 1 capital
            'RCOA5311': 'RC-R.3',     # Tier 1 capital
            'RCOA3792': 'RC-R.4',     # Tier 2 capital
            'RCOA3128': 'RC-R.5',     # Total capital
            'RCOA2170': 'RC-R.6',     # Total assets
            'RCOA3814': 'RC-R.7',     # Total risk-weighted assets
            'RCOAP793': 'RC-R.8',     # Common equity tier 1 capital ratio
            'RCOAP794': 'RC-R.9',     # Tier 1 capital ratio
            'RCOAP795': 'RC-R.10',    # Total capital ratio
            'RCOAP796': 'RC-R.11',    # Leverage ratio
        })
        
        # Schedule RC-L - Derivatives and Off-Balance Sheet
        self.mappings['RCL'] = {
            'RCON3814': 'RC-L.1.a',   # Financial standby letters of credit
            'RCON3815': 'RC-L.1.b',   # Performance standby letters of credit
            'RCON3816': 'RC-L.1.c',   # Commercial letters of credit
            'RCON3411': 'RC-L.2',     # Unused commitments
            'RCON8723': 'RC-L.11.a',  # Interest rate derivatives notional
            'RCON8724': 'RC-L.11.b',  # Foreign exchange derivatives notional
            'RCON8725': 'RC-L.11.c',  # Equity derivatives notional
            'RCON8726': 'RC-L.11.d',  # Commodity derivatives notional
        }
        
        # Schedule RC-T - Fiduciary and Related Services
        self.mappings['RCT'] = {
            'RIADB904': 'RC-T.1',     # Personal trust and agency accounts income
            'RIADB905': 'RC-T.2',     # Employee benefit-defined contribution income
            'RIADB906': 'RC-T.3',     # Employee benefit-defined benefit income
            'RIADB907': 'RC-T.4',     # Other retirement accounts income
            'RIADB909': 'RC-T.5',     # Custody and safekeeping accounts income
            'RIADA491': 'RC-T.10',    # Net fiduciary income
            'RIADC058': 'RC-T.11',    # Fiduciary expenses
        }
        
        # Schedule RC-O - Other Data
        self.mappings['RCO'] = {
            'RCON6810': 'RC-O.1',     # Total deposit liabilities before exclusions
            'RCONF049': 'RC-O.2',     # Total allowable exclusions
            'RCONF045': 'RC-O.3',     # Total daily average deposits
            'RCONF048': 'RC-O.4',     # Quarter-end total deposits
            'RCONG641': 'RC-O.M.1',   # Number of deposit accounts $250,000 or less
            'RCONG645': 'RC-O.M.2',   # Number of deposit accounts more than $250,000
        }
        
        # Log mapping statistics
        total_mappings = sum(len(schedule) for schedule in self.mappings.values())
        self.logger.info(f"✓ Initialized {total_mappings} line item mappings across {len(self.mappings)} schedules")
    
    def get_line_item(self, schedule_code, mdrm_code):
        """Get hierarchical line item for MDRM code within schedule"""
        # Handle schedule code variations
        base_schedule = self._normalize_schedule_code(schedule_code)
        
        schedule_mappings = self.mappings.get(base_schedule, {})
        
        # First check exact match
        if mdrm_code in schedule_mappings:
            return schedule_mappings[mdrm_code]
        
        # Check without prefix (RCON/RCFD are often interchangeable)
        code_number = mdrm_code[4:] if len(mdrm_code) > 4 else mdrm_code
        
        for code, line_item in schedule_mappings.items():
            if code.endswith(code_number):
                return line_item
        
        # Return empty string if no mapping found
        return ""
    
    def _normalize_schedule_code(self, schedule_code):
        """Normalize schedule code to base form"""
        # Remove part numbers and extra characters
        if not schedule_code:
            return ""
            
        # Convert to uppercase
        code = schedule_code.upper()
        
        # Handle special cases
        if code.startswith('RCR'):
            return 'RCR'
        elif code.startswith('RCL'):
            return 'RCL'
        elif code.startswith('RCO'):
            return 'RCO'
        elif code.startswith('RCT'):
            return 'RCT'
        elif code.startswith('RCN'):
            return 'RCN'
        
        # Remove numbers at the end (e.g., RCT2 -> RCT)
        return code.rstrip('0123456789')
    
    def parse_schedule_code(self, filename):
        """Extract schedule code from filename"""
        # Enhanced pattern matching
        patterns = [
            r'Schedule\s+([A-Z]+\d*)',
            r'Call\s+Schedule\s+([A-Z]+\d*)',
            r'FFIEC\s+CDR\s+Call\s+Schedule\s+([A-Z]+\d*)',
            r'Schedule([A-Z]+\d*)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, filename, re.IGNORECASE)
            if match:
                return match.group(1).upper()
        
        # Extended schedule codes list
        schedule_codes = [
            'GL', 'RC', 'RCA', 'RCB', 'RCC', 'RCD', 'RCE', 'RCF', 'RCG', 
            'RCH', 'RCK', 'RCL', 'RCM', 'RCN', 'RCO', 'RCP', 'RCQ', 'RCR', 
            'RCS', 'RCT', 'RCV', 'RI', 'RIA', 'RIB', 'RIC', 'RIE'
        ]
        
        # Check for schedule codes anywhere in filename
        filename_upper = filename.upper()
        for code in schedule_codes:
            if code in filename_upper:
                # Make sure it's not part of a larger word
                idx = filename_upper.find(code)
                if idx > 0 and filename_upper[idx-1].isalnum():
                    continue
                if idx + len(code) < len(filename_upper) and filename_upper[idx + len(code)].isalpha():
                    continue
                return code
        
        return None
    
class BulkDataProcessor:
    """
    Main processor for FFIEC Bulk Data files
    """
    
    def __init__(self, dictionary_path=None, logger=None):
        self.logger = logger or logging.getLogger('FIRE.BulkData')
        self.dictionary = EnhancedMDRMDictionary(dictionary_path, logger)
        self.line_mapper = LineItemMapper(logger)
        self.data_frames = {}
        self.processed_data = []
        
        # Load institution names
        self.institution_names = self._load_institution_names()
        
        # Share institution names with the dictionary
        if hasattr(self.dictionary, 'institution_lookup'):
            self.dictionary.institution_lookup.update(self.institution_names)
        
    def _load_institution_names(self):
        """Load institution name mappings"""
        institution_names = {}
        
        # Try to load from lookup file
        lookup_path = os.path.join(
            os.path.dirname(__file__), 
            "dictionaries", 
            "institution_lookup.json"
        )
        
        if os.path.exists(lookup_path):
            try:
                with open(lookup_path, 'r', encoding='utf-8') as f:
                    institution_names = json.load(f)
                    self.logger.info(f"✓ Loaded {len(institution_names)} institution names from lookup")
            except Exception as e:
                self.logger.warning(f"Could not load institution names: {e}")
        else:
            self.logger.warning(f"Institution lookup file not found at: {lookup_path}")
        
        return institution_names
      
    def process_bulk_file(self, filepath, target_rssd_id=None):
        """
        Process a single bulk data file
        
        Args:
            filepath: Path to the bulk data file
            target_rssd_id: Optional RSSD ID to filter for specific institution
            
        Returns:
            pd.DataFrame: Processed data in 6-column format
        """
        # Check file size to determine which processor to use
        file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
        
        # Use enhanced processing for files > 5MB or if validation is important
        if file_size_mb > 5:
            self.logger.info(f"📄 File size {file_size_mb:.1f}MB - using enhanced processor")
            return self.process_bulk_file_enhanced(filepath, target_rssd_id)
        
        # Otherwise, use original processing for small files
        self.logger.info(f"📄 Processing file: {os.path.basename(filepath)}")
        
        try:
            # Detect schedule from filename
            schedule_code = self.line_mapper.parse_schedule_code(os.path.basename(filepath))
            if not schedule_code:
                self.logger.warning(f"⚠️ Could not detect schedule code from: {os.path.basename(filepath)}")
                schedule_code = "Unknown"
            
            schedule_info = self.dictionary.get_schedule_info(schedule_code)
            schedule_name = schedule_info.get('name', schedule_code)
            
            self.logger.info(f"📊 Detected Schedule: {schedule_code} - {schedule_name}")
            
            # Read the file
            df = pd.read_csv(filepath, sep='\t', dtype=str, low_memory=False)
            self.logger.info(f"✓ Loaded {len(df)} rows, {len(df.columns)} columns")
            
            # Filter by RSSD ID if specified
            if target_rssd_id:
                df = df[df['IDRSSD'] == str(target_rssd_id)]
                self.logger.info(f"✓ Filtered to {len(df)} rows for RSSD ID: {target_rssd_id}")
            
            if df.empty:
                self.logger.warning("⚠️ No data found after filtering")
                return pd.DataFrame()
            
            # Convert to 6-column format
            result_data = self._convert_to_six_column_format(df, schedule_code)
            
            return pd.DataFrame(result_data)
            
        except Exception as e:
            self.logger.error(f"✗ Error processing file: {str(e)}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()
    
    def _convert_to_six_column_format(self, df, schedule_code):
        """
        Convert bulk data to 6-column format:
        RSSDID | Name | Line Item | Description | MDRM Code | Amount
        """
        result_rows = []
        
        # Get schedule info
        schedule_info = self.dictionary.get_schedule_info(schedule_code)
        
        # Process each row
        for idx, row in df.iterrows():
            rssd_id = row.get('IDRSSD', '')
            
            # Get institution name (placeholder - would be loaded from separate file)
            inst_name = self.dictionary.get_institution_name(rssd_id)
            
            # Process each column (except IDRSSD)
            for col in df.columns:
                if col == 'IDRSSD':
                    continue
                
                # Check if this is an MDRM code column
                if re.match(r'^(RCON|RCFD|RIAD|RCFN|RCOA|RCOB|RCOC|RCOD)[A-Z0-9]+$', col):
                    value = row[col]
                    
                    # Skip empty values
                    if pd.isna(value) or str(value).strip() == '':
                        continue
                    
                    # Get line item number
                    line_item = self.line_mapper.get_line_item(schedule_code, col)
                    
                    # Get description from MDRM dictionary
                    description = self.dictionary.get_mdrm_description(col)
                    
                    # Create row in 6-column format
                    result_rows.append({
                        'RSSDID': rssd_id,
                        'Name': inst_name,
                        'Line Item': line_item,
                        'Description': description,
                        'MDRM Code': col,
                        'Amount': value
                    })
      
        self.logger.info(f"✓ Converted to {len(result_rows)} rows in 6-column format")
        return result_rows
    
    def _convert_to_six_column_format_enhanced(self, df, schedule_code):
        """
        Enhanced conversion to 6-column format with validation and error handling
        
        Improvements:
        - Better data type handling
        - Amount validation and formatting
        - Missing data handling
        - Progress tracking for large files
        """
        result_rows = []
        
        # Get schedule info
        schedule_info = self.dictionary.get_schedule_info(schedule_code)
        
        # Pre-compile regex for MDRM codes
        mdrm_pattern = re.compile(r'^(RCON|RCFD|RIAD|RCFN|RCOA|RCOB|RCOC|RCOD)[A-Z0-9]+$')
        
        # Track processing stats
        total_rows = len(df)
        processed_rows = 0
        skipped_cells = 0
        invalid_amounts = 0
        
        # Process in chunks for large files
        chunk_size = 1000
        
        for chunk_start in range(0, total_rows, chunk_size):
            chunk_end = min(chunk_start + chunk_size, total_rows)
            chunk = df.iloc[chunk_start:chunk_end]
            
            # Log progress for large files
            if total_rows > 10000 and chunk_start % 10000 == 0:
                self.logger.info(f"Processing rows {chunk_start:,} - {chunk_end:,} of {total_rows:,}")
            
            for idx, row in chunk.iterrows():
                rssd_id = str(row.get('IDRSSD', '')).strip()
                
                # Validate RSSD ID
                if not rssd_id or rssd_id == 'nan':
                    continue
                
                # Get institution name
                inst_name = self.dictionary.get_institution_name(rssd_id)
                
                # Process each column
                for col in df.columns:
                    if col == 'IDRSSD':
                        continue
                    
                    # Check if this is an MDRM code column
                    if mdrm_pattern.match(col):
                        value = row[col]
                        
                        # Enhanced value validation
                        if pd.isna(value) or str(value).strip() in ['', '.', 'NA', 'N/A']:
                            skipped_cells += 1
                            continue
                        
                        # Clean and validate amount
                        amount_str = str(value).strip()
                        
                        # Handle special values
                        if amount_str.upper() in ['ND', 'NR', 'CONF']:
                            # ND = No Data, NR = Not Reported, CONF = Confidential
                            continue
                        
                        # Validate numeric format
                        try:
                            # Remove any non-numeric characters except -, ., and ,
                            cleaned_amount = re.sub(r'[^\d\-.,]', '', amount_str)
                            
                            # Handle parentheses for negatives
                            if amount_str.startswith('(') and amount_str.endswith(')'):
                                cleaned_amount = '-' + cleaned_amount
                            
                            # Validate it's a number
                            float(cleaned_amount.replace(',', ''))
                        except ValueError:
                            invalid_amounts += 1
                            self.logger.debug(f"Invalid amount format: {amount_str} for {col}")
                            continue
                        
                        # Get line item number
                        line_item = self.line_mapper.get_line_item(schedule_code, col)
                        
                        # Get description from MDRM dictionary
                        description = self.dictionary.get_mdrm_description(col)
                        
                        # If no description, try to generate one
                        if not description:
                            description = self._generate_description(col, schedule_code)
                        
                        # Create row in 6-column format
                        result_rows.append({
                            'RSSDID': rssd_id,
                            'Name': inst_name,
                            'Line Item': line_item,
                            'Description': description,
                            'MDRM Code': col,
                            'Amount': cleaned_amount
                        })
                
                processed_rows += 1
        
        # Log processing statistics
        self.logger.info(f"✓ Processed {processed_rows:,} rows, skipped {skipped_cells:,} empty cells")
        if invalid_amounts > 0:
            self.logger.warning(f"⚠️ Found {invalid_amounts:,} invalid amount values")
        
        self.logger.info(f"✓ Converted to {len(result_rows):,} rows in 6-column format")
        
        return result_rows

    def _generate_description(self, mdrm_code, schedule_code):
        """Generate a fallback description for unmapped MDRM codes"""
        # Extract components
        prefix = mdrm_code[:4] if len(mdrm_code) > 4 else mdrm_code
        
        # Common prefixes
        prefix_map = {
            'RCON': 'Domestic',
            'RCFD': 'Consolidated',
            'RIAD': 'Income Statement',
            'RCFN': 'Foreign',
            'RCOA': 'Regulatory',
            'RCOB': 'Other',
            'RCOC': 'Credit',
            'RCOD': 'Derivative'
        }
        
        prefix_desc = prefix_map.get(prefix, prefix)
        
        # Add schedule context
        schedule_desc = self.dictionary.get_schedule_info(schedule_code).get('name', schedule_code)
        
        return f"{prefix_desc} item - {schedule_desc}"
    
    def validate_data_quality(self, df):
        """Validate data quality and provide warnings"""
        issues = []
        
        # Check for required columns
        if 'IDRSSD' not in df.columns:
            issues.append("Missing required IDRSSD column")
        
        # Check for empty dataframe
        if df.empty:
            issues.append("No data rows found")
            return issues
        
        # Check for institutions
        unique_institutions = df['IDRSSD'].nunique()
        if unique_institutions == 0:
            issues.append("No valid institution IDs found")
        
        # Check for MDRM codes
        mdrm_columns = [col for col in df.columns if re.match(r'^(RCON|RCFD|RIAD|RCFN)', col)]
        if not mdrm_columns:
            issues.append("No MDRM code columns found")
        
        # Check data density
        if len(mdrm_columns) > 0:
            total_cells = len(df) * len(mdrm_columns)
            non_empty_cells = 0
            
            for col in mdrm_columns:
                non_empty_cells += df[col].notna().sum()
            
            density = (non_empty_cells / total_cells) * 100 if total_cells > 0 else 0
            
            if density < 1:
                issues.append(f"Very low data density: {density:.1f}%")
            elif density < 10:
                issues.append(f"Low data density: {density:.1f}%")
        
        return issues
    
    def process_bulk_file_enhanced(self, filepath, target_rssd_id=None):
        """
        Enhanced bulk file processing with better error handling and validation
        """
        self.logger.info(f"📄 Processing file: {os.path.basename(filepath)}")
        file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
        self.logger.info(f"📊 File size: {file_size_mb:.1f} MB")
        
        try:
            # Detect schedule from filename
            schedule_code = self.line_mapper.parse_schedule_code(os.path.basename(filepath))
            if not schedule_code:
                self.logger.warning(f"⚠️ Could not detect schedule code from: {os.path.basename(filepath)}")
                schedule_code = "Unknown"
            
            schedule_info = self.dictionary.get_schedule_info(schedule_code)
            schedule_name = schedule_info.get('name', schedule_code)
            
            self.logger.info(f"📊 Detected Schedule: {schedule_code} - {schedule_name}")
            
            # Read the file with optimizations for large files
            if file_size_mb > 10:
                self.logger.info(f"⏳ Reading large file (this may take a moment)...")
            
            # Use chunking for very large files
            if file_size_mb > 50:
                chunks = []
                chunk_iter = pd.read_csv(filepath, sep='\t', dtype=str, low_memory=False, chunksize=10000)
                
                for i, chunk in enumerate(chunk_iter):
                    chunks.append(chunk)
                    if i % 10 == 0:
                        self.logger.info(f"  Read {(i+1)*10000:,} rows...")
                
                df = pd.concat(chunks, ignore_index=True)
            else:
                # Read normally for smaller files
                df = pd.read_csv(filepath, sep='\t', dtype=str, low_memory=False)
            
            self.logger.info(f"✓ Loaded {len(df):,} rows, {len(df.columns)} columns")
            
            # Validate data quality
            issues = self.validate_data_quality(df)
            if issues:
                self.logger.warning("⚠️ Data quality issues found:")
                for issue in issues:
                    self.logger.warning(f"  - {issue}")
            
            # Filter by RSSD ID if specified
            if target_rssd_id:
                df_filtered = df[df['IDRSSD'] == str(target_rssd_id)]
                self.logger.info(f"✓ Filtered to {len(df_filtered):,} rows for RSSD ID: {target_rssd_id}")
                
                if df_filtered.empty:
                    self.logger.warning("⚠️ No data found for specified RSSD ID")
                    # Show available RSSD IDs
                    available_ids = df['IDRSSD'].value_counts().head(10)
                    self.logger.info("Available RSSD IDs in file (top 10):")
                    for rssd_id, count in available_ids.items():
                        self.logger.info(f"  {rssd_id}: {count:,} rows")
                    return pd.DataFrame()
                
                df = df_filtered
            
            # Convert to 6-column format with enhanced processing
            result_data = self._convert_to_six_column_format_enhanced(df, schedule_code)
            
            # Create DataFrame and add metadata
            result_df = pd.DataFrame(result_data)
            
            # Add metadata attributes
            result_df.attrs['schedule_code'] = schedule_code
            result_df.attrs['schedule_name'] = schedule_name
            result_df.attrs['source_file'] = os.path.basename(filepath)
            result_df.attrs['processing_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            return result_df
            
        except Exception as e:
            self.logger.error(f"✗ Error processing file: {str(e)}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return pd.DataFrame()
        
    def process_files_parallel(self, file_list, target_rssd_id=None, num_workers=None, progress_callback=None):
        """
        Process multiple files in parallel with schedule code tracking
        
        Args:
            file_list: List of file paths to process
            target_rssd_id: Optional RSSD ID to filter
            num_workers: Number of parallel workers (default: CPU count - 1)
            progress_callback: Optional callback for progress updates
            
        Returns:
            pd.DataFrame: Combined results with schedule information
        """
        if num_workers is None:
            num_workers = min(mp.cpu_count() - 1, 4)
        
        self.logger.info(f"🚀 Starting parallel processing with {num_workers} workers")
        
        # Create a wrapper function that includes schedule code extraction
        def process_file_with_schedule(filepath):
            """Process file and add schedule code to results"""
            try:
                # Extract schedule code
                schedule_code = self.line_mapper.parse_schedule_code(os.path.basename(filepath))
                if not schedule_code:
                    schedule_code = "Unknown"
                
                # Process the file
                file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
                
                if file_size_mb > 5:
                    df = self.process_bulk_file_enhanced(filepath, target_rssd_id)
                else:
                    df = self.process_bulk_file(filepath, target_rssd_id)
                
                # Add schedule code to all rows
                if not df.empty:
                    df['schedule_code'] = schedule_code
                    df['source_file'] = os.path.basename(filepath)
                
                return df
                
            except Exception as e:
                self.logger.error(f"Error in parallel processing of {filepath}: {str(e)}")
                return pd.DataFrame()
        
        # Process files in parallel
        with mp.Pool(num_workers) as pool:
            # Create progress tracking
            results = []
            total_files = len(file_list)
            
            # Use imap for better progress tracking
            for i, result in enumerate(pool.imap(process_file_with_schedule, file_list)):
                results.append(result)
                
                # Send progress update if callback provided
                if progress_callback and (i + 1) % 5 == 0 or (i + 1) == total_files:
                    # Extract schedule code from the filename for progress display
                    current_file = file_list[i] if i < len(file_list) else file_list[-1]
                    schedule_code = self.line_mapper.parse_schedule_code(os.path.basename(current_file))
                    
                    progress_callback({
                        'current_file': i + 1,
                        'total_files': total_files,
                        'current_schedule': schedule_code or "Unknown",
                        'percentage': ((i + 1) / total_files) * 100,
                        'message': f"Processed {i + 1} of {total_files} files"
                    })
                
                # Progress update every 10 files
                if (i + 1) % 10 == 0 or (i + 1) == total_files:
                    self.logger.info(f"  Progress: {i + 1}/{total_files} files processed")
        
        # Combine results
        non_empty_results = [r for r in results if not r.empty]
        
        if non_empty_results:
            combined_df = pd.concat(non_empty_results, ignore_index=True)
            
            # Optimize memory usage
            combined_df = self.optimize_dataframe_memory(combined_df)
            
            self.logger.info(f"✓ Combined {len(non_empty_results)} files into {len(combined_df):,} rows")
            return combined_df
        else:
            self.logger.warning("⚠️ No data found in parallel processing")
            return pd.DataFrame()
      
      
    def optimize_dataframe_memory(self, df):
        """Optimize DataFrame memory usage"""
        start_mem = df.memory_usage(deep=True).sum() / 1024**2
        
        # Optimize object columns
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].astype('category')
        
        # Optimize numeric columns
        for col in df.select_dtypes(include=['int']).columns:
            df[col] = pd.to_numeric(df[col], downcast='integer')
        
        for col in df.select_dtypes(include=['float']).columns:
            df[col] = pd.to_numeric(df[col], downcast='float')
        
        end_mem = df.memory_usage(deep=True).sum() / 1024**2
        self.logger.info(f"💾 Memory optimized: {start_mem:.1f}MB → {end_mem:.1f}MB ({(1-end_mem/start_mem)*100:.1f}% reduction)")
        
        return df

    def process_large_file_chunked(self, filepath, target_rssd_id=None, chunk_size=50000):
        """Process very large files in chunks"""
        self.logger.info(f"📦 Processing large file in chunks of {chunk_size:,} rows")
        
        chunks = []
        for chunk in pd.read_csv(filepath, sep='\t', chunksize=chunk_size, dtype=str, low_memory=False):
            if target_rssd_id:
                chunk = chunk[chunk['IDRSSD'] == str(target_rssd_id)]
            
            if not chunk.empty:
                # Process chunk
                schedule_code = self.line_mapper.parse_schedule_code(os.path.basename(filepath))
                processed = self._convert_to_six_column_format_enhanced(chunk, schedule_code)
                chunks.append(pd.DataFrame(processed))
            
            # Free memory
            gc.collect()
        
        # Combine all chunks
        if chunks:
            result = pd.concat(chunks, ignore_index=True)
            return self.optimize_dataframe_memory(result)
        
        return pd.DataFrame()
    
    def process_directory(self, directory_path, target_rssd_id=None, file_pattern="*.txt", use_parallel=True, progress_callback=None):
        """
        Process all bulk data files in a directory with parallel processing support
        
        Args:
            directory_path: Path to directory containing bulk files
            target_rssd_id: Optional RSSD ID to filter
            file_pattern: Pattern to match files (default: *.txt)
            use_parallel: Whether to use parallel processing (default: True)
            progress_callback: Optional callback for progress updates
            
        Returns:
            Dict[str, pd.DataFrame]: Dictionary of schedule_code -> DataFrame
        """
        self.logger.info(f"📁 Processing directory: {directory_path}")
        
        results = {}
        
        # Find all matching files
        pattern = os.path.join(directory_path, file_pattern)
        files = glob.glob(pattern)
        
        self.logger.info(f"✓ Found {len(files)} files to process")
        
        # Determine if we should use parallel processing
        # Use parallel for > 5 files or if total size > 50MB
        total_size_mb = sum(os.path.getsize(f) / (1024 * 1024) for f in files)
        should_use_parallel = use_parallel and (len(files) > 5 or total_size_mb > 50)
        
        if should_use_parallel:
            self.logger.info(f"🚀 Using parallel processing ({mp.cpu_count()} cores available)")
            self.logger.info(f"📊 Total data size: {total_size_mb:.1f} MB")
            
            # Process files in parallel with progress tracking
            processed_files = self.process_files_parallel(files, target_rssd_id, progress_callback=progress_callback)
            
            # Group by schedule code
            schedule_groups = {}
            for _, row in processed_files.iterrows():
                # Extract schedule code from the row data
                # Assuming the data has been tagged with schedule info during processing
                schedule_code = row.get('schedule_code', 'Unknown')
                
                if schedule_code not in schedule_groups:
                    schedule_groups[schedule_code] = []
                schedule_groups[schedule_code].append(row)
            
            # Convert groups to DataFrames
            for schedule_code, rows in schedule_groups.items():
                if rows:
                    results[schedule_code] = pd.DataFrame(rows)
                    self.logger.info(f"  ✓ Schedule {schedule_code}: {len(rows)} rows")
                    
        else:
            # Use sequential processing for smaller datasets
            self.logger.info(f"📄 Using sequential processing")
            
            for idx, filepath in enumerate(sorted(files)):
                try:
                    # Extract schedule code
                    schedule_code = self.line_mapper.parse_schedule_code(os.path.basename(filepath))
                    if not schedule_code:
                        schedule_code = f"File_{len(results)}"
                    
                    # Send progress update if callback provided
                    if progress_callback:
                        progress_callback({
                            'current_file': idx + 1,
                            'total_files': len(files),
                            'current_schedule': schedule_code,
                            'percentage': ((idx + 1) / len(files)) * 100,
                            'message': f"Processing Schedule {schedule_code}"
                        })
                    
                    # Process file
                    self.logger.info(f"  Processing: {os.path.basename(filepath)}")
                    
                    # Check file size to determine processing method
                    file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
                    
                    if file_size_mb > 100:
                        # Use chunked processing for very large files
                        self.logger.info(f"    Using chunked processing (file size: {file_size_mb:.1f} MB)")
                        df = self.process_large_file_chunked(filepath, target_rssd_id)
                    else:
                        # Use standard processing
                        df = self.process_bulk_file(filepath, target_rssd_id)
                    
                    if not df.empty:
                        # Optimize memory if DataFrame is large
                        if len(df) > 10000:
                            self.logger.info(f"    Optimizing memory usage...")
                            df = self.optimize_dataframe_memory(df)
                        
                        results[schedule_code] = df
                        self.logger.info(f"    ✓ Added {len(df)} rows")
                        
                except Exception as e:
                    self.logger.error(f"  ✗ Error processing {filepath}: {str(e)}")
                    continue
        
        # Summary statistics
        total_rows = sum(len(df) for df in results.values())
        total_memory_mb = sum(df.memory_usage(deep=True).sum() / 1024**2 for df in results.values())
        
        self.logger.info(f"✅ Processed {len(results)} schedules")
        self.logger.info(f"📊 Total rows: {total_rows:,}")
        self.logger.info(f"💾 Total memory usage: {total_memory_mb:.1f} MB")
        
        return results
   
    def save_to_excel(self, data_dict, output_file, institution_name=None, include_enhancements=True, rssd_id=None):
        """
        Save processed data to Excel with formatting and optional enhancement sheets
        
        Args:
            data_dict: Dictionary of schedule_code -> DataFrame
            output_file: Output Excel filename
            institution_name: Optional institution name for metadata
            include_enhancements: Whether to include Executive Dashboard and Key Metrics sheets
            rssd_id: RSSD ID for single institution (needed for enhancements)
        """
        self.logger.info(f"💾 Saving to Excel: {output_file}")
        
        try:
            # Create output directory if it doesn't exist
            output_dir = os.path.dirname(output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                self.logger.info(f"Created output directory: {output_dir}")
            
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Add metadata sheet with enhancement parameters
            ws_meta = wb.create_sheet("Report Info")
            self._add_metadata_sheet(ws_meta, data_dict, institution_name, 
                                include_enhancements=include_enhancements,
                                rssd_id=rssd_id)
            
            # Add enhancement sheets for single institution if requested
            if include_enhancements and data_dict:
                try:
                    self.logger.info("📊 Creating enhanced analysis sheets for single institution...")
                    
                    # Create a format matching multi-institution structure
                    single_inst_results = {
                        rssd_id or 'SINGLE': {
                            'name': institution_name or 'Institution',
                            'data': data_dict,
                            'rssd_id': rssd_id or 'SINGLE'
                        }
                    }
                    
                    # Create Excel Enhancement Processor
                    from bulk_data_processor import ExcelEnhancementProcessor
                    enhancer = ExcelEnhancementProcessor(self.logger)
                    
                    # Executive Dashboard
                    self.logger.info("📊 Creating Executive Dashboard...")
                    ws_dashboard = wb.create_sheet("Executive Dashboard")
                    enhancer.create_executive_dashboard(ws_dashboard, single_inst_results, self)
                    
                    # Key Metrics
                    self.logger.info("📈 Creating Key Metrics sheet...")
                    ws_metrics = wb.create_sheet("Key Metrics")
                    enhancer.create_key_metrics_sheet(ws_metrics, single_inst_results, self)
                    
                    self.logger.info("✅ Enhanced analysis sheets created successfully")
                    
                except Exception as e:
                    self.logger.warning(f"⚠️ Could not create enhancement sheets: {str(e)}")
                    # Remove any partially created sheets
                    if "Executive Dashboard" in wb.sheetnames:
                        wb.remove(wb["Executive Dashboard"])
                    if "Key Metrics" in wb.sheetnames:
                        wb.remove(wb["Key Metrics"])
            
            # Add data sheets
            for schedule_code, df in data_dict.items():
                # Get schedule info
                schedule_info = self.dictionary.get_schedule_info(schedule_code)
                sheet_name = schedule_code[:31]  # Excel limit
                
                self.logger.info(f"Adding sheet: {sheet_name} with {len(df)} rows")
                
                ws = wb.create_sheet(sheet_name)
                self._write_schedule_to_sheet(ws, df, schedule_code, schedule_info)
            
            # Save workbook
            wb.save(output_file)
            self.logger.info(f"✅ Excel file saved: {output_file}")
            
        except PermissionError as e:
            self.logger.error(f"Permission denied saving file: {output_file}")
            self.logger.error(f"Make sure the file is not open in Excel")
            raise Exception(f"Cannot save file - it may be open in Excel: {output_file}")
        except Exception as e:
            self.logger.error(f"Error saving to Excel: {str(e)}")
            self.logger.error(f"Data dict keys: {list(data_dict.keys())}")
            self.logger.error(f"Output file: {output_file}")
            raise
    
    def _add_metadata_sheet(self, ws, data_dict, institution_name, include_enhancements=False, rssd_id=None):
        """Add metadata sheet to workbook with optional quick links to enhancement sheets"""
        # Title
        ws['A1'] = "🔥 FIRE - Bulk Data Analysis Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        # Metadata
        metadata = [
            ['Report Type', 'FFIEC Call Report - Bulk Data'],
            ['Institution', institution_name or 'Multiple Institutions'],
            ['Processing Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Schedules Processed', len(data_dict)],
            ['Total Rows', sum(len(df) for df in data_dict.values())],
        ]
        
        if rssd_id:
            metadata.insert(2, ['RSSD ID', rssd_id])
        
        row = 3
        for label, value in metadata:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Quick Links section if enhancements are included
        if include_enhancements:
            row += 1
            ws.cell(row=row, column=1, value="Quick Links:").font = Font(bold=True, size=12)
            row += 1
            
            # Executive Dashboard link
            cell = ws.cell(row=row, column=1, value="📊 Executive Dashboard")
            cell.hyperlink = "#'Executive Dashboard'!A1"
            cell.font = Font(color="0563C1", underline="single", bold=True)
            ws.cell(row=row, column=2, value="Key performance indicators and visual summary")
            row += 1
            
            # Key Metrics link
            cell = ws.cell(row=row, column=1, value="📈 Key Metrics & Analysis")
            cell.hyperlink = "#'Key Metrics'!A1"
            cell.font = Font(color="0563C1", underline="single", bold=True)
            ws.cell(row=row, column=2, value="Detailed financial ratios and ALM metrics")
            row += 2
        
        # Schedule summary with hyperlinks
        ws.cell(row=row+1, column=1, value="Schedules:").font = Font(bold=True)
        ws.cell(row=row+1, column=2, value="(Click schedule code to jump to tab)").font = Font(italic=True, color="666666")
        row += 2

        for schedule_code, df in data_dict.items():
            schedule_info = self.dictionary.get_schedule_info(schedule_code)
            
            # Create hyperlink to the schedule tab
            sheet_name = schedule_code[:31]  # Excel sheet name limit
            cell = ws.cell(row=row, column=1, value=schedule_code)
            cell.hyperlink = f"#{sheet_name}!A1"
            cell.font = Font(color="0563C1", underline="single")  # Blue, underlined
            
            ws.cell(row=row, column=2, value=schedule_info.get('name', ''))
            ws.cell(row=row, column=3, value=f"{len(df)} rows")
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
    
    def _write_schedule_to_sheet(self, ws, df, schedule_code, schedule_info):
        """Write schedule data to worksheet with formatting"""
        # Add headers
        headers = ['RSSDID', 'Name', 'Line Item', 'Description', 'MDRM Code', 'Amount']
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add "Back to Report Info" link in cell H1
        back_cell = ws.cell(row=1, column=8, value="← Back to Report Info")
        back_cell.hyperlink = "#'Report Info'!A1"
        back_cell.font = Font(color="0563C1", underline="single", bold=True)
        back_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Write data
        for row_idx, row in df.iterrows():
            for col_idx, header in enumerate(headers, 1):
                value = row.get(header, '')
                cell = ws.cell(row=row_idx+2, column=col_idx, value=value)
                
                # Format numbers in Amount column
                if header == 'Amount' and value:
                    try:
                        # Convert to number and format
                        num_value = float(str(value).replace(',', ''))
                        cell.value = num_value
                        cell.number_format = '#,##0'
                    except:
                        pass
                
                # Center align certain columns
                if header in ['RSSDID', 'Line Item', 'MDRM Code']:
                    cell.alignment = Alignment(horizontal='center')
        
        # Apply alternating row colors for better readability
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_num in range(3, len(df) + 2):  # Start from row 3 (first data row after header)
            if row_num % 2 == 0:  # Even rows get light gray background
                for col_num in range(1, 7):  # Columns A through F
                    ws.cell(row=row_num, column=col_num).fill = light_fill
        
        # Auto-fit column widths based on content
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            # Set minimum and maximum widths for each column
            col_letter = column_cells[0].column_letter
            
            if col_letter == 'A':  # RSSDID
                adjusted_width = min(max(length, 10), 15)
            elif col_letter == 'B':  # Name
                adjusted_width = min(max(length, 30), 50)
            elif col_letter == 'C':  # Line Item
                adjusted_width = min(max(length, 10), 20)
            elif col_letter == 'D':  # Description
                adjusted_width = min(max(length, 40), 70)
            elif col_letter == 'E':  # MDRM Code
                adjusted_width = min(max(length, 12), 20)
            elif col_letter == 'F':  # Amount
                adjusted_width = min(max(length, 15), 25)
            else:
                adjusted_width = min(max(length, 10), 30)
            
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=6):
            for cell in row:
                cell.border = thin_border
        
        # Add summary row if there are numeric values in Amount column
        if len(df) > 0:
            summary_row = len(df) + 3
            ws.cell(row=summary_row, column=4, value="Total:").font = Font(bold=True)
            ws.cell(row=summary_row, column=5, value=f"=COUNTA(F2:F{len(df)+1})").font = Font(bold=True)
            ws.cell(row=summary_row, column=5).alignment = Alignment(horizontal='center')
            
            # Add border to summary row
            for col in range(4, 7):
                ws.cell(row=summary_row, column=col).border = Border(
                    top=Side(style='double'),
                    bottom=Side(style='thin')
                )
        
        # Freeze top row and first two columns (RSSDID and Name)
        ws.freeze_panes = 'C2'
        
        # Add auto-filter to all data columns
        ws.auto_filter.ref = f"A1:F{len(df) + 1}"
        
        self.logger.info(f"✓ Written {len(df)} rows to sheet: {schedule_code}")

# Enhanced Excel processor with ALM metrics
class ExcelEnhancementProcessor:
    """Handles creation of Executive Dashboard and Key Metrics sheets with ALM metrics"""
    
    def __init__(self, logger=None):
        self.logger = logger or logging.getLogger('FIRE.ExcelEnhancements')
        
        # Define key MDRM codes for metric calculations
        self.metric_codes = {
            # Balance Sheet (RC) - Existing codes
            'total_assets': ['RCFD2170', 'RCON2170'],
            'total_loans': ['RCFD2122', 'RCON2122'], 
            'total_deposits': ['RCON2200', 'RCFN2200', 'RCON6631', 'RCFN6631'],
            'total_equity': ['RCFD3210', 'RCON3210'],
            
            # Income Statement (RI) - Existing codes
            'net_income': ['RIAD4340'],
            'net_interest_income': ['RIAD4074'],
            'noninterest_income': ['RIAD4079'],
            'noninterest_expense': ['RIAD4093'],
            'provision_loan_losses': ['RIAD4230'],
            
            # Asset Quality (RC-N) - Existing codes
            'nonaccrual_loans': ['RCFD1403', 'RCON1403'],
            'past_due_90': ['RCFD1407', 'RCON1407'],
            
            # Capital Ratios (RC-R) - Existing codes
            'tier1_capital_ratio': ['RCOA7206'],
            'total_capital_ratio': ['RCOA7205'],
            'leverage_ratio': ['RCOA7204'],
            'cet1_ratio': ['RCOA7206'],
            
            # NEW: Treasury Risk & ALM Metrics
            # Earning assets and interest-bearing liabilities
            'earning_assets': ['RCFD3381', 'RCON3381'],
            'interest_bearing_liabilities': ['RCFD3353', 'RCON3353'],
            
            # Rate-sensitive assets and liabilities
            'rate_sensitive_assets': ['RCFDA564', 'RCONA564'],
            'rate_sensitive_liabilities': ['RCFDA579', 'RCONA579'],
            
            # Liquidity items
            'cash_and_due': ['RCON0081', 'RCON0071'],  # Cash and balances due
            'securities_available_for_sale': ['RCON1773', 'RCFD1773'],
            'securities_held_to_maturity': ['RCON1754', 'RCFD1754'],
            'federal_funds_sold': ['RCON1287', 'RCFD1287'],
            
            # Funding sources
            'large_time_deposits': ['RCON2604', 'RCFN2604'],  # Time deposits > $250K
            'wholesale_funding': ['RCON2800', 'RCFD2800'],  # Federal funds purchased
            'other_borrowed_money': ['RCON2930', 'RCFD2930'],
            
            # Interest income/expense components
            'interest_income_loans': ['RIAD4107'],
            'interest_income_securities': ['RIAD4020'],
            'interest_expense_deposits': ['RIAD4170'],
            'interest_expense_borrowed': ['RIAD4180'],
            
            # Asset composition
            'commercial_loans': ['RCON1600', 'RCFD1600'],
            'real_estate_loans': ['RCON1410', 'RCFD1410'],
            'consumer_loans': ['RCON1975', 'RCFD1975'],
            'trading_assets': ['RCON3545', 'RCFD3545'],
        }

    def create_executive_dashboard(self, ws, all_institution_results, processor):
        """Create Executive Dashboard sheet with visual summary including ALM metrics"""
        try:
            # Import required for charts
            from openpyxl.chart import BarChart, PieChart, Reference
            
            # Title and styling
            ws['A1'] = "Executive Dashboard"
            ws['A1'].font = Font(bold=True, size=18, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            ws.merge_cells('A1:H1')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Report date
            ws['A2'] = f"Report Date: {datetime.now().strftime('%B %d, %Y')}"
            ws['A2'].font = Font(italic=True, size=11)
            ws.merge_cells('A2:H2')
            
            # Extract metrics for each institution
            metrics_data = self._extract_key_metrics(all_institution_results, processor)
            
            # Section 1: Total Assets Comparison (Bar Chart)
            self._create_assets_comparison(ws, metrics_data, start_row=4)
            
            # Section 2: Key Financial Ratios Table
            self._create_ratios_table(ws, metrics_data, start_row=20)
            
            # NEW Section 3: Interest Rate Risk Summary
            self._create_interest_rate_risk_summary(ws, metrics_data, start_row=30)
            
            # NEW Section 4: Liquidity Risk Indicators
            self._create_liquidity_indicators(ws, metrics_data, start_row=40)
            
            # Section 5: Asset Quality Indicators - move down
            self._create_asset_quality_section(ws, metrics_data, start_row=50)
            
            # Section 6: Top 5 Balance Sheet Items - move down
            self._create_top_items_section(ws, all_institution_results, processor, start_row=60)
            
            # Apply column widths
            column_widths = {'A': 30, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Add navigation link
            ws['H3'] = "← Back to Summary"
            ws['H3'].hyperlink = "#Summary!A1"
            ws['H3'].font = Font(color="0563C1", underline="single", bold=True)
            
            self.logger.info("✓ Created Executive Dashboard with ALM metrics")
            
        except Exception as e:
            self.logger.error(f"Error creating Executive Dashboard: {str(e)}")
            raise

    def create_key_metrics_sheet(self, ws, all_institution_results, processor):
        """Create Key Metrics sheet with detailed financial ratios including ALM metrics"""
        try:
            # Title
            ws['A1'] = "Key Financial Metrics & Peer Analysis"
            ws['A1'].font = Font(bold=True, size=18, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            ws.merge_cells('A1:H1')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Extract metrics
            metrics_data = self._extract_key_metrics(all_institution_results, processor)
            
            # Section 1: Profitability Metrics
            self._create_profitability_metrics(ws, metrics_data, start_row=3)
            
            # Section 2: Efficiency Metrics
            self._create_efficiency_metrics(ws, metrics_data, start_row=13)
            
            # Section 3: Capital Adequacy
            self._create_capital_metrics(ws, metrics_data, start_row=23)
            
            # NEW Section 4: Interest Rate Risk Metrics
            self._create_interest_rate_risk_metrics(ws, metrics_data, start_row=33)
            
            # NEW Section 5: Liquidity Metrics
            self._create_liquidity_metrics_section(ws, metrics_data, start_row=45)
            
            # NEW Section 6: Balance Sheet Composition
            self._create_balance_sheet_composition(ws, metrics_data, start_row=57)
            
            # Section 7: Asset Quality Metrics - move down
            self._create_detailed_asset_quality(ws, metrics_data, start_row=69)
            
            # Section 8: Peer Group Statistics - move down
            self._create_peer_statistics(ws, metrics_data, start_row=79)
            
            # Apply formatting
            self._apply_metrics_formatting(ws)
            
            # Add navigation
            ws['H2'] = "← Back to Summary"
            ws['H2'].hyperlink = "#Summary!A1"
            ws['H2'].font = Font(color="0563C1", underline="single", bold=True)
            
            self.logger.info("✓ Created Key Metrics sheet with ALM metrics")
            
        except Exception as e:
            self.logger.error(f"Error creating Key Metrics sheet: {str(e)}")
            raise

    def _extract_key_metrics(self, all_institution_results, processor):
        """Extract key metrics from institution data"""
        metrics = {}
        
        for rssd_id, inst_data in all_institution_results.items():
            inst_metrics = {
                'name': inst_data['name'],
                'rssd_id': rssd_id
            }
            
            # Extract values from schedules
            for metric_name, mdrm_codes in self.metric_codes.items():
                value = self._find_metric_value(inst_data['data'], mdrm_codes)
                inst_metrics[metric_name] = value
            
            # Calculate derived metrics
            inst_metrics = self._calculate_derived_metrics(inst_metrics)
            
            # NEW: Calculate ALM metrics
            inst_metrics = self._calculate_alm_metrics(inst_metrics)
            
            metrics[rssd_id] = inst_metrics
        
        return metrics
    
    def _find_metric_value(self, schedule_data, mdrm_codes):
        """Find metric value from schedule data using MDRM codes"""
        # Special handling for total deposits (sum of components)
        if 'RCON2200' in mdrm_codes:  # Indicates deposit calculation
            noninterest = self._find_single_value(schedule_data, ['RCON2200', 'RCFN2200'])
            interest = self._find_single_value(schedule_data, ['RCON6631', 'RCFN6631'])
            return (noninterest or 0) + (interest or 0)
        
        # Standard single value lookup
        return self._find_single_value(schedule_data, mdrm_codes)

    def _find_single_value(self, schedule_data, mdrm_codes):
        """Helper to find a single value"""
        for schedule_code, df in schedule_data.items():
            for mdrm_code in mdrm_codes:
                mask = df['MDRM Code'] == mdrm_code
                if mask.any():
                    value = df.loc[mask, 'Amount'].iloc[0]
                    try:
                        return float(str(value).replace(',', ''))
                    except:
                        continue
        return 0
    
    def _calculate_derived_metrics(self, metrics):
        """Calculate financial ratios from base metrics"""
        # ROA (Return on Assets)
        if metrics.get('net_income') and metrics.get('total_assets'):
            # Annualize if quarterly data
            metrics['roa'] = (metrics['net_income'] * 4) / metrics['total_assets'] * 100
        else:
            metrics['roa'] = 0
        
        # ROE (Return on Equity)
        if metrics.get('net_income') and metrics.get('total_equity'):
            metrics['roe'] = (metrics['net_income'] * 4) / metrics['total_equity'] * 100
        else:
            metrics['roe'] = 0
        
        # Net Interest Margin
        if metrics.get('net_interest_income') and metrics.get('total_assets'):
            metrics['nim'] = (metrics['net_interest_income'] * 4) / metrics['total_assets'] * 100
        else:
            metrics['nim'] = 0
        
        # Efficiency Ratio
        if (metrics.get('net_interest_income', 0) + metrics.get('noninterest_income', 0)) > 0:
            metrics['efficiency_ratio'] = (
                metrics.get('noninterest_expense', 0) / 
                (metrics.get('net_interest_income', 0) + metrics.get('noninterest_income', 0))
            ) * 100
        else:
            metrics['efficiency_ratio'] = 0
        
        # Loan to Deposit Ratio
        if metrics.get('total_deposits'):
            metrics['loan_deposit_ratio'] = (
                metrics.get('total_loans', 0) / metrics['total_deposits']
            ) * 100
        else:
            metrics['loan_deposit_ratio'] = 0
        
        # NPL Ratio (Non-performing loans)
        if metrics.get('total_loans'):
            npl = metrics.get('nonaccrual_loans', 0) + metrics.get('past_due_90', 0)
            metrics['npl_ratio'] = (npl / metrics['total_loans']) * 100
        else:
            metrics['npl_ratio'] = 0
        
        return metrics
    
    def _calculate_alm_metrics(self, metrics):
        """Calculate ALM and treasury risk metrics"""
        # Interest Rate Sensitivity Ratio
        if metrics.get('rate_sensitive_assets') and metrics.get('rate_sensitive_liabilities'):
            metrics['ir_sensitivity_ratio'] = (
                metrics['rate_sensitive_assets'] / metrics['rate_sensitive_liabilities']
            ) * 100
        else:
            metrics['ir_sensitivity_ratio'] = 0
        
        # Liquid Assets Ratio
        liquid_assets = (
            metrics.get('cash_and_due', 0) +
            metrics.get('securities_available_for_sale', 0) +
            metrics.get('federal_funds_sold', 0)
        )
        if metrics.get('total_assets'):
            metrics['liquid_assets_ratio'] = (liquid_assets / metrics['total_assets']) * 100
        else:
            metrics['liquid_assets_ratio'] = 0
        
        # Wholesale Funding Dependency
        wholesale_funds = (
            metrics.get('wholesale_funding', 0) +
            metrics.get('other_borrowed_money', 0) +
            metrics.get('large_time_deposits', 0)
        )
        if metrics.get('total_assets'):
            metrics['wholesale_funding_ratio'] = (wholesale_funds / metrics['total_assets']) * 100
        else:
            metrics['wholesale_funding_ratio'] = 0
        
        # Earning Asset Yield
        if metrics.get('interest_income_loans') and metrics.get('earning_assets'):
            # Annualize if quarterly
            metrics['earning_asset_yield'] = (
                (metrics['interest_income_loans'] * 4) / metrics['earning_assets']
            ) * 100
        else:
            metrics['earning_asset_yield'] = 0
        
        # Cost of Funds
        total_interest_expense = (
            metrics.get('interest_expense_deposits', 0) +
            metrics.get('interest_expense_borrowed', 0)
        )
        if total_interest_expense and metrics.get('interest_bearing_liabilities'):
            metrics['cost_of_funds'] = (
                (total_interest_expense * 4) / metrics['interest_bearing_liabilities']
            ) * 100
        else:
            metrics['cost_of_funds'] = 0
        
        # Interest Rate Spread
        metrics['interest_rate_spread'] = (
            metrics.get('earning_asset_yield', 0) - metrics.get('cost_of_funds', 0)
        )
        
        # Fixed vs Variable Rate Mix (simplified calculation)
        # In practice, this would require more detailed MDRM codes
        fixed_rate_assets = metrics.get('securities_held_to_maturity', 0)
        variable_rate_assets = metrics.get('commercial_loans', 0) * 0.7  # Assume 70% variable
        total_rate_assets = fixed_rate_assets + variable_rate_assets
        
        if total_rate_assets > 0:
            metrics['fixed_rate_asset_pct'] = (fixed_rate_assets / total_rate_assets) * 100
            metrics['variable_rate_asset_pct'] = (variable_rate_assets / total_rate_assets) * 100
        else:
            metrics['fixed_rate_asset_pct'] = 0
            metrics['variable_rate_asset_pct'] = 0
        
        # Interest Rate Gap (simplified)
        metrics['interest_rate_gap'] = (
            metrics.get('rate_sensitive_assets', 0) -
            metrics.get('rate_sensitive_liabilities', 0)
        )
        
        # Gap Ratio
        if metrics.get('total_assets'):
            metrics['gap_ratio'] = (metrics['interest_rate_gap'] / metrics['total_assets']) * 100
        else:
            metrics['gap_ratio'] = 0
        
        # Deposit Stability Ratio (core deposits / total deposits)
        core_deposits = metrics.get('total_deposits', 0) - metrics.get('large_time_deposits', 0)
        if metrics.get('total_deposits'):
            metrics['deposit_stability_ratio'] = (core_deposits / metrics['total_deposits']) * 100
        else:
            metrics['deposit_stability_ratio'] = 0
        
        return metrics

    def _create_assets_comparison(self, ws, metrics_data, start_row):
        """Create total assets comparison with bar chart"""
        from openpyxl.chart import BarChart, Reference
        
        ws.cell(row=start_row, column=1, value="Total Assets Comparison").font = Font(bold=True, size=14)
        
        # Headers
        row = start_row + 2
        ws.cell(row=row, column=1, value="Institution")
        ws.cell(row=row, column=2, value="Total Assets ($000)")
        ws.cell(row=row, column=3, value="% of Largest")
        
        # Apply header formatting
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        assets_data = [(m['name'], m.get('total_assets', 0)) for m in metrics_data.values()]
        assets_data.sort(key=lambda x: x[1], reverse=True)
        
        max_assets = assets_data[0][1] if assets_data else 1
        
        for i, (name, assets) in enumerate(assets_data):
            row = start_row + 3 + i
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=assets).number_format = '#,##0'
            ws.cell(row=row, column=3, value=f"{(assets/max_assets*100):.1f}%")
        
        # Create bar chart
        chart = BarChart()
        chart.title = "Total Assets by Institution"
        chart.style = 10
        chart.height = 8
        chart.width = 10
        
        # Add data
        data = Reference(ws, min_col=2, min_row=start_row+2, max_row=start_row+2+len(assets_data))
        cats = Reference(ws, min_col=1, min_row=start_row+3, max_row=start_row+2+len(assets_data))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Add data labels
        from openpyxl.chart.label import DataLabelList
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
                
        # Format y-axis for millions/billions
        chart.y_axis.number_format = '#,##0,,"M"'
        chart.y_axis.title = "Total Assets (Millions)"
        
        ws.add_chart(chart, f"E{start_row}")
  
    def _create_ratios_table(self, ws, metrics_data, start_row):
        """Create key financial ratios table with single/multi institution support"""
        ws.cell(row=start_row, column=1, value="Key Financial Ratios").font = Font(bold=True, size=14)
        
        # Detect if single or multi institution
        is_single_institution = len(metrics_data) == 1
        
        # Headers
        row = start_row + 2
        if is_single_institution:
            headers = ["Metric", "Value", "Status"]
        else:
            headers = ["Metric", "Primary"] + [f"Peer {i}" for i in range(1, len(metrics_data))] + ["Peer Avg"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Ratio rows
        ratios = [
            ("ROA (%)", 'roa', '0.00'),
            ("ROE (%)", 'roe', '0.00'),
            ("Net Interest Margin (%)", 'nim', '0.00'),
            ("Efficiency Ratio (%)", 'efficiency_ratio', '0.0'),
            ("Tier 1 Capital Ratio (%)", 'tier1_capital_ratio', '0.00'),
            ("Loan/Deposit Ratio (%)", 'loan_deposit_ratio', '0.0')
        ]
        
        inst_list = list(metrics_data.values())
        
        for i, (label, metric_key, fmt) in enumerate(ratios):
            row = start_row + 3 + i
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            
            if is_single_institution:
                # Single institution mode
                value = inst_list[0].get(metric_key, 0)
                cell = ws.cell(row=row, column=2, value=value)
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal='center')
                
                # Status column
                status, color = self._get_ratio_status(metric_key, value)
                status_cell = ws.cell(row=row, column=3, value=status)
                status_cell.font = Font(color=color, bold=True)
                status_cell.alignment = Alignment(horizontal='center')
            else:
                # Multi institution mode (existing logic)
                values = []
                for col, inst in enumerate(inst_list, 2):
                    value = inst.get(metric_key, 0)
                    values.append(value)
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.number_format = fmt
                    cell.alignment = Alignment(horizontal='center')
                
                # Peer average (excluding primary)
                if len(values) > 1:
                    peer_avg = sum(values[1:]) / len(values[1:])
                    cell = ws.cell(row=row, column=len(inst_list)+2, value=peer_avg)
                    cell.number_format = fmt
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(italic=True)

    def _get_ratio_status(self, metric_key, value):
        """Determine status and color for financial ratios based on industry benchmarks"""
        # Industry benchmark ranges (these would ideally come from a database)
        benchmarks = {
            'roa': {
                'excellent': (1.5, float('inf')),
                'good': (1.0, 1.5),
                'adequate': (0.5, 1.0),
                'needs_improvement': (-float('inf'), 0.5)
            },
            'roe': {
                'excellent': (15.0, float('inf')),
                'good': (10.0, 15.0),
                'adequate': (5.0, 10.0),
                'needs_improvement': (-float('inf'), 5.0)
            },
            'nim': {
                'excellent': (4.0, float('inf')),
                'good': (3.0, 4.0),
                'adequate': (2.0, 3.0),
                'needs_improvement': (-float('inf'), 2.0)
            },
            'efficiency_ratio': {
                'excellent': (-float('inf'), 50.0),
                'good': (50.0, 60.0),
                'adequate': (60.0, 70.0),
                'needs_improvement': (70.0, float('inf'))
            },
            'tier1_capital_ratio': {
                'excellent': (12.0, float('inf')),
                'good': (10.0, 12.0),
                'adequate': (8.0, 10.0),
                'needs_improvement': (-float('inf'), 8.0)
            },
            'loan_deposit_ratio': {
                'excellent': (70.0, 90.0),
                'good': (60.0, 70.0) if value < 70 else (90.0, 100.0),
                'adequate': (50.0, 60.0) if value < 70 else (100.0, 110.0),
                'needs_improvement': (-float('inf'), 50.0) if value < 70 else (110.0, float('inf'))
            }
        }
        
        # Get benchmarks for this metric
        metric_benchmarks = benchmarks.get(metric_key, {})
        
        # Determine status
        for status, (min_val, max_val) in [
            ('Excellent', metric_benchmarks.get('excellent', (float('inf'), float('inf')))),
            ('Good', metric_benchmarks.get('good', (float('inf'), float('inf')))),
            ('Adequate', metric_benchmarks.get('adequate', (float('inf'), float('inf')))),
            ('Needs Improvement', metric_benchmarks.get('needs_improvement', (float('inf'), float('inf'))))
        ]:
            if min_val <= value < max_val:
                colors = {
                    'Excellent': '008000',  # Green
                    'Good': '0563C1',       # Blue
                    'Adequate': 'FFA500',   # Orange
                    'Needs Improvement': 'FF0000'  # Red
                }
                return status, colors.get(status, '000000')
        
        return 'N/A', '666666'  # Gray

    def _create_peer_statistics(self, ws, metrics_data, start_row):
        """Create peer group statistical analysis (adapted for single institution)"""
        ws.cell(row=start_row, column=1, value="Peer Group Statistical Analysis").font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells(f'A{start_row}:E{start_row}')
        
        # For single institution, show industry benchmarks instead
        if len(metrics_data) == 1:
            row = start_row + 2
            ws.cell(row=row, column=1, value="Industry Benchmarks").font = Font(bold=True, size=12)
            row += 1
            
            # Show benchmark ranges
            benchmarks = [
                ("ROA (%)", "1.0 - 1.5", "Industry average for well-performing banks"),
                ("ROE (%)", "10 - 15", "Typical range for efficient capital utilization"),
                ("Net Interest Margin (%)", "3.0 - 4.0", "Standard range for commercial banks"),
                ("Efficiency Ratio (%)", "50 - 60", "Lower is better; best-in-class below 50%"),
                ("NPL Ratio (%)", "< 2.0", "Below 2% indicates good asset quality"),
                ("Tier 1 Capital (%)", "> 10", "Regulatory minimum is 6%; well-capitalized > 10%")
            ]
            
            headers = ["Metric", "Benchmark Range", "Description"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            for i, (metric, benchmark, description) in enumerate(benchmarks):
                row_num = row + 1 + i
                ws.cell(row=row_num, column=1, value=metric).font = Font(bold=True)
                ws.cell(row=row_num, column=2, value=benchmark)
                ws.cell(row=row_num, column=3, value=description)
        else:
            # Multi-institution statistics (existing logic)
            # Calculate statistics
            row = start_row + 2
            headers = ["Metric", "Mean", "Median", "Std Dev", "Primary Percentile"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Key metrics for statistical analysis
            stat_metrics = [
                ("Total Assets ($MM)", 'total_assets', 1000000),
                ("ROA (%)", 'roa', 1),
                ("ROE (%)", 'roe', 1),
                ("Efficiency Ratio (%)", 'efficiency_ratio', 1),
                ("NPL Ratio (%)", 'npl_ratio', 1),
            ]
            
            primary_inst = list(metrics_data.values())[0]
            
            for i, (label, key, divisor) in enumerate(stat_metrics):
                row = start_row + 3 + i
                values = [inst.get(key, 0) / divisor for inst in metrics_data.values()]
                
                if values:
                    mean_val = np.mean(values)
                    median_val = np.median(values)
                    std_val = np.std(values)
                    
                    # Calculate percentile for primary institution
                    primary_val = primary_inst.get(key, 0) / divisor
                    percentile = (sum(1 for v in values if v <= primary_val) / len(values)) * 100
                    
                    ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                    ws.cell(row=row, column=2, value=mean_val).number_format = '0.00'
                    ws.cell(row=row, column=3, value=median_val).number_format = '0.00'
                    ws.cell(row=row, column=4, value=std_val).number_format = '0.00'
                    ws.cell(row=row, column=5, value=f"{percentile:.0f}th")
        
    def _create_interest_rate_risk_summary(self, ws, metrics_data, start_row):
        """Create interest rate risk summary panel for Executive Dashboard"""
        ws.cell(row=start_row, column=1, value="Interest Rate Risk Summary").font = Font(bold=True, size=14)
        
        # Headers
        row = start_row + 2
        headers = ["Institution", "IR Sensitivity", "Gap Ratio (%)", "Asset Yield (%)", "Cost of Funds (%)"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for i, inst in enumerate(metrics_data.values()):
            row = start_row + 3 + i
            ws.cell(row=row, column=1, value=inst['name'])
            ws.cell(row=row, column=2, value=inst.get('ir_sensitivity_ratio', 0)).number_format = '0.00'
            ws.cell(row=row, column=3, value=inst.get('gap_ratio', 0)).number_format = '0.00'
            ws.cell(row=row, column=4, value=inst.get('earning_asset_yield', 0)).number_format = '0.00'
            ws.cell(row=row, column=5, value=inst.get('cost_of_funds', 0)).number_format = '0.00'
            
            # Conditional formatting for gap ratio
            gap_cell = ws.cell(row=row, column=3)
            gap_value = inst.get('gap_ratio', 0)
            if abs(gap_value) > 20:
                gap_cell.font = Font(color="FF0000", bold=True)  # Red for high gap
            elif abs(gap_value) > 10:
                gap_cell.font = Font(color="FFA500")  # Orange for moderate gap

    def _create_liquidity_indicators(self, ws, metrics_data, start_row):
        """Create liquidity risk indicators for Executive Dashboard"""
        ws.cell(row=start_row, column=1, value="Liquidity Risk Indicators").font = Font(bold=True, size=14)
        
        # Headers
        row = start_row + 2
        headers = ["Institution", "Liquid Assets (%)", "Wholesale Funding (%)", "Deposit Stability (%)"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows with risk indicators
        for i, inst in enumerate(metrics_data.values()):
            row = start_row + 3 + i
            ws.cell(row=row, column=1, value=inst['name'])
            
            # Liquid Assets Ratio
            liquid_ratio = inst.get('liquid_assets_ratio', 0)
            cell = ws.cell(row=row, column=2, value=liquid_ratio)
            cell.number_format = '0.0'
            if liquid_ratio < 10:
                cell.font = Font(color="FF0000")  # Red for low liquidity
            elif liquid_ratio < 20:
                cell.font = Font(color="FFA500")  # Orange for moderate
            else:
                cell.font = Font(color="008000")  # Green for good
            
            # Wholesale Funding
            wholesale_ratio = inst.get('wholesale_funding_ratio', 0)
            cell = ws.cell(row=row, column=3, value=wholesale_ratio)
            cell.number_format = '0.0'
            if wholesale_ratio > 30:
                cell.font = Font(color="FF0000")  # Red for high dependency
            elif wholesale_ratio > 20:
                cell.font = Font(color="FFA500")  # Orange
            
            # Deposit Stability
            stability = inst.get('deposit_stability_ratio', 0)
            cell = ws.cell(row=row, column=4, value=stability)
            cell.number_format = '0.0'
            if stability < 70:
                cell.font = Font(color="FF0000")  # Red for low stability
            elif stability < 80:
                cell.font = Font(color="FFA500")  # Orange
            else:
                cell.font = Font(color="008000")  # Green

    def _create_asset_quality_section(self, ws, metrics_data, start_row):
        """Create asset quality indicators section"""
        ws.cell(row=start_row, column=1, value="Asset Quality Indicators").font = Font(bold=True, size=14)
        
        # Calculate NPL ratios
        row = start_row + 2
        headers = ["Institution", "NPL Ratio (%)", "Status"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data rows with conditional formatting
        for i, inst in enumerate(metrics_data.values()):
            row = start_row + 3 + i
            npl_ratio = inst.get('npl_ratio', 0)
            
            ws.cell(row=row, column=1, value=inst['name'])
            ws.cell(row=row, column=2, value=npl_ratio).number_format = '0.00'
            
            # Status with color coding
            if npl_ratio < 1:
                status = "Excellent"
                color = "008000"  # Green
            elif npl_ratio < 2:
                status = "Good"
                color = "FFA500"  # Orange
            else:
                status = "Needs Attention"
                color = "FF0000"  # Red
            
            status_cell = ws.cell(row=row, column=3, value=status)
            status_cell.font = Font(color=color, bold=True)
    
    def _create_top_items_section(self, ws, all_institution_results, processor, start_row):
        """Create top 5 balance sheet items section"""
        ws.cell(row=start_row, column=1, value="Top 5 Balance Sheet Items (Primary Institution)").font = Font(bold=True, size=14)
        
        # Get primary institution data
        primary_data = list(all_institution_results.values())[0]
        
        # Find RC schedule data
        rc_data = None
        for schedule_code, df in primary_data['data'].items():
            if schedule_code == 'RC':
                rc_data = df
                break
        
        if rc_data is not None and not rc_data.empty:
            # Get top 5 items by amount
            rc_sorted = rc_data.copy()
            rc_sorted['Amount_num'] = pd.to_numeric(rc_sorted['Amount'].astype(str).str.replace(',', ''), errors='coerce')
            top_items = rc_sorted.nlargest(5, 'Amount_num')
            
            # Headers
            row = start_row + 2
            headers = ["Line Item", "Description", "Amount ($000)"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Data rows
            for i, (idx, item) in enumerate(top_items.iterrows()):
                row = start_row + 3 + i
                ws.cell(row=row, column=1, value=item.get('Line Item', ''))
                ws.cell(row=row, column=2, value=item.get('Description', '')[:50])  # Truncate long descriptions
                ws.cell(row=row, column=3, value=item.get('Amount_num', 0)).number_format = '#,##0'
                
    def _create_profitability_metrics(self, ws, metrics_data, start_row):
        """Create profitability metrics section"""
        ws.cell(row=start_row, column=1, value="Profitability Metrics").font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells(f'A{start_row}:H{start_row}')
        
        # Section styling
        for col in range(1, 9):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Add detailed profitability metrics
        row = start_row + 2
        metrics = [
            ("Net Income ($000)", 'net_income', '#,##0'),
            ("Net Interest Income ($000)", 'net_interest_income', '#,##0'),
            ("Non-Interest Income ($000)", 'noninterest_income', '#,##0'),
            ("Pre-Tax Income ($000)", None, '#,##0'),  # Calculate if tax data available
            ("Operating Income ($000)", None, '#,##0'),  # Calculate if available
        ]
        
        self._create_metric_rows(ws, metrics_data, metrics, row)
    
    def _create_efficiency_metrics(self, ws, metrics_data, start_row):
        """Create efficiency metrics section"""
        ws.cell(row=start_row, column=1, value="Efficiency Metrics").font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells(f'A{start_row}:H{start_row}')
        
        # Section styling
        for col in range(1, 9):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row = start_row + 2
        metrics = [
            ("Efficiency Ratio (%)", 'efficiency_ratio', '0.0'),
            ("Operating Expense Ratio (%)", None, '0.00'),
            ("Cost of Funds (%)", 'cost_of_funds', '0.00'),
            ("Non-Interest Expense ($000)", 'noninterest_expense', '#,##0'),
        ]
                
        self._create_metric_rows(ws, metrics_data, metrics, row)
    
    def _create_capital_metrics(self, ws, metrics_data, start_row):
        """Create capital adequacy metrics section"""
        ws.cell(row=start_row, column=1, value="Capital Adequacy").font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells(f'A{start_row}:H{start_row}')
        
        # Section styling
        for col in range(1, 9):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row = start_row + 2
        metrics = [
            ("CET1 Ratio (%)", 'cet1_ratio', '0.00'),
            ("Tier 1 Capital Ratio (%)", 'tier1_capital_ratio', '0.00'),
            ("Total Capital Ratio (%)", 'total_capital_ratio', '0.00'),
            ("Leverage Ratio (%)", 'leverage_ratio', '0.00'),
            ("Total Equity ($000)", 'total_equity', '#,##0'),
        ]
                
        self._create_metric_rows(ws, metrics_data, metrics, row)
    
    def _create_interest_rate_risk_metrics(self, ws, metrics_data, start_row):
        """Create detailed interest rate risk metrics section"""
        ws.cell(row=start_row, column=1, value="Interest Rate Risk Metrics").font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells(f'A{start_row}:H{start_row}')
        
        # Section styling
        for col in range(1, 9):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row = start_row + 2
        metrics = [
            ("Earning Asset Yield (%)", 'earning_asset_yield', '0.00'),
            ("Cost of Interest-Bearing Liabilities (%)", 'cost_of_funds', '0.00'),
            ("Interest Rate Spread (%)", 'interest_rate_spread', '0.00'),
            ("Rate-Sensitive Assets ($000)", 'rate_sensitive_assets', '#,##0'),
            ("Rate-Sensitive Liabilities ($000)", 'rate_sensitive_liabilities', '#,##0'),
            ("Interest Rate Gap ($000)", 'interest_rate_gap', '#,##0'),
            ("Gap Ratio (%)", 'gap_ratio', '0.00'),
            ("IR Sensitivity Ratio", 'ir_sensitivity_ratio', '0.00'),
        ]
        
        self._create_metric_rows(ws, metrics_data, metrics, row)

    def _create_liquidity_metrics_section(self, ws, metrics_data, start_row):
        """Create detailed liquidity metrics section"""
        ws.cell(row=start_row, column=1, value="Liquidity Metrics").font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells(f'A{start_row}:H{start_row}')
        
        # Section styling
        for col in range(1, 9):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row = start_row + 2
        metrics = [
            ("Cash & Due From Banks ($000)", 'cash_and_due', '#,##0'),
            ("Securities Available for Sale ($000)", 'securities_available_for_sale', '#,##0'),
            ("Liquid Assets / Total Assets (%)", 'liquid_assets_ratio', '0.00'),
            ("Wholesale Funding / Total Assets (%)", 'wholesale_funding_ratio', '0.00'),
            ("Large Time Deposits ($000)", 'large_time_deposits', '#,##0'),
            ("Deposit Stability Ratio (%)", 'deposit_stability_ratio', '0.00'),
        ]
        
        self._create_metric_rows(ws, metrics_data, metrics, row)

    def _create_balance_sheet_composition(self, ws, metrics_data, start_row):
        """Create balance sheet composition analysis"""
        ws.cell(row=start_row, column=1, value="Balance Sheet Composition Analysis").font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells(f'A{start_row}:H{start_row}')
        
        # Section styling
        for col in range(1, 9):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row = start_row + 2
        
        # Asset Mix Ratios
        for inst in metrics_data.values():
            if inst.get('total_assets', 0) > 0:
                inst['loans_to_assets'] = (inst.get('total_loans', 0) / inst['total_assets']) * 100
                inst['securities_to_assets'] = ((inst.get('securities_available_for_sale', 0) +
                                                inst.get('securities_held_to_maturity', 0)) / inst['total_assets']) * 100
                inst['cash_to_assets'] = (inst.get('cash_and_due', 0) / inst['total_assets']) * 100
        
        asset_metrics = [
            ("Loans / Total Assets (%)", 'loans_to_assets', '0.0'),
            ("Securities / Total Assets (%)", 'securities_to_assets', '0.0'),
            ("Cash / Total Assets (%)", 'cash_to_assets', '0.0'),
            ("Fixed Rate Assets (%)", 'fixed_rate_asset_pct', '0.0'),
            ("Variable Rate Assets (%)", 'variable_rate_asset_pct', '0.0'),
        ]
        
        self._create_metric_rows(ws, metrics_data, asset_metrics, row)

    def _create_detailed_asset_quality(self, ws, metrics_data, start_row):
        """Create detailed asset quality metrics"""
        ws.cell(row=start_row, column=1, value="Asset Quality Metrics").font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells(f'A{start_row}:H{start_row}')
        
        # Section styling
        for col in range(1, 9):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row = start_row + 2
        metrics = [
            ("NPL Ratio (%)", 'npl_ratio', '0.00'),
            ("Non-Accrual Loans ($000)", 'nonaccrual_loans', '#,##0'),
            ("Past Due 90+ Days ($000)", 'past_due_90', '#,##0'),
            ("Provision for Loan Losses ($000)", 'provision_loan_losses', '#,##0'),
            ("Net Charge-offs ($000)", None, '#,##0'),  # If available
        ]
                
        self._create_metric_rows(ws, metrics_data, metrics, row)
    
    def _create_peer_statistics(self, ws, metrics_data, start_row):
        """Create peer group statistical analysis"""
        ws.cell(row=start_row, column=1, value="Peer Group Statistical Analysis").font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells(f'A{start_row}:E{start_row}')
        
        # Calculate statistics
        row = start_row + 2
        headers = ["Metric", "Mean", "Median", "Std Dev", "Primary Percentile"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Key metrics for statistical analysis
        stat_metrics = [
            ("Total Assets ($MM)", 'total_assets', 1000000),
            ("ROA (%)", 'roa', 1),
            ("ROE (%)", 'roe', 1),
            ("Efficiency Ratio (%)", 'efficiency_ratio', 1),
            ("NPL Ratio (%)", 'npl_ratio', 1),
        ]
        
        primary_inst = list(metrics_data.values())[0]
        
        for i, (label, key, divisor) in enumerate(stat_metrics):
            row = start_row + 3 + i
            values = [inst.get(key, 0) / divisor for inst in metrics_data.values()]
            
            if values:
                mean_val = np.mean(values)
                median_val = np.median(values)
                std_val = np.std(values)
                
                # Calculate percentile for primary institution
                primary_val = primary_inst.get(key, 0) / divisor
                percentile = (sum(1 for v in values if v <= primary_val) / len(values)) * 100
                
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=mean_val).number_format = '0.00'
                ws.cell(row=row, column=3, value=median_val).number_format = '0.00'
                ws.cell(row=row, column=4, value=std_val).number_format = '0.00'
                ws.cell(row=row, column=5, value=f"{percentile:.0f}th")
   
    def _create_metric_rows(self, ws, metrics_data, metrics_list, start_row):
        """Helper to create metric rows with consistent formatting"""
        # Headers
        inst_list = list(metrics_data.values())
        headers = ["Metric", inst_list[0]['name']] + [inst['name'] for inst in inst_list[1:]]
        
        # Add "Peer Avg" column if more than one institution
        if len(inst_list) > 1:
            headers.append("Peer Avg")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for i, (label, metric_key, fmt) in enumerate(metrics_list):
            row = start_row + 1 + i
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            
            if metric_key:
                values = []
                for col, inst in enumerate(inst_list, 2):
                    value = inst.get(metric_key, 0)
                    values.append(value)
                    cell = ws.cell(row=row, column=col, value=value)
                    # Apply number format regardless of value
                    cell.number_format = fmt
                    cell.alignment = Alignment(horizontal='center')
                
                # Peer average (excluding primary)
                if len(values) > 1:
                    peer_avg = sum(values[1:]) / len(values[1:])
                    cell = ws.cell(row=row, column=len(inst_list)+2, value=peer_avg)
                    # Apply number format regardless of value
                    cell.number_format = fmt
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(italic=True)
            else:
                # For rows without metric_key (calculated fields), leave cells empty
                for col in range(2, len(headers) + 1):
                    ws.cell(row=row, column=col, value="")
            
            # Apply alternating row colors
            if i % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                    )
                    
    def _apply_metrics_formatting(self, ws):
        """Apply consistent formatting to metrics sheet"""
        # Column widths
        column_widths = {
            'A': 35,  # Metric names
            'B': 20,  # Primary institution
            'C': 20,  # Peer 1
            'D': 20,  # Peer 2
            'E': 20,  # Peer 3
            'F': 15,  # Peer Avg
            'G': 15,  # Extra
            'H': 15   # Extra
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Add borders to all cells with data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=8):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border

    def _create_asset_liability_pie_chart(self, ws, metrics_data, position):
        """Create pie chart showing fixed vs variable rate asset mix"""
        from openpyxl.chart import PieChart, Reference
        
        # Get primary institution data
        primary_inst = list(metrics_data.values())[0]
        
        # Create data for chart
        data = [
            ['Asset Type', 'Percentage'],
            ['Fixed Rate', primary_inst.get('fixed_rate_asset_pct', 0)],
            ['Variable Rate', primary_inst.get('variable_rate_asset_pct', 0)]
        ]
        
        # Write data to hidden area
        start_row = 100  # Use row 100+ for chart data
        for i, row_data in enumerate(data):
            for j, value in enumerate(row_data):
                ws.cell(row=start_row + i, column=10 + j, value=value)
        
        # Create pie chart
        pie = PieChart()
        pie.title = "Fixed vs Variable Rate Asset Mix"
        pie.style = 10
        pie.height = 7
        pie.width = 10
        
        # Add data
        labels = Reference(ws, min_col=10, min_row=start_row+1, max_row=start_row+2)
        data = Reference(ws, min_col=11, min_row=start_row, max_row=start_row+2)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        
        # Add data labels
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        
        ws.add_chart(pie, position)


# Example usage and testing
if __name__ == "__main__":
    # Setup logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Initialize processor
    processor = BulkDataProcessor()
    
    # Example: Process a single file
    sample_file = "FFIEC CDR Call Schedule RCT 03312025(2 of 2).txt"
    if os.path.exists(sample_file):
        df = processor.process_bulk_file(sample_file, target_rssd_id="2277860")  # Capital One
        print(df.head())
    
    # Example: Process directory
    # results = processor.process_directory("./bulk_data/", target_rssd_id="2277860")
    # processor.save_to_excel(results, "CapitalOne_CallReport_Q1_2025.xlsx", "Capital One Financial Corp")