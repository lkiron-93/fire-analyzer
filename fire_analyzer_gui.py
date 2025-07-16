"""
FIRE - Financial Institution Regulatory Extractor
A comprehensive GUI application for extracting and analyzing regulatory filings and financial documents

Features:
- Multi-company support with SEC EDGAR API integration
- Support for various filing types (10-K, 10-Q, 8-K, etc.)
- Call Report support (PDF, XBRL, SDF formats)
- Local file analysis with document type identification
- Advanced table extraction with formatting preservation
- Multiple export formats (Excel, CSV, JSON)
- Modern dark theme UI with intuitive navigation

Author: FIRE Development Team
Version: 2.0
"""

# ===== STANDARD LIBRARY IMPORTS =====
# These are built into Python - no installation needed
import json          # For reading/writing JSON data
import os            # For file system operations
import queue         # For thread-safe communication between GUI and background tasks
import subprocess    # For running external commands (like pip install)
import sys           # For system-specific parameters and functions
import re            # For regular expression pattern matching
import threading     # For running tasks in background without freezing GUI
import tkinter as tk # Main GUI framework
import sqlite3       # For database operations
import webbrowser    # For opening web pages
from datetime import datetime  # For timestamps and date formatting
from tkinter import filedialog, messagebox, scrolledtext, ttk  # Additional GUI components

# ===== THIRD-PARTY IMPORTS =====
# These will be dynamically imported after checking if they're installed
# This prevents the app from crashing if libraries are missing
# - pandas: Data manipulation and analysis
# - requests: HTTP library for API calls
# - beautifulsoup4: HTML/XML parsing
# - openpyxl: Excel file handling
# - lxml: XML parser
# - numpy: Numerical operations
# - yfinance: Yahoo Finance API (optional)
# - xlsxwriter: Enhanced Excel formatting
# - pdfplumber: PDF parsing for Call Reports

# ===== LOCAL IMPORTS =====
# Import our custom modules for bulk data processing
from bulk_file_manager import BulkDataOrganizer, BulkFileManager, BulkFileMetadata
# For Excel Prototype Test
from bulk_data_processor import ExcelEnhancementProcessor

# ===== GLOBAL VARIABLES =====
# Dictionary to store imported libraries after verification
# This allows us to use libraries dynamically without causing import errors
imported_libs = {}

def check_and_import_libraries():
    """
    Check if required libraries are installed and import them dynamically
    
    This function attempts to import all required packages and keeps track
    of any missing ones. It uses dynamic imports to avoid errors if packages
    are not installed.
    
    Returns:
        list: Names of packages that failed to import
    """
    # Define required packages with their import names
    # Format: 'package_name': 'import_name'
    required_packages = {
        'requests': 'requests',           # HTTP library for API calls
        'pandas': 'pandas',               # Data manipulation
        'beautifulsoup4': 'bs4',          # HTML/XML parsing
        'openpyxl': 'openpyxl',          # Excel file handling
        'lxml': 'lxml',                   # XML parser
        'numpy': 'numpy',                 # Numerical operations
        'yfinance': 'yfinance',           # Yahoo Finance API (optional)
        'xlsxwriter': 'xlsxwriter',       # Enhanced Excel formatting
        'pdfplumber': 'pdfplumber'        # PDF parsing for Call Reports
    }
    
    missing_packages = []
    
    # Try to import each package
    for package_name, import_name in required_packages.items():
        try:
            if import_name == 'bs4':
                # Special handling for BeautifulSoup
                # We need to import the BeautifulSoup class from the bs4 module
                imported_libs['BeautifulSoup'] = __import__('bs4', fromlist=['BeautifulSoup']).BeautifulSoup
            else:
                # Standard import for other packages
                imported_libs[import_name] = __import__(import_name)
        except ImportError:
            # Package not installed - add to missing list
            missing_packages.append(package_name)
    
    return missing_packages

class FIREAnalyzer:
    """
    Main application class for the FIRE (Financial Institution Regulatory Extractor) GUI
    
    This class creates and manages the entire GUI application, including:
    - Tab-based interface for different workflows
    - Library installation management
    - Filing analysis coordination
    - Export functionality
    - Thread management for non-blocking operations
    """
    
    def __init__(self, root):
        """
        Initialize the FIRE Analyzer application
        
        Args:
            root: Tkinter root window object
        """
        # Store reference to the main window
        self.root = root
        self.root.title("🔥 FIRE - Financial Institution Regulatory Extractor")
        
        # Set larger default window size and minimum size
        self.root.geometry("1200x850")
        self.root.minsize(1000, 700)
        
        # Start maximized on Windows for better user experience
        if sys.platform == 'win32':
            self.root.state('zoomed')
        
        # Try to set application icon if available
        try:
            self.root.iconbitmap('icon.ico')
        except:
            pass  # Icon file not found, continue without it
        
        # Initialize the style system for modern appearance
        self.style = ttk.Style()
        self.style.theme_use('clam')  # Modern theme

        # Define color scheme for dark theme
        # These colors create a professional, modern appearance
        self.bg_color = "#1e1e1e"          # Dark background
        self.card_color = "#2d2d2d"        # Card backgrounds
        self.primary_color = "#007acc"     # Blue accent (Microsoft VSCode blue)
        self.secondary_color = "#4fc3f7"   # Light blue
        self.success_color = "#4caf50"     # Green for success states
        self.error_color = "#f44336"       # Red for errors
        self.text_color = "#ffffff"        # White text
        self.muted_color = "#b0bec5"       # Muted/secondary text

        # Apply the color scheme
        self.root.configure(bg=self.bg_color)
        self.configure_modern_styles()
        
        # Initialize queue for thread communication
        # This allows background threads to send updates to the GUI safely
        self.queue = queue.Queue()
        
        # Check which required libraries are missing
        self.missing_packages = check_and_import_libraries()
        
        # Build the GUI interface
        self.create_gui()
        
        # Start monitoring the queue for messages from background threads
        # Check every 100ms for new messages
        self.root.after(100, self.process_queue)
        
    def configure_modern_styles(self):
        """
        Configure TTK styles for a modern, professional appearance
        
        This method sets up custom styles for all TTK widgets used in the
        application, creating a cohesive dark theme design.
        """
        # Configure notebook (tab container) style
        self.style.configure('TNotebook', 
                            background=self.bg_color,
                            borderwidth=0)
        self.style.configure('TNotebook.Tab',
                            background=self.card_color,
                            foreground=self.text_color,
                            padding=[20, 10],
                            font=('Arial', 10, 'bold'))
        # Tab appearance changes on hover and selection
        self.style.map('TNotebook.Tab',
                      background=[('selected', self.primary_color),
                                ('active', self.secondary_color)])
        
        # Configure frame styles for grouping elements
        self.style.configure('TLabelFrame',
                            background=self.bg_color,
                            foreground=self.text_color,
                            borderwidth=1,
                            relief='solid')
        self.style.configure('TLabelFrame.Label',
                            background=self.bg_color,
                            foreground=self.primary_color,
                            font=('Arial', 11, 'bold'))
        
        # Configure button styles
        self.style.configure('TButton',
                            background=self.primary_color,
                            foreground='white',
                            borderwidth=0,
                            focuscolor='none',
                            font=('Arial', 10))
        # Button appearance changes on interaction
        self.style.map('TButton',
                      background=[('active', self.secondary_color),
                                ('pressed', '#005999')])
        
        # Configure entry widget styles
        self.style.configure('TEntry',
                            fieldbackground=self.card_color,
                            foreground=self.text_color,
                            borderwidth=1,
                            insertcolor=self.text_color)
        
        # Configure combobox (dropdown) styles
        self.style.configure('TCombobox',
                            fieldbackground=self.card_color,
                            foreground=self.text_color,
                            borderwidth=1)
        
    def create_gui(self):
        """
        Create the main GUI interface with four tabs
        
        The interface is organized into tabs:
        1. Setup & Configuration - Library installation and instructions
        2. Filing Analysis - Company search and document analysis
        3. Results & Export - View results and export options
        4. Bulk Data Processing - Process FFIEC bulk data files
        """
        # Create the main notebook widget for tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Create frames for each tab
        self.setup_tab = ttk.Frame(self.notebook)
        self.filing_tab = ttk.Frame(self.notebook)
        self.results_tab = ttk.Frame(self.notebook)
        self.bulk_data_tab = ttk.Frame(self.notebook)  # New feature for bulk data
        
        # Add tabs to notebook with descriptive names
        self.notebook.add(self.setup_tab, text="Setup & Configuration")
        self.notebook.add(self.filing_tab, text="Filing Analysis")
        self.notebook.add(self.results_tab, text="Results & Export")
        self.notebook.add(self.bulk_data_tab, text="Bulk Data Processing")
        
        # Add this line - bind tab change event
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        
        # Populate each tab with its content
        self.create_setup_tab()
        self.create_filing_tab()
        self.create_results_tab()
        self.create_bulk_data_tab()
        
        # Disable filing and results tabs if required packages are missing
        # This prevents errors from attempting to use uninstalled libraries
        if self.missing_packages:
            self.notebook.tab(1, state='disabled')
            self.notebook.tab(2, state='disabled')
    
    def create_setup_tab(self):
        """
        Create the setup and configuration tab
        
        This tab contains:
        - Application title and branding
        - Library installation status and controls
        - User guide and instructions
        """
        # Title section with white background for contrast
        title_frame = tk.Frame(self.setup_tab, bg='white', height=80)
        title_frame.pack(fill='x', padx=20, pady=(20, 10))
        title_frame.pack_propagate(False)  # Maintain fixed height
        
        # Main application title
        title_label = tk.Label(
            title_frame,
            text="🔥 FIRE Analyzer",
            font=('Arial', 24, 'bold'),
            bg='white',
            fg=self.primary_color
        )
        title_label.pack(pady=20)
        
        # Subtitle
        subtitle_label = tk.Label(
            title_frame,
            text="Financial Institution Regulatory Extractor",
            font=('Arial', 12),
            bg='white',
            fg='gray'
        )
        subtitle_label.pack()
        
        # Library Status Frame - Shows installation status
        lib_frame = ttk.LabelFrame(self.setup_tab, text="Library Status", padding=20)
        lib_frame.pack(fill='x', padx=20, pady=10)
        
        if self.missing_packages:
            # Show missing packages and install button
            status_text = f"Missing packages: {', '.join(self.missing_packages)}"
            status_color = self.error_color
            
            # Create install button
            self.install_button = tk.Button(
                lib_frame,
                text="Install Required Libraries",
                command=self.install_packages,
                bg=self.secondary_color,
                fg='white',
                font=('Arial', 12, 'bold'),
                padx=20,
                pady=10,
                cursor='hand2'
            )
            self.install_button.pack(pady=10)
        else:
            # All packages installed
            status_text = "✓ All required libraries are installed"
            status_color = self.success_color
        
        # Status label showing current state
        self.status_label = tk.Label(
            lib_frame,
            text=status_text,
            font=('Arial', 12),
            fg=status_color
        )
        self.status_label.pack(pady=5)
        
        # Progress indicators (hidden until installation starts)
        self.progress_frame = tk.Frame(lib_frame)
        self.progress_frame.pack(fill='x', pady=10)
        
        self.progress_bar = ttk.Progressbar(
            self.progress_frame,
            mode='indeterminate',
            length=400
        )
        
        self.progress_label = tk.Label(
            self.progress_frame,
            text="",
            font=('Arial', 10)
        )
        
        # User guide section
        instructions_frame = ttk.LabelFrame(self.setup_tab, text="Getting Started", padding=20)
        instructions_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Comprehensive instructions for users
        instructions = """
🔥 FIRE (Financial Institution Regulatory Extractor) - User Guide

📋 QUICK START:
1. Install Libraries: Click 'Install Required Libraries' if needed
2. Go to 'Filing Analysis' tab to begin

📊 TWO ANALYSIS METHODS:

METHOD 1: Live SEC API (Recommended for SEC Filings)
- Uses your email for SEC API authentication 
- Enter company ticker (e.g., AAPL, COF, MSFT)
- Click 'Search' then 'Analyze Filing'
- Downloads latest filings automatically from SEC.gov

METHOD 2: Local File Upload (For SEC Filings & Call Reports)
- Download filing manually first
- Use 'Local File' option in Filing Analysis tab
- Specify document type (Call Report, 10-K, etc.)
- Supported formats:
  • SEC Filings: HTML (.htm), XBRL (.xbrl), XML (.xml)
  • Call Reports: PDF (.pdf), XBRL (.xbrl), SDF (.sdf, .txt)

📁 WHERE TO GET FILES:
- SEC Filings: sec.gov/edgar/search-filings
- Call Reports: cdr.ffiec.gov/public/
  • Download options: XBRL, PDF, or SDF formats
  • All three Call Report formats are supported!

✨ KEY FEATURES:
- Automatic financial table extraction
- Excel export with original formatting preserved
- Support for 10-K, 10-Q, 8-K, and Call Reports
- MDRM dictionary integration (8,863+ codes)
- Handles XBRL, PDF, and SDF Call Report formats
- Professional-grade analysis tools
- Custom company naming for exports

📊 CALL REPORT FEATURES:
- XBRL: Best for structured data extraction
- PDF: Best for preserving visual layout with schedule detection
- SDF: Best for raw data processing
- Auto-populates blank descriptions using MDRM dictionary

⚠️ NOTE: Live API requires email authentication (already configured)
"""
        
        # Create scrollable text widget for instructions
        instructions_text = tk.Text(
            instructions_frame,
            wrap='word',
            height=12,
            font=('Arial', 10),
            bg='white'
        )
        instructions_text.pack(fill='both', expand=True)
        instructions_text.insert('1.0', instructions)
        instructions_text.config(state='disabled')  # Make read-only
    
    def create_filing_tab(self):
        """
        Create the filing analysis tab
        
        This tab contains:
        - Company search functionality
        - Document source selection (API vs local file)
        - Filing type selection
        - Analysis options
        - Progress monitoring
        """
        # Company selection frame - for searching companies
        company_frame = ttk.LabelFrame(self.filing_tab, text="Company Selection", padding=15)
        company_frame.pack(fill='x', padx=20, pady=(20, 10))
        
        # Search method selection (ticker, name, or CIK)
        tk.Label(company_frame, text="Search by:", font=('Arial', 10)).grid(row=0, column=0, sticky='w', padx=5)
        
        self.search_method = tk.StringVar(value="ticker")
        
        # Radio buttons for search method
        tk.Radiobutton(
            company_frame,
            text="Ticker Symbol",
            variable=self.search_method,
            value="ticker",
            command=self.update_search_field
        ).grid(row=0, column=1, padx=5)
        
        tk.Radiobutton(
            company_frame,
            text="Company Name",
            variable=self.search_method,
            value="name",
            command=self.update_search_field
        ).grid(row=0, column=2, padx=5)
        
        tk.Radiobutton(
            company_frame,
            text="CIK Number",
            variable=self.search_method,
            value="cik",
            command=self.update_search_field
        ).grid(row=0, column=3, padx=5)
        
        # Search input field
        tk.Label(company_frame, text="Enter Value:", font=('Arial', 10)).grid(row=1, column=0, sticky='w', padx=5, pady=10)
        
        self.search_entry = ttk.Entry(company_frame, width=30, font=('Arial', 10))
        self.search_entry.grid(row=1, column=1, columnspan=2, padx=5, pady=10)
        
        # Search button
        self.search_button = ttk.Button(
            company_frame,
            text="Search",
            command=self.search_company
        )
        self.search_button.grid(row=1, column=3, padx=5, pady=10)
        
        # Company info display (shows search results)
        self.company_info_label = tk.Label(
            company_frame,
            text="",
            font=('Arial', 10),
            fg=self.primary_color
        )
        self.company_info_label.grid(row=2, column=0, columnspan=4, pady=5)
        
        # Document selection frame
        doc_frame = ttk.LabelFrame(self.filing_tab, text="Document Selection", padding=15)
        doc_frame.pack(fill='x', padx=20, pady=10)
        
        # Document source selection (SEC API or local file)
        tk.Label(doc_frame, text="Source:", font=('Arial', 10)).grid(row=0, column=0, sticky='w', padx=5)
        
        self.doc_source = tk.StringVar(value="sec")
        
        tk.Radiobutton(
            doc_frame,
            text="SEC Filings",
            variable=self.doc_source,
            value="sec",
            command=self.update_doc_types
        ).grid(row=0, column=1, padx=5)
        
        tk.Radiobutton(
            doc_frame,
            text="Local File",
            variable=self.doc_source,
            value="local",
            command=self.update_doc_types
        ).grid(row=0, column=2, padx=5)
        
        # Filing type dropdown (for SEC API)
        tk.Label(doc_frame, text="Filing Type:", font=('Arial', 10)).grid(row=1, column=0, sticky='w', padx=5, pady=10)
        
        self.filing_type = tk.StringVar()
        self.filing_dropdown = ttk.Combobox(
            doc_frame,
            textvariable=self.filing_type,
            width=28,
            state='readonly'
        )
        self.filing_dropdown.grid(row=1, column=1, columnspan=2, padx=5, pady=10)
        self.update_doc_types()  # Initialize dropdown values
        
        # Date range selection
        tk.Label(doc_frame, text="Date Range:", font=('Arial', 10)).grid(row=2, column=0, sticky='w', padx=5)
        
        self.date_range = tk.StringVar(value="latest")
        ttk.Combobox(
            doc_frame,
            textvariable=self.date_range,
            values=["Latest", "Last 2 Years", "Last 5 Years", "All Available"],
            width=28,
            state='readonly'
        ).grid(row=2, column=1, columnspan=2, padx=5)
        
        # Local file selection frame (hidden by default)
        self.local_file_frame = tk.Frame(doc_frame)
        self.local_file_frame.grid(row=3, column=0, columnspan=4, pady=10)

        # Row 1: File selection and document type on same row
        file_row_frame = tk.Frame(self.local_file_frame)
        file_row_frame.pack(fill='x')

        # File path section
        tk.Label(
            file_row_frame,
            text="File:",
            font=('Arial', 10)
        ).pack(side='left', padx=(5, 5))

        self.file_path_var = tk.StringVar()
        self.file_entry = ttk.Entry(
            file_row_frame,
            textvariable=self.file_path_var,
            width=35
        )
        self.file_entry.pack(side='left', padx=(0, 5))

        ttk.Button(
            file_row_frame,
            text="Browse",
            command=self.browse_file
        ).pack(side='left', padx=(0, 15))

        # Document type section on same row
        tk.Label(
            file_row_frame,
            text="Document Type:",
            font=('Arial', 10)
        ).pack(side='left', padx=(10, 5))

        self.local_doc_type = tk.StringVar()
        self.doc_type_entry = ttk.Entry(
            file_row_frame,
            textvariable=self.local_doc_type,
            width=20,
            font=('Arial', 10)
        )
        self.doc_type_entry.pack(side='left')

        # Row 2: Quick select buttons
        quick_select_frame = tk.Frame(self.local_file_frame)
        quick_select_frame.pack(fill='x', pady=(5, 0))

        tk.Label(
            quick_select_frame,
            text="Quick select:",
            font=('Arial', 9),
            fg='gray'
        ).pack(side='left', padx=(5, 10))

        # Create buttons for common document types
        for doc_type in ["Call Report", "10-K", "10-Q", "8-K"]:
            tk.Button(
                quick_select_frame,
                text=doc_type,
                command=lambda dt=doc_type: self.local_doc_type.set(dt),
                bg=self.secondary_color,
                fg='white',
                font=('Arial', 9),
                padx=15,
                pady=2,
                cursor='hand2',
                relief='flat'
            ).pack(side='left', padx=2)

        # Initially hide local file frame
        self.local_file_frame.grid_remove()
        
        # Analysis options frame
        options_frame = ttk.LabelFrame(self.filing_tab, text="Analysis Options", padding=15)
        options_frame.pack(fill='x', padx=20, pady=10)
        
        # Checkbox options for analysis
        self.extract_tables = tk.BooleanVar(value=True)
        self.extract_text = tk.BooleanVar(value=False)
        self.preserve_formatting = tk.BooleanVar(value=True)
        self.create_summary = tk.BooleanVar(value=True)
        
        tk.Checkbutton(
            options_frame,
            text="Extract Financial Tables",
            variable=self.extract_tables
        ).grid(row=0, column=0, sticky='w', padx=10, pady=5)
        
        tk.Checkbutton(
            options_frame,
            text="Extract Text Sections",
            variable=self.extract_text
        ).grid(row=0, column=1, sticky='w', padx=10, pady=5)
        
        tk.Checkbutton(
            options_frame,
            text="Preserve Original Formatting",
            variable=self.preserve_formatting
        ).grid(row=1, column=0, sticky='w', padx=10, pady=5)
        
        tk.Checkbutton(
            options_frame,
            text="Create Summary Report",
            variable=self.create_summary
        ).grid(row=1, column=1, sticky='w', padx=10, pady=5)
        
        # Main analyze button
        self.analyze_button = tk.Button(
            self.filing_tab,
            text="Analyze Filing",
            command=self.analyze_filing,
            bg=self.success_color,
            fg='white',
            font=('Arial', 14, 'bold'),
            padx=30,
            pady=12,
            cursor='hand2'
        )
        self.analyze_button.pack(pady=20)
        
        # Progress display area
        self.analysis_progress = scrolledtext.ScrolledText(
            self.filing_tab,
            height=8,
            wrap='word',
            font=('Courier', 9)
        )
        self.analysis_progress.pack(fill='both', expand=True, padx=20, pady=(0, 20))
    
    def create_results_tab(self):
        """
        Create the results and export tab
        
        This tab contains:
        - Analysis summary display
        - Company identifier input for file naming
        - Export format options
        - Output location selection
        - Results preview table
        """
        # Results summary frame
        summary_frame = ttk.LabelFrame(self.results_tab, text="Analysis Summary", padding=15)
        summary_frame.pack(fill='x', padx=20, pady=(20, 10))
        
        # Text widget for summary display
        self.summary_text = tk.Text(
            summary_frame,
            height=6,
            wrap='word',
            font=('Arial', 10)
        )
        self.summary_text.pack(fill='x')
        
        # Export options frame
        export_frame = ttk.LabelFrame(self.results_tab, text="Export Options", padding=15)
        export_frame.pack(fill='x', padx=20, pady=10)

        # Company name input for file naming
        company_name_frame = tk.Frame(export_frame)
        company_name_frame.grid(row=0, column=0, columnspan=4, sticky='w', pady=(0, 10))

        tk.Label(
            company_name_frame, 
            text="Company Identifier for Filename:", 
            font=('Arial', 10, 'bold')
        ).pack(side='left', padx=5)

        self.company_identifier = tk.StringVar()
        self.company_entry = ttk.Entry(
            company_name_frame,
            textvariable=self.company_identifier,
            width=30,
            font=('Arial', 10)
        )
        self.company_entry.pack(side='left', padx=5)

        tk.Label(
            company_name_frame,
            text="(e.g., COF, Capital_One, JPM)",
            font=('Arial', 9),
            fg='gray'
        ).pack(side='left', padx=5)

        # Export format selection
        tk.Label(export_frame, text="Export Format:", font=('Arial', 10)).grid(row=1, column=0, sticky='w', padx=5)
        
        self.export_format = tk.StringVar(value="excel_formatted")
        formats = [
            ("Excel with Formatting", "excel_formatted"),
            ("Excel Basic", "excel_basic"),
            ("CSV Files", "csv"),
            ("JSON", "json"),
            ("All Formats", "all")
        ]
        
        # Create radio buttons for each format
        for i, (text, value) in enumerate(formats):
            tk.Radiobutton(
                export_frame,
                text=text,
                variable=self.export_format,
                value=value
            ).grid(row=(i//3) + 1, column=(i%3)+1, sticky='w', padx=10, pady=5)
        
        # Output folder selection
        tk.Label(export_frame, text="Output Folder:", font=('Arial', 10)).grid(row=3, column=0, sticky='w', padx=5, pady=10)
        
        # Default to user's desktop
        self.output_path_var = tk.StringVar(value=os.path.join(os.path.expanduser("~"), "Desktop", "FIRE_Analysis"))
        
        output_frame = tk.Frame(export_frame)
        output_frame.grid(row=3, column=1, columnspan=3, sticky='w', pady=10)
        
        ttk.Entry(
            output_frame,
            textvariable=self.output_path_var,
            width=50
        ).pack(side='left', padx=5)
        
        ttk.Button(
            output_frame,
            text="Browse",
            command=self.browse_output_folder
        ).pack(side='left')
        
        # Export button
        self.export_button = tk.Button(
            export_frame,
            text="Export Results",
            command=self.export_results,
            bg=self.secondary_color,
            fg='white',
            font=('Arial', 12, 'bold'),
            padx=20,
            pady=10,
            cursor='hand2',
            state='disabled'  # Disabled until analysis completes
        )
        self.export_button.grid(row=4, column=0, columnspan=4, pady=20)
        
        # Results preview frame
        preview_frame = ttk.LabelFrame(self.results_tab, text="Results Preview", padding=15)
        preview_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # Create treeview widget for table listing
        self.results_tree = ttk.Treeview(preview_frame, height=10)
        self.results_tree.pack(fill='both', expand=True)
        
        # Add scrollbars to treeview
        v_scroll = ttk.Scrollbar(preview_frame, orient='vertical', command=self.results_tree.yview)
        v_scroll.pack(side='right', fill='y')
        self.results_tree.configure(yscrollcommand=v_scroll.set)
        
        h_scroll = ttk.Scrollbar(preview_frame, orient='horizontal', command=self.results_tree.xview)
        h_scroll.pack(side='bottom', fill='x')
        self.results_tree.configure(xscrollcommand=h_scroll.set)
    
    def install_packages(self):
        """
        Initiate package installation process
        
        This method starts the installation of missing packages in a separate
        thread to avoid freezing the GUI.
        """
        # Disable install button to prevent multiple clicks
        self.install_button.config(state='disabled')
        
        # Show progress indicators
        self.progress_bar.pack()
        self.progress_bar.start()
        self.progress_label.pack(pady=5)
        
        # Run installation in separate thread
        thread = threading.Thread(target=self._install_packages_thread)
        thread.daemon = True  # Thread will close when main program exits
        thread.start()
    
    def _install_packages_thread(self):
        """
        Thread function to install missing packages using pip
        
        This runs in the background to install each missing package
        and sends status updates to the main GUI thread via the queue.
        """
        try:
            for package in self.missing_packages:
                # Send progress update
                self.queue.put(('progress', f"Installing {package}..."))
                
                # Run pip install command
                result = subprocess.run(
                    [sys.executable, "-m", "pip", "install", package],
                    capture_output=True,
                    text=True
                )
                
                if result.returncode != 0:
                    # Installation failed
                    self.queue.put(('error', f"Failed to install {package}: {result.stderr}"))
                    return
            
            # All packages installed successfully
            self.queue.put(('success', "All packages installed successfully!"))
            
            # Re-check imports
            global imported_libs
            missing = check_and_import_libraries()
            
            if not missing:
                # Enable other tabs if all packages now available
                self.queue.put(('enable_tabs', None))
                
        except Exception as e:
            self.queue.put(('error', f"Installation error: {str(e)}"))
    
    def process_queue(self):
        """
        Process messages from background threads
        
        This method runs periodically to check for messages from background
        threads and update the GUI accordingly. It handles various message
        types for different operations.
        
        CRITICAL: This method must have proper try/except alignment!
        """
        try:
            while True:
                # Get message from queue (non-blocking)
                msg_type, msg_data = self.queue.get_nowait()
                
                if msg_type == 'progress':
                    # Update progress label and analysis progress
                    if hasattr(self, 'progress_label'):
                        self.progress_label.config(text=msg_data)
                    self.update_analysis_progress(msg_data)
                    
                    # Update bulk progress display
                    if hasattr(self, 'bulk_progress'):
                        self.update_bulk_progress(msg_data)
                        
                        # Parse progress data for summary panel updates
                        if hasattr(self, 'progress_summary_frame'):
                            # Show progress summary frame during processing
                            if not self.progress_summary_frame.winfo_viewable():
                                self.progress_summary_frame.pack(fill='x', padx=20, pady=10, before=self.bulk_process_button)
                            
                            # Update progress statistics
                            if "Schedule" in msg_data and ":" in msg_data:
                                # Extract schedule code
                                schedule_match = re.search(r'Schedule (\w+)', msg_data)
                                if schedule_match:
                                    self.current_schedule_label.config(text=schedule_match.group(1))
                            
                            # Update file progress
                            file_match = re.search(r'File (\d+)/(\d+)', msg_data)
                            if file_match:
                                current = int(file_match.group(1))
                                total = int(file_match.group(2))
                                self.files_progress_label.config(text=f"{current} / {total}")
                                
                                # Update progress bar
                                if total > 0:
                                    progress_percent = (current / total) * 100
                                    self.bulk_progress_bar['value'] = progress_percent
                                    self.overall_progress_label.config(text=f"{progress_percent:.1f}%")
                            
                            # Update completed schedules count
                            if "✓ Completed" in msg_data:
                                current_text = self.schedules_completed_label.cget("text")
                                try:
                                    current_count = int(current_text)
                                    self.schedules_completed_label.config(text=str(current_count + 1))
                                except ValueError:
                                    self.schedules_completed_label.config(text="1")
                
                elif msg_type == 'success':
                    # Installation completed successfully
                    self.progress_bar.stop()
                    self.progress_bar.pack_forget()
                    self.progress_label.config(text=msg_data, fg=self.success_color)
                    self.status_label.config(
                        text="✓ All required libraries are installed",
                        fg=self.success_color
                    )
                    self.install_button.pack_forget()
                
                elif msg_type == 'error':
                    # Handle errors
                    if hasattr(self, 'progress_bar'):
                        self.progress_bar.stop()
                        self.progress_bar.pack_forget()
                    messagebox.showerror("Error", msg_data)
                    if hasattr(self, 'install_button'):
                        self.install_button.config(state='normal')
                    # Also re-enable bulk process button if it was a bulk processing error
                    if hasattr(self, 'bulk_process_button'):
                        self.bulk_process_button.config(state='normal', text="Process Bulk Data")
                
                elif msg_type == 'enable_tabs':
                    # Enable filing and results tabs
                    self.notebook.tab(1, state='normal')
                    self.notebook.tab(2, state='normal')
                    self.notebook.select(1)  # Switch to filing tab
                
                elif msg_type == 'analysis_complete':
                    # Analysis finished successfully
                    self.analyze_button.config(state='normal', text="Analyze Filing")
                    self.export_button.config(state='normal')
                    self.notebook.select(2)  # Switch to results tab
                    self.display_results(msg_data)
                
                elif msg_type == 'bulk_complete':
                    # Bulk processing completed
                    self.bulk_process_button.config(state='normal', text="Process Bulk Data")
                    
                    # Update progress to 100%
                    if hasattr(self, 'bulk_progress_bar'):
                        self.bulk_progress_bar['value'] = 100
                        self.overall_progress_label.config(text="100%")
                    
                    #REPLACE
                    # Build completion message based on processing mode
                    if 'institutions' in msg_data:
                        # Multi-institution mode
                        completion_msg = (
                            f"✅ Multi-Institution Comparison Complete!\n\n" +
                            f"Institutions Compared: {msg_data['institutions']}\n" +
                            f"Schedules per Institution: {msg_data['schedules']}\n" +
                            f"Total Data Rows: {msg_data['total_rows']:,}\n\n" +
                            f"Output: {msg_data['output_path']}"
                        )
                    else:
                        # Single institution mode
                        completion_msg = (
                            f"✅ Bulk data processed successfully!\n\n" +
                            f"Schedules: {msg_data['schedules']}\n" +
                            f"Total Rows: {msg_data['total_rows']:,}\n\n" +
                            f"Output: {msg_data['output_path']}"
                        )

                    result = messagebox.showinfo(
                        "Processing Complete",
                        completion_msg,
                        type='okcancel'
                    )
                    
                    # Hide progress summary after completion
                    if hasattr(self, 'progress_summary_frame'):
                        self.progress_summary_frame.pack_forget()
                        
                        # Reset progress labels
                        self.current_schedule_label.config(text="Not started")
                        self.files_progress_label.config(text="0 / 0")
                        self.schedules_completed_label.config(text="0")
                        self.overall_progress_label.config(text="0%")
                        self.bulk_progress_bar['value'] = 0
                    
                    if result == 'ok':
                        # Open output file based on operating system
                        if sys.platform == 'win32':
                            os.startfile(msg_data['output_path'])
                        elif sys.platform == 'darwin':
                            subprocess.run(['open', msg_data['output_path']])
                        else:
                            subprocess.run(['xdg-open', msg_data['output_path']])
         
                elif msg_type == 'show_resume_button':
                    # Show resume button
                    if hasattr(self, 'bulk_resume_button'):
                        self.bulk_resume_button.pack(side='left', padx=5)
                        
                #updated
                elif msg_type == 'show_retry_button':
                    # Show retry button
                    if hasattr(self, 'bulk_retry_button'):
                        self.bulk_retry_button.pack(side='left', padx=5)
                            
        except queue.Empty:
            # No messages to process - this is normal
            pass
        
        # Schedule next check after 100ms
        self.root.after(100, self.process_queue)
    
    def update_search_field(self):
        """
        Update search field placeholder based on selected search method
        
        This provides user guidance on the expected input format.
        """
        method = self.search_method.get()
        placeholders = {
            'ticker': 'e.g., AAPL, MSFT, JPM',
            'name': 'e.g., Apple Inc., Microsoft',
            'cik': 'e.g., 0000320193'
        }
        # TODO: Implement actual placeholder update
        # Tkinter Entry doesn't support placeholders natively
    
    def update_doc_types(self):
        """
        Update document type options based on selected source
        
        Shows different options for SEC API vs local file upload.
        """
        source = self.doc_source.get()
        
        if source == 'sec':
            # SEC filing types available via API
            types = ['10-K', '10-Q', '8-K', 'DEF 14A', '20-F', 'All Types']
            self.filing_dropdown['values'] = types
            self.filing_type.set('10-K')
            # Hide local file selection
            if hasattr(self, 'local_file_frame'):
                self.local_file_frame.grid_remove()
    
        elif source == 'local':
            # Clear dropdown for local files
            self.filing_dropdown['values'] = []
            self.filing_type.set('')
            # Show local file selection
            if hasattr(self, 'local_file_frame'):
                self.local_file_frame.grid()
    
    def browse_file(self):
        """
        Open file dialog for selecting local filing document
        
        Supports various file types for SEC filings and Call Reports.
        """
        filename = filedialog.askopenfilename(
            title="Select SEC Filing or Call Report",
            filetypes=[
                ("SEC Filing files", "*.htm *.html"),
                ("XBRL files", "*.xbrl *.xml"),
                ("Call Report PDF", "*.pdf"),
                ("Call Report SDF", "*.sdf *.txt"),
                ("All supported files", "*.htm *.html *.xml *.xbrl *.pdf *.sdf *.txt"),
                ("All files", "*.*")
            ]
        )
        if filename:
            self.file_path_var.set(filename)
    
    def browse_output_folder(self):
        """
        Open folder dialog for selecting export destination
        """
        folder = filedialog.askdirectory(title="Select Output Folder")
        if folder:
            self.output_path_var.set(folder)
    
    def search_company(self):
        """
        Search for company information using SEC EDGAR API
        
        This is a placeholder that would connect to the actual SEC API
        to search for companies by ticker, name, or CIK.
        """
        search_value = self.search_entry.get().strip()
        if not search_value:
            messagebox.showwarning("Input Required", "Please enter a search value")
            return
        
        # TODO: Implement actual SEC EDGAR API search
        # For now, show placeholder result
        self.company_info_label.config(
            text=f"Found: {search_value.upper()} - Example Corporation\nCIK: 0000000000 | Exchange: NYSE"
        )
    
    def update_analysis_progress(self, message):
        """
        Update the analysis progress display with timestamped message
        
        Args:
            message: Progress message to display
        """
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.analysis_progress.insert('end', f"{timestamp} - {message}\n")
        self.analysis_progress.see('end')  # Auto-scroll to bottom
    
    def analyze_filing(self):
        """
        Start the filing analysis process
        
        Validates inputs and starts analysis in a background thread.
        """
        # Validate inputs based on source
        if self.doc_source.get() == 'local':
            if not self.file_path_var.get():
                messagebox.showwarning("Input Required", "Please select a file")
                return
            # Check if document type is specified
            if not self.local_doc_type.get():
                messagebox.showwarning("Input Required", "Please specify the document type")
                return
        else:
            if not self.search_entry.get():
                messagebox.showwarning("Input Required", "Please search for a company first")
                return
        
        # Clear previous progress
        self.analysis_progress.delete('1.0', 'end')
        
        # Disable button and update text
        self.analyze_button.config(state='disabled', text="Analyzing...")
        
        # Run analysis in separate thread
        thread = threading.Thread(target=self._analyze_filing_thread)
        thread.daemon = True
        thread.start()
    
    def _analyze_filing_thread(self):
        """
        Thread function for filing analysis
        
        This performs the actual analysis using the enhanced_scraper module
        and sends progress updates to the GUI.
        """
        try:
            # Import the scraper module
            from enhanced_scraper import EnhancedFIREScraper    
            
            self.queue.put(('progress', 'Starting analysis...'))
            
            # Create scraper instance based on source
            if self.doc_source.get() == 'local':
                # Local file analysis
                # Get document type from user input
                doc_type = self.local_doc_type.get() or "Financial Document"
                
                # Create company info with document type
                company_info = {
                    'name': 'Local Company',
                    'ticker': 'LOCAL',
                    'cik': 'N/A',
                    'doc_type': doc_type
                }
                
                scraper = EnhancedFIREScraper(
                    company_info=company_info,
                    local_file_path=self.file_path_var.get()
                )
                
                # Store document type in metadata
                scraper.metadata['form_type'] = doc_type
            else:
                # SEC EDGAR API analysis
                self.queue.put(('progress', 'Fetching filing from SEC EDGAR...'))

                # Create scraper
                scraper = EnhancedFIREScraper()

                # Set company by ticker
                search_value = self.search_entry.get().strip()
                if not scraper.set_company(ticker=search_value):
                    self.queue.put(('error', f'Could not find company with ticker: {search_value}'))
                    return

                # Get filing URL
                filing_url = scraper.get_filing_url(self.filing_type.get())
                if not filing_url:
                    self.queue.put(('error', 'Could not find the requested filing'))
                    return

                scraper.filing_url = filing_url
            
            self.queue.put(('progress', 'Loading filing document...'))
            
            # Perform table extraction
            if scraper.scrape_all_tables():
                self.queue.put(('progress', f'Found {len(scraper.tables)} financial tables'))
                
                # Package results
                results = {
                    'tables': scraper.tables,
                    'summary': f"Extracted {len(scraper.tables)} tables",
                    'company': self.search_entry.get() or 'Local File',
                    'doc_type': scraper.metadata.get('form_type', 'Unknown')
                }
                
                # Store the scraper for export
                self.current_scraper = scraper
                
                self.queue.put(('analysis_complete', results))
            else:
                self.queue.put(('error', 'Failed to extract tables from filing'))
                
        except Exception as e:
            self.queue.put(('error', f'Analysis error: {str(e)}'))
            # Re-enable analyze button
            self.analyze_button.config(state='normal', text="Analyze Filing")
    
    def display_results(self, results):
        """
        Display analysis results in the results tab
        
        Args:
            results: Dictionary containing analysis results
        """
        # Store results for export
        self.current_results = results
        
        # Pre-populate company identifier if available
        if 'company' in results and results['company'] != 'Local File':
            self.company_identifier.set(results['company'])
        
        # Update summary text
        doc_type = results.get('doc_type', 'Document')
        summary = f"""
Analysis Complete!

Company: {results['company']}
Document Type: {doc_type}
Tables Found: {len(results['tables'])}

Summary:
{results['summary']}
"""
        self.summary_text.delete('1.0', 'end')
        self.summary_text.insert('1.0', summary)
        
        # Clear and update results tree
        self.results_tree.delete(*self.results_tree.get_children())
        
        # Configure columns
        self.results_tree['columns'] = ('Type', 'Rows', 'Columns', 'Section')
        self.results_tree.heading('#0', text='Table Name')
        self.results_tree.heading('Type', text='Type')
        self.results_tree.heading('Rows', text='Rows')
        self.results_tree.heading('Columns', text='Columns')
        self.results_tree.heading('Section', text='Section')
        
        # Add table information to tree
        for i, table in enumerate(results['tables']):
            data = table['data']['data']
            rows = len(data)
            cols = len(data[0]) if data else 0
            
            self.results_tree.insert(
                '',
                'end',
                text=table['name'],
                values=('Financial', rows, cols, table['section'])
            )
    
    def export_results(self):
        """
        Export results to selected format
        
        Shows progress window and handles the export process.
        """
        # Validate company identifier
        if not self.company_identifier.get().strip():
            messagebox.showwarning("Input Required", 
                                 "Please enter a company identifier for the filename")
            return
        
        export_format = self.export_format.get()
        output_path = self.output_path_var.get()
        
        # Create output directory
        os.makedirs(output_path, exist_ok=True)
        
        # Show export progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Exporting Results")
        progress_window.geometry("400x150")
        
        tk.Label(
            progress_window,
            text="Exporting results...",
            font=('Arial', 12)
        ).pack(pady=20)
        
        export_progress = ttk.Progressbar(
            progress_window,
            mode='indeterminate',
            length=300
        )
        export_progress.pack(pady=10)
        export_progress.start()
        
        # Perform export after short delay (for visual feedback)
        self.root.after(500, lambda: self._complete_export(progress_window, output_path))
    
    def _complete_export(self, progress_window, output_path):
        """
        Complete the export process
        
        Args:
            progress_window: Progress dialog to close
            output_path: Directory to save files
        """
        progress_window.destroy()

        # Perform actual export
        try:
            # Check if we have results to export
            if hasattr(self, 'current_results'):
                tables = self.current_results.get('tables', [])
                
                # Import the scraper module
                from enhanced_scraper import EnhancedFIREScraper
                
                # Use existing scraper if available, or create new one
                if hasattr(self, 'current_scraper'):
                    scraper = self.current_scraper
                else:
                    # Create a new scraper instance with the tables
                    scraper = EnhancedFIREScraper()
                    scraper.tables = tables
                
                # Update company info with user-specified identifier
                company_id = self.company_identifier.get().strip()
                company_id = company_id.replace(' ', '_').replace('/', '_').replace('\\', '_')
                
                scraper.company_info = {
                    'name': company_id,
                    'ticker': company_id,
                    'cik': 'N/A'
                }
                
                # Update metadata
                scraper.metadata.update({
                    'company': company_id,
                    'ticker': company_id,
                    'form_type': self.current_results.get('doc_type', '10-K')
                })
                
                # Generate filenames with company identifier
                date_str = datetime.now().strftime('%Y%m%d')
                base_filename = f"{company_id}_financial_tables_{date_str}"
                
                # Export based on selected format
                export_format = self.export_format.get()
                
                if export_format == 'excel_formatted':
                    output_file = os.path.join(output_path, f'{base_filename}_formatted.xlsx')
                    scraper.save_to_excel_formatted(output_file)
                elif export_format == 'excel_basic':
                    output_file = os.path.join(output_path, f'{base_filename}_basic.xlsx')
                    scraper.save_to_excel_basic(output_file)
                elif export_format == 'csv':
                    csv_dir = os.path.join(output_path, f'{company_id}_csv_tables_{date_str}')
                    scraper.save_to_csv(csv_dir)
                elif export_format == 'json':
                    output_file = os.path.join(output_path, f'{company_id}_financial_data_{date_str}.json')
                    scraper.save_to_json(output_file)
                elif export_format == 'all':
                    # Export all formats
                    scraper.save_to_excel_formatted(os.path.join(output_path, f'{base_filename}_formatted.xlsx'))
                    scraper.save_to_excel_basic(os.path.join(output_path, f'{base_filename}_basic.xlsx'))
                    scraper.save_to_csv(os.path.join(output_path, f'{company_id}_csv_tables_{date_str}'))
                    scraper.save_to_json(os.path.join(output_path, f'{company_id}_financial_data_{date_str}.json'))
                    
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export: {str(e)}")
            return
        
        # Show success message
        message = f"Results exported successfully!\n\nFiles saved to:\n{output_path}"
        result = messagebox.showinfo("Export Complete", message, type='okcancel')
        
        if result == 'ok':
            # Open output folder in file explorer
            if sys.platform == 'win32':
                os.startfile(output_path)
            elif sys.platform == 'darwin':
                subprocess.run(['open', output_path])
            else:
                subprocess.run(['xdg-open', output_path])

    # ===== BULK DATA PROCESSING METHODS =====
    # The following methods handle bulk FFIEC data processing functionality
    def create_bulk_data_tab(self):
        """
        Create the Bulk Data processing tab with scrollable content
        
        This tab allows users to process FFIEC bulk data downloads:
        - Tab-delimited text files from FFIEC
        - Convert to standardized Excel format
        - Add MDRM code descriptions
        - Filter by institution
        """
        
        # Create scrollable container
        container = ttk.Frame(self.bulk_data_tab)
        container.pack(fill='both', expand=True)
        
        # Create canvas and scrollbar
        canvas = tk.Canvas(container, bg=self.bg_color)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        # Configure scrolling
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Enable mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Use scrollable_frame as parent for all content
        parent_frame = scrollable_frame
        
        # ===== Quick Start / Auto-detection section =====
        detection_frame = ttk.LabelFrame(parent_frame, text="Quick Start", padding=15)
        detection_frame.pack(fill='x', padx=20, pady=(20, 10))
        
        # Check for existing data
        found_data = self.check_for_bulk_data()
        
        if found_data:
            tk.Label(
                detection_frame,
                text=f"✓ Found {len(found_data)} potential FFIEC data folder(s):",
                font=('Arial', 10, 'bold'),
                fg=self.success_color
            ).pack(anchor='w')
            
            # Dropdown to select found data
            self.found_data_var = tk.StringVar()
            found_dropdown = ttk.Combobox(
                detection_frame,
                textvariable=self.found_data_var,
                values=found_data,
                width=60,
                state='readonly'
            )
            found_dropdown.pack(fill='x', pady=5)
            if found_data:
                found_dropdown.set(found_data[0])
            
            ttk.Button(
                detection_frame,
                text="Use Selected Folder",
                command=lambda: self.bulk_dir_path.set(self.found_data_var.get())
            ).pack(pady=5)
            
        else:
            # No data found - show download prompt
            tk.Label(
                detection_frame,
                text="⚠️ No FFIEC bulk data found in common locations",
                font=('Arial', 10, 'bold'),
                fg=self.error_color
            ).pack(anchor='w')
            
            tk.Label(
                detection_frame,
                text="You need to download FFIEC bulk data first:",
                font=('Arial', 10)
            ).pack(anchor='w', pady=(5, 10))
            
            # Download guide button
            download_frame = tk.Frame(detection_frame)
            download_frame.pack()
            
            tk.Button(
                download_frame,
                text="📖 View Download Guide",
                command=self.show_download_guide,
                bg=self.primary_color,
                fg='white',
                font=('Arial', 11, 'bold'),
                padx=20,
                pady=8,
                cursor='hand2'
            ).pack(side='left', padx=5)
            
            tk.Button(
                download_frame,
                text="🌐 Open FFIEC Website",
                command=lambda: webbrowser.open("https://cdr.ffiec.gov/public/PWS/DownloadBulkData.aspx"),
                bg=self.secondary_color,
                fg='white',
                font=('Arial', 11, 'bold'),
                padx=20,
                pady=8,
                cursor='hand2'
            ).pack(side='left', padx=5)
        
        # Bulk Data Overview Frame
        overview_frame = ttk.LabelFrame(parent_frame, text="FFIEC Bulk Data Processor", padding=15)
        overview_frame.pack(fill='x', padx=20, pady=(20, 10))
        
        # Description
        desc_text = """📊 FFIEC Bulk Data Downloads provide structured Call Report data for all financial institutions.
        
    This feature processes tab-delimited text files from FFIEC's bulk download service and converts them to a 
    standardized 6-column Excel format with proper MDRM code descriptions and hierarchical line item numbering.

    Key Features:
    - Process individual files or entire quarterly datasets (47 files)
    - Filter by specific institution (RSSD ID)
    - Automatic MDRM code description lookup
    - Hierarchical line item numbering (RC.1, RC.2, RI.1, etc.)
    - Professional Excel output with formatting
    - Resume interrupted processing
    - Retry failed files
        """
        
        desc_label = tk.Label(
            overview_frame,
            text=desc_text,
            font=('Arial', 10),
            justify='left',
            wraplength=700
        )
        desc_label.pack(padx=10, pady=10)
        
        # Input Selection Frame
        input_frame = ttk.LabelFrame(parent_frame, text="Data Source", padding=15)
        input_frame.pack(fill='x', padx=20, pady=10)
        
        # Processing mode selection
        tk.Label(input_frame, text="Processing Mode:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=5)
        
        self.bulk_mode = tk.StringVar(value="single")
        
        tk.Radiobutton(
            input_frame,
            text="Single File",
            variable=self.bulk_mode,
            value="single",
            command=self.update_bulk_input_fields
        ).grid(row=0, column=1, padx=5)
        
        tk.Radiobutton(
            input_frame,
            text="Directory (Multiple Files)",
            variable=self.bulk_mode,
            value="directory",
            command=self.update_bulk_input_fields
        ).grid(row=0, column=2, padx=5)
        
        # File/Directory selection
        self.bulk_file_frame = tk.Frame(input_frame)
        self.bulk_file_frame.grid(row=1, column=0, columnspan=4, pady=10, sticky='ew')
        
        tk.Label(self.bulk_file_frame, text="Select File:", font=('Arial', 10)).pack(side='left', padx=5)
        
        self.bulk_file_path = tk.StringVar()
        self.bulk_file_entry = ttk.Entry(
            self.bulk_file_frame,
            textvariable=self.bulk_file_path,
            width=50
        )
        self.bulk_file_entry.pack(side='left', padx=5)
        
        ttk.Button(
            self.bulk_file_frame,
            text="Browse",
            command=self.browse_bulk_file
        ).pack(side='left')
        
        # Directory selection frame (hidden by default)
        self.bulk_dir_frame = tk.Frame(input_frame)
        
        tk.Label(self.bulk_dir_frame, text="Select Directory:", font=('Arial', 10)).pack(side='left', padx=5)
        
        self.bulk_dir_path = tk.StringVar()
        self.bulk_dir_entry = ttk.Entry(
            self.bulk_dir_frame,
            textvariable=self.bulk_dir_path,
            width=50
        )
        self.bulk_dir_entry.pack(side='left', padx=5)
        
        ttk.Button(
            self.bulk_dir_frame,
            text="Browse",
            command=self.browse_bulk_directory
        ).pack(side='left')

        # Quarter Selection Frame
        quarter_frame = ttk.LabelFrame(parent_frame, text="Quarter Selection", padding=15)
        quarter_frame.pack(fill='x', padx=20, pady=10)

        # Quarter dropdown
        quarter_select_frame = tk.Frame(quarter_frame)
        quarter_select_frame.pack(fill='x', pady=5)

        tk.Label(
            quarter_select_frame,
            text="Select Quarter:",
            font=('Arial', 10, 'bold')
        ).pack(side='left', padx=5)

        self.selected_quarter = tk.StringVar()
        self.quarter_dropdown = ttk.Combobox(
            quarter_select_frame,
            textvariable=self.selected_quarter,
            width=20,
            state='readonly'
        )
        self.quarter_dropdown.pack(side='left', padx=5)
        
        # Add this line - bind quarter selection change
        self.quarter_dropdown.bind('<<ComboboxSelected>>', lambda e: self.update_quarter_info())

        ttk.Button(
            quarter_select_frame,
            text="Scan for Quarters",
            command=self.refresh_quarters
        ).pack(side='left', padx=10)

        # Quarter info display
        self.quarter_info_frame = tk.Frame(quarter_frame)
        self.quarter_info_frame.pack(fill='x', pady=10)

        self.quarter_info_label = tk.Label(
            self.quarter_info_frame,
            text="Select a directory and click 'Scan for Quarters' to see available data",
            font=('Arial', 9),
            justify='left',
            fg='gray'
        )
        self.quarter_info_label.pack()
        
        # Filter Options Frame
        filter_frame = ttk.LabelFrame(parent_frame, text="Filter Options", padding=15)
        filter_frame.pack(fill='x', padx=20, pady=10)
        
       # Institution mode selection
        mode_frame = tk.Frame(filter_frame)
        mode_frame.pack(fill='x', pady=(5, 10))

        tk.Label(
            mode_frame,
            text="Processing Mode:",
            font=('Arial', 10, 'bold')
        ).pack(side='left', padx=5)

        self.institution_mode = tk.StringVar(value="single")

        tk.Radiobutton(
            mode_frame,
            text="Single Institution (with RSSD filter)",
            variable=self.institution_mode,
            value="single",
            command=self.toggle_mode_interface
        ).pack(side='left', padx=10)

        tk.Radiobutton(
            mode_frame,
            text="Multi-Institution Comparison (2-4 institutions)",
            variable=self.institution_mode,
            value="multi",
            command=self.toggle_mode_interface
        ).pack(side='left', padx=10)
        
        # Institution filter
        inst_filter_frame = tk.Frame(filter_frame)
        inst_filter_frame.pack(fill='x', pady=5)

        tk.Label(
            inst_filter_frame,
            text="Filter by Institution (Optional):",
            font=('Arial', 10, 'bold')
        ).pack(side='left', padx=5)

        tk.Label(
            inst_filter_frame,
            text="RSSD ID:",
            font=('Arial', 10)
        ).pack(side='left', padx=(20, 5))

        self.bulk_rssd_id = tk.StringVar()
        self.bulk_rssd_entry = ttk.Entry(
            inst_filter_frame,
            textvariable=self.bulk_rssd_id,
            width=15
        )
        self.bulk_rssd_entry.pack(side='left', padx=5)

        tk.Label(
            inst_filter_frame,
            text="OR",
            font=('Arial', 10, 'italic'),
            fg='gray'
        ).pack(side='left', padx=10)

        # Institution name for output
        name_frame = tk.Frame(filter_frame)
        name_frame.pack(fill='x', pady=5)

        tk.Label(
            name_frame,
            text="Institution Name:",
            font=('Arial', 10)
        ).pack(side='left', padx=5)

        self.bulk_inst_name = tk.StringVar()
        self.bulk_inst_name_entry = ttk.Entry(
            name_frame,
            textvariable=self.bulk_inst_name,
            width=40
        )
        self.bulk_inst_name_entry.pack(side='left', padx=5)

        # Add name search button
        ttk.Button(
            name_frame,
            text="Find RSSD",
            command=self.find_rssd_by_name
        ).pack(side='left', padx=5)

        tk.Label(
            name_frame,
            text="(partial name ok, case-insensitive)",
            font=('Arial', 9),
            fg='gray'
        ).pack(side='left', padx=5)

        # Help text
        help_frame = tk.Frame(filter_frame)
        help_frame.pack(fill='x', pady=(5, 10))

        tk.Label(
            help_frame,
            text="💡 Enter either RSSD ID or Institution Name. If name matches multiple institutions, you'll be prompted to select.",
            font=('Arial', 9),
            fg='blue',
            wraplength=600,
            justify='left'
        ).pack(padx=5)
        
        #Text Prompt to User
        # Multi-mode disabled notice (initially hidden)
        self.multi_mode_notice_frame = tk.Frame(filter_frame)
        self.multi_mode_notice_label = tk.Label(
            self.multi_mode_notice_frame,
            text="ℹ️ These fields are disabled in Multi-Institution Comparison mode. Use the selection interface below.",
            font=('Arial', 9, 'italic'),
            fg='gray',
            wraplength=600,
            justify='left'
        )
        self.multi_mode_notice_label.pack(padx=5)
        
        # Don't pack the frame initially - it will be shown/hidden by toggle_mode_interface
        # Multi-Institution Selection Frame (initially hidden)
        self.multi_inst_frame = tk.Frame(filter_frame)
        # Don't pack it yet - it will be shown when multi mode is selected

        # Primary institution
        primary_frame = ttk.LabelFrame(self.multi_inst_frame, text="Primary Institution", padding=10)
        primary_frame.pack(fill='x', pady=(0, 10))

        primary_input_frame = tk.Frame(primary_frame)
        primary_input_frame.pack(fill='x')

        tk.Label(primary_input_frame, text="RSSD ID or Name:", font=('Arial', 10)).pack(side='left', padx=5)
        self.primary_inst_var = tk.StringVar()
        self.primary_inst_entry = ttk.Entry(primary_input_frame, textvariable=self.primary_inst_var, width=30)
        self.primary_inst_entry.pack(side='left', padx=5)

        ttk.Button(
            primary_input_frame,
            text="Search",
            command=lambda: self.search_institution('primary')
        ).pack(side='left', padx=5)

        self.primary_inst_label = tk.Label(primary_frame, text="", font=('Arial', 9), fg='gray')
        self.primary_inst_label.pack(anchor='w', padx=5, pady=(5, 0))

        # Peer institutions
        peers_frame = ttk.LabelFrame(self.multi_inst_frame, text="Peer Institutions (1-3 additional)", padding=10)
        peers_frame.pack(fill='x', pady=(0, 10))

        # Add peer controls
        add_peer_frame = tk.Frame(peers_frame)
        add_peer_frame.pack(fill='x', pady=(0, 10))

        tk.Label(add_peer_frame, text="Add Peer:", font=('Arial', 10)).pack(side='left', padx=5)
        self.peer_search_var = tk.StringVar()
        self.peer_search_entry = ttk.Entry(add_peer_frame, textvariable=self.peer_search_var, width=30)
        self.peer_search_entry.pack(side='left', padx=5)

        ttk.Button(
            add_peer_frame,
            text="Search & Add",
            command=self.add_peer_institution
        ).pack(side='left', padx=5)

        ttk.Button(
            add_peer_frame,
            text="Suggest Similar Banks",
            command=self.suggest_similar_banks
        ).pack(side='left', padx=10)

        # Selected peers list
        tk.Label(peers_frame, text="Selected Peers:", font=('Arial', 10, 'bold')).pack(anchor='w', padx=5)

        # Create a frame for the peer list
        self.peers_list_frame = tk.Frame(peers_frame)
        self.peers_list_frame.pack(fill='both', expand=True, padx=5, pady=5)

        # Create scrollable listbox for peers
        list_container = tk.Frame(self.peers_list_frame)
        list_container.pack(fill='both', expand=True)

        scrollbar = ttk.Scrollbar(list_container)
        scrollbar.pack(side='right', fill='y')

        self.peers_listbox = tk.Listbox(
            list_container,
            height=4,
            yscrollcommand=scrollbar.set,
            font=('Arial', 9)
        )
        self.peers_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.peers_listbox.yview)

        # Remove button
        ttk.Button(
            self.peers_list_frame,
            text="Remove Selected",
            command=self.remove_selected_peer
        ).pack(pady=(5, 0))

        # Store peer data
        self.selected_peers = []

        # Multi-institution validation warning
        self.multi_warning_label = tk.Label(
            self.multi_inst_frame,
            text="",
            font=('Arial', 9),
            fg='orange',
            wraplength=600
        )
        self.multi_warning_label.pack(fill='x', padx=5, pady=(5, 0))
        
        # Dictionary Options Frame
        dict_frame = ttk.LabelFrame(parent_frame, text="MDRM Dictionary", padding=15)
        dict_frame.pack(fill='x', padx=20, pady=10)
        
        # Dictionary status
        self.dict_status_label = tk.Label(
            dict_frame,
            text="Dictionary Status: Not Loaded",
            font=('Arial', 10),
            fg=self.error_color
        )
        self.dict_status_label.pack(pady=5)
        
        # Dictionary path
        dict_path_frame = tk.Frame(dict_frame)
        dict_path_frame.pack(fill='x')
        
        tk.Label(
            dict_path_frame,
            text="Dictionary File:",
            font=('Arial', 10)
        ).pack(side='left', padx=5)
        
        self.bulk_dict_path = tk.StringVar(
            value=os.path.join(os.path.dirname(__file__), "dictionaries", "call_report_mdrm_dictionary.json")
        )
        ttk.Entry(
            dict_path_frame,
            textvariable=self.bulk_dict_path,
            width=50
        ).pack(side='left', padx=5)
        
        ttk.Button(
            dict_path_frame,
            text="Browse",
            command=self.browse_dictionary_file
        ).pack(side='left', padx=5)
        
        ttk.Button(
            dict_path_frame,
            text="Load Dictionary",
            command=self.load_mdrm_dictionary
        ).pack(side='left', padx=10)
        
        # Progress Summary Frame (initially hidden)
        progress_summary_frame = ttk.LabelFrame(parent_frame, text="Processing Status", padding=10)
        progress_summary_frame.pack(fill='x', padx=20, pady=10)

        # Create a grid layout for progress stats
        stats_frame = tk.Frame(progress_summary_frame)
        stats_frame.pack(fill='x')

        # Current Schedule
        tk.Label(stats_frame, text="Current Schedule:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=5)
        self.current_schedule_label = tk.Label(stats_frame, text="Not started", font=('Arial', 10))
        self.current_schedule_label.grid(row=0, column=1, sticky='w', padx=5)

        # Files Progress
        tk.Label(stats_frame, text="Files Progress:", font=('Arial', 10, 'bold')).grid(row=0, column=2, sticky='w', padx=20)
        self.files_progress_label = tk.Label(stats_frame, text="0 / 0", font=('Arial', 10))
        self.files_progress_label.grid(row=0, column=3, sticky='w', padx=5)

        # Schedules Completed
        tk.Label(stats_frame, text="Schedules Completed:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=5)
        self.schedules_completed_label = tk.Label(stats_frame, text="0", font=('Arial', 10))
        self.schedules_completed_label.grid(row=1, column=1, sticky='w', padx=5)

        # Overall Progress
        tk.Label(stats_frame, text="Overall Progress:", font=('Arial', 10, 'bold')).grid(row=1, column=2, sticky='w', padx=20)
        self.overall_progress_label = tk.Label(stats_frame, text="0%", font=('Arial', 10))
        self.overall_progress_label.grid(row=1, column=3, sticky='w', padx=5)

        # Progress bar
        self.bulk_progress_bar = ttk.Progressbar(
            progress_summary_frame,
            mode='determinate',
            length=400
        )
        self.bulk_progress_bar.pack(pady=10, fill='x')

        # Initially hide the progress summary (show only during processing)
        progress_summary_frame.pack_forget()
        self.progress_summary_frame = progress_summary_frame
        
        # Processing History Frame
        history_frame = ttk.LabelFrame(parent_frame, text="Processing History", padding=15)
        history_frame.pack(fill='x', padx=20, pady=10)

        # Status text display
        self.processing_status_text = tk.Text(
            history_frame,
            height=4,
            wrap='word',
            font=('Arial', 10),
            bg='white',
            state='disabled'  # Make it read-only
        )
        self.processing_status_text.pack(fill='x', padx=5, pady=5)

        # Clear history button
        clear_history_frame = tk.Frame(history_frame)
        clear_history_frame.pack(fill='x', pady=5)

        ttk.Button(
            clear_history_frame,
            text="Clear History",
            command=self.clear_processing_history
        ).pack(side='right', padx=5)

        # Store reference to history frame
        self.processing_history_frame = history_frame
        
        # Process button section with resume/retry options
        process_button_frame = tk.Frame(parent_frame)
        process_button_frame.pack(pady=20)
        
        # Main process button
        self.bulk_process_button = tk.Button(
            process_button_frame,
            text="Process Bulk Data",
            command=self.process_bulk_data,
            bg=self.success_color,
            fg='white',
            font=('Arial', 14, 'bold'),
            padx=30,
            pady=12,
            cursor='hand2'
        )
        self.bulk_process_button.pack(side='left', padx=5)
        
        # Resume button (initially hidden)
        self.bulk_resume_button = tk.Button(
            process_button_frame,
            text="Resume Processing",
            command=lambda: self.process_bulk_data(resume_mode=True),
            bg=self.primary_color,
            fg='white',
            font=('Arial', 12, 'bold'),
            padx=20,
            pady=10,
            cursor='hand2'
        )
        # Will be shown when there are pending files
        
        # Retry failed button (initially hidden)
        self.bulk_retry_button = tk.Button(
            process_button_frame,
            text="Retry Failed Files",
            command=lambda: self.process_bulk_data(retry_failed=True),
            bg=self.error_color,
            fg='white',
            font=('Arial', 12, 'bold'),
            padx=20,
            pady=10,
            cursor='hand2'
        )
        # Will be shown when there are failed files
        
        # Add debug frame
        debug_frame = tk.Frame(parent_frame)
        debug_frame.pack(pady=10)
        
        tk.Button(
            debug_frame,
            text="📋 View Log File",
            command=self.view_log_file,
            bg='gray',
            fg='white',
            font=('Arial', 10),
            padx=15,
            pady=5,
            cursor='hand2'
        ).pack(side='left', padx=5)
        
        tk.Button(
            debug_frame,
            text="📁 Open Logs Folder",
            command=self.open_logs_folder,
            bg='gray',
            fg='white',
            font=('Arial', 10),
            padx=15,
            pady=5,
            cursor='hand2'
        ).pack(side='left', padx=5)
        
        # Progress display
        self.bulk_progress = scrolledtext.ScrolledText(
            parent_frame,
            height=8,
            wrap='word',
            font=('Courier', 9)
        )
        self.bulk_progress.pack(fill='both', expand=True, padx=20, pady=(0, 20))  
        
        # Configure text formatting tags
        self.configure_bulk_progress_tags()

    def find_rssd_by_name(self):
        """Find RSSD ID by institution name (case-insensitive, partial match)"""
        search_name = self.bulk_inst_name.get().strip()
        
        if not search_name:
            messagebox.showwarning("Input Required", "Please enter an institution name to search")
            return
        
        # Load institution lookup if not already loaded
        if not hasattr(self, 'institution_lookup'):
            self.load_institution_lookup()
        
        # Search for matches (case-insensitive)
        search_lower = search_name.lower()
        matches = []
        
        for rssd_id, inst_name in self.institution_lookup.items():
            if search_lower in inst_name.lower():
                matches.append((rssd_id, inst_name))
        
        if not matches:
            messagebox.showinfo(
                "No Matches", 
                f"No institutions found matching '{search_name}'.\n\n"
                "Try:\n"
                "• A shorter search term\n"
                "• Checking spelling\n"
                "• Using the RSSD ID instead"
            )
            return
        
        elif len(matches) == 1:
            # Single match - auto-populate
            rssd_id, inst_name = matches[0]
            self.bulk_rssd_id.set(rssd_id)
            self.bulk_inst_name.set(inst_name)
            messagebox.showinfo(
                "Match Found",
                f"Found: {inst_name}\nRSSD ID: {rssd_id}"
            )
        
        else:
            # Multiple matches - show selection dialog
            self.show_institution_selector(matches)
    
    def show_institution_selector(self, matches):
        """Show dialog to select from multiple institution matches"""
        # Create selection window
        selector = tk.Toplevel(self.root)
        selector.title("Select Institution")
        selector.geometry("600x400")
        
        # Header
        tk.Label(
            selector,
            text=f"Found {len(matches)} matching institutions:",
            font=('Arial', 12, 'bold')
        ).pack(pady=10)
        
        # Create frame for listbox and scrollbar
        list_frame = tk.Frame(selector)
        list_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        
        # Listbox
        listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            font=('Arial', 10),
            height=15
        )
        listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Add items sorted by institution name
        for rssd_id, inst_name in sorted(matches, key=lambda x: x[1]):
            listbox.insert('end', f"{inst_name} (RSSD: {rssd_id})")
        
        # Selection handler
        def on_select():
            selection = listbox.curselection()
            if selection:
                idx = selection[0]
                rssd_id, inst_name = matches[idx]
                self.bulk_rssd_id.set(rssd_id)
                self.bulk_inst_name.set(inst_name)
                selector.destroy()
            else:
                messagebox.showwarning("No Selection", "Please select an institution")
        
        # Buttons
        button_frame = tk.Frame(selector)
        button_frame.pack(pady=10)
        
        ttk.Button(
            button_frame,
            text="Select",
            command=on_select
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame,
            text="Cancel",
            command=selector.destroy
        ).pack(side='left', padx=5)
        
        # Make window modal
        selector.transient(self.root)
        selector.grab_set()

    def load_institution_lookup(self):
        """Load institution lookup dictionary"""
        lookup_path = os.path.join(
            os.path.dirname(__file__), 
            "dictionaries", 
            "institution_lookup.json"
        )
        
        try:
            with open(lookup_path, 'r', encoding='utf-8') as f:
                self.institution_lookup = json.load(f)
            self.update_bulk_progress(f"Loaded {len(self.institution_lookup)} institution names")
        except Exception as e:
            self.institution_lookup = {}
            messagebox.showerror("Error", f"Failed to load institution names: {str(e)}")
    
    def search_institution(self, target='primary'):
        """Search for institution and update display"""
        if target == 'primary':
            search_term = self.primary_inst_var.get().strip()
        else:
            search_term = self.peer_search_var.get().strip()
        
        if not search_term:
            messagebox.showwarning("Input Required", "Please enter an RSSD ID or institution name")
            return
        
        # Load institution lookup if needed
        if not hasattr(self, 'institution_lookup'):
            self.load_institution_lookup()
        
        # Check if it's an RSSD ID (all digits)
        if search_term.isdigit():
            rssd_id = search_term
            inst_name = self.institution_lookup.get(rssd_id, f"Institution {rssd_id}")
            
            if target == 'primary':
                self.primary_inst_label.config(
                    text=f"✓ {inst_name} (RSSD: {rssd_id})",
                    fg=self.success_color
                )
                # Store primary institution data
                self.primary_institution = {'rssd_id': rssd_id, 'name': inst_name}
                # ADD THIS LINE:
                self.check_multi_institution_validation()
        else:
            # Search by name
            search_lower = search_term.lower()
            matches = []
            
            for rssd_id, inst_name in self.institution_lookup.items():
                if search_lower in inst_name.lower():
                    matches.append((rssd_id, inst_name))
            
            if not matches:
                messagebox.showinfo("No Matches", f"No institutions found matching '{search_term}'")
            elif len(matches) == 1:
                rssd_id, inst_name = matches[0]
                if target == 'primary':
                    self.primary_inst_label.config(
                        text=f"✓ {inst_name} (RSSD: {rssd_id})",
                        fg=self.success_color
                    )
                    self.primary_institution = {'rssd_id': rssd_id, 'name': inst_name}
                    # ADD THIS LINE:
                    self.check_multi_institution_validation()
            else:
                # Multiple matches - show selector
                self.show_multi_institution_selector(matches, target)
                
    def show_multi_institution_selector(self, matches, target):
        """Show dialog to select from multiple institution matches for multi-institution mode"""
        selector = tk.Toplevel(self.root)
        selector.title("Select Institution")
        selector.geometry("600x400")
        
        tk.Label(
            selector,
            text=f"Found {len(matches)} matching institutions:",
            font=('Arial', 12, 'bold')
        ).pack(pady=10)
        
        list_frame = tk.Frame(selector)
        list_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        
        listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            font=('Arial', 10),
            height=15
        )
        listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=listbox.yview)
        
        for rssd_id, inst_name in sorted(matches, key=lambda x: x[1]):
            listbox.insert('end', f"{inst_name} (RSSD: {rssd_id})")
        
        def on_select():
            selection = listbox.curselection()
            if selection:
                idx = selection[0]
                rssd_id, inst_name = matches[idx]
                
                if target == 'primary':
                    self.primary_inst_label.config(
                        text=f"✓ {inst_name} (RSSD: {rssd_id})",
                        fg=self.success_color
                    )
                    self.primary_institution = {'rssd_id': rssd_id, 'name': inst_name}
                else:
                    # Add as peer
                    self.add_peer_to_list(rssd_id, inst_name)
                
                selector.destroy()
            else:
                messagebox.showwarning("No Selection", "Please select an institution")
        
        button_frame = tk.Frame(selector)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="Select", command=on_select).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Cancel", command=selector.destroy).pack(side='left', padx=5)
        
        selector.transient(self.root)
        selector.grab_set()
        
    def add_peer_institution(self):
        """Add a peer institution to the comparison list"""
        search_term = self.peer_search_var.get().strip()
        
        if not search_term:
            messagebox.showwarning("Input Required", "Please enter an RSSD ID or institution name")
            return
        
        # Check if we already have 3 peers
        if len(self.selected_peers) >= 3:
            messagebox.showwarning("Limit Reached", "Maximum 3 peer institutions allowed")
            return
        
        # Load institution lookup if needed
        if not hasattr(self, 'institution_lookup'):
            self.load_institution_lookup()
        
        # Search for institution
        if search_term.isdigit():
            rssd_id = search_term
            inst_name = self.institution_lookup.get(rssd_id, f"Institution {rssd_id}")
            self.add_peer_to_list(rssd_id, inst_name)
        else:
            # Search by name
            search_lower = search_term.lower()
            matches = []
            
            for rssd_id, inst_name in self.institution_lookup.items():
                if search_lower in inst_name.lower():
                    matches.append((rssd_id, inst_name))
            
            if not matches:
                messagebox.showinfo("No Matches", f"No institutions found matching '{search_term}'")
            elif len(matches) == 1:
                rssd_id, inst_name = matches[0]
                self.add_peer_to_list(rssd_id, inst_name)
            else:
                # Multiple matches - show selector
                self.show_multi_institution_selector(matches, 'peer')
                
    def add_peer_to_list(self, rssd_id, inst_name):
        """Add peer to the selected peers list"""
        # Check for duplicates
        if any(p['rssd_id'] == rssd_id for p in self.selected_peers):
            messagebox.showwarning("Duplicate", f"{inst_name} is already in the peer list")
            return
        
        # Check if it's the same as primary
        if hasattr(self, 'primary_institution') and self.primary_institution.get('rssd_id') == rssd_id:
            messagebox.showwarning("Duplicate", "Cannot add primary institution as a peer")
            return
        
        # Add to list
        self.selected_peers.append({'rssd_id': rssd_id, 'name': inst_name})
        self.peers_listbox.insert('end', f"{inst_name} (RSSD: {rssd_id})")
        
        # Clear search box
        self.peer_search_var.set("")
        
        # Update warning if needed
        self.check_multi_institution_validation()
        
    def remove_selected_peer(self):
        """Remove selected peer from the list"""
        selection = self.peers_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a peer to remove")
            return
        
        idx = selection[0]
        self.peers_listbox.delete(idx)
        del self.selected_peers[idx]
        
        # Update validation
        self.check_multi_institution_validation()
        
    def suggest_similar_banks(self):
        """Suggest similar banks based on primary institution (placeholder)"""
        if not hasattr(self, 'primary_institution'):
            messagebox.showwarning("No Primary Institution", "Please select a primary institution first")
            return
        
        # This is a placeholder - in a real implementation, you would:
        # 1. Look up the asset size of the primary institution
        # 2. Find institutions with similar asset sizes
        # 3. Maybe consider geographic proximity or business model
        
        messagebox.showinfo(
            "Coming Soon", 
            "Bank suggestion feature will be implemented in a future update.\n\n"
            "It will suggest banks with similar asset sizes and business models."
        )
        
    def check_multi_institution_validation(self):
        """Check validation for multi-institution selection"""
        warning_msgs = []
        
        # Check minimum institutions
        total_institutions = 1 if hasattr(self, 'primary_institution') else 0
        total_institutions += len(self.selected_peers)
        
        if total_institutions < 2:
            warning_msgs.append(f"⚠️ Minimum 2 institutions required (currently have {total_institutions})")
        elif total_institutions > 4:
            warning_msgs.append(f"⚠️ Maximum 4 institutions allowed (currently have {total_institutions})")
        
        # Update warning label
        if warning_msgs:
            self.multi_warning_label.config(text="\n".join(warning_msgs), fg='orange')
        else:
            self.multi_warning_label.config(text="✓ Valid institution selection", fg=self.success_color)
        
    def check_for_bulk_data(self):
        """Check common locations for FFIEC bulk data"""
        common_locations = [
            os.path.join(os.path.expanduser("~"), "Desktop"),
            os.path.join(os.path.expanduser("~"), "Downloads"),
            os.path.join(os.path.expanduser("~"), "Documents", "FFIEC_Data"),
            "C:\\FFIEC_Data",
            os.path.join(os.path.dirname(__file__), "sample_data")
        ]
        
        found_locations = []
        
        for location in common_locations:
            if os.path.exists(location):
                # Look for FFIEC bulk data folders
                try:
                    for item in os.listdir(location):
                        if "FFIEC" in item and "CDR" in item and os.path.isdir(os.path.join(location, item)):
                            found_locations.append(os.path.join(location, item))
                except PermissionError:
                    continue
        
        return found_locations

    def show_download_guide(self):
        """Open the bulk data download guide"""
        guide_path = os.path.join(os.path.dirname(__file__), "docs", "bulk_data_download_guide.md")
        
        if os.path.exists(guide_path):
            # Open in default markdown viewer or browser
            if sys.platform == 'win32':
                os.startfile(guide_path)
            elif sys.platform == 'darwin':
                subprocess.run(['open', guide_path])
            else:
                subprocess.run(['xdg-open', guide_path])
        else:
            # Show inline instructions if guide not found
            messagebox.showinfo(
                "Download Instructions",
                "To download FFIEC Bulk Data:\n\n"
                "1. Visit: https://cdr.ffiec.gov/public/PWS/DownloadBulkData.aspx\n"
                "2. Select 'Call Reports -- Single Period'\n"
                "3. Choose your reporting period\n"
                "4. Select 'Tab Delimited' format\n"
                "5. Click Download\n"
                "6. Extract the ZIP file\n\n"
                "Then return here and browse to the extracted folder."
            )
         
    def toggle_mode_interface(self):
        """Toggle between single and multi-institution interfaces"""
        if self.institution_mode.get() == "single":
            # Enable single institution interface
            self.bulk_rssd_entry.config(state='normal')
            self.bulk_inst_name_entry.config(state='normal')
            
            # Find and enable the "Find RSSD" button
            for widget in self.bulk_inst_name_entry.master.winfo_children():
                if isinstance(widget, ttk.Button) and widget['text'] == "Find RSSD":
                    widget.config(state='normal')
                    self.find_rssd_button = widget  # Store reference for later
                    break
            
            # Hide multi-institution interface
            if hasattr(self, 'multi_inst_frame'):
                self.multi_inst_frame.pack_forget()
                
            # Hide the disabled notice
            if hasattr(self, 'multi_mode_notice_frame'):
                self.multi_mode_notice_frame.pack_forget()
                
        else:
            # Disable single institution interface
            self.bulk_rssd_entry.config(state='disabled')
            self.bulk_inst_name_entry.config(state='disabled')
            
            # Clear the fields
            self.bulk_rssd_id.set("")
            self.bulk_inst_name.set("")
            
            # Find and disable the "Find RSSD" button
            for widget in self.bulk_inst_name_entry.master.winfo_children():
                if isinstance(widget, ttk.Button) and widget['text'] == "Find RSSD":
                    widget.config(state='disabled')
                    self.find_rssd_button = widget  # Store reference for later
                    break
            
            # Show the disabled notice (simplified pack without 'after')
            if hasattr(self, 'multi_mode_notice_frame'):
                self.multi_mode_notice_frame.pack(fill='x', pady=(5, 10))
            
            # Show multi-institution interface
            if hasattr(self, 'multi_inst_frame'):
                self.multi_inst_frame.pack(fill='x', pady=(10, 0))
     
    def update_bulk_input_fields(self):
        """Toggle between file and directory input fields"""
        if self.bulk_mode.get() == "single":
            self.bulk_dir_frame.grid_remove()
            self.bulk_file_frame.grid()
        else:
            self.bulk_file_frame.grid_remove()
            self.bulk_dir_frame.grid(row=1, column=0, columnspan=4, pady=10, sticky='ew')

    def browse_bulk_file(self):
        """Browse for a single bulk data file"""
        filename = filedialog.askopenfilename(
            title="Select FFIEC Bulk Data File",
            filetypes=[
                ("Text files", "*.txt"),
                ("All files", "*.*")
            ]
        )
        if filename:
            self.bulk_file_path.set(filename)

    def browse_bulk_directory(self):
        """Browse for directory containing bulk data files"""
        directory = filedialog.askdirectory(title="Select Directory with FFIEC Bulk Data Files")
        if directory:
            self.bulk_dir_path.set(directory)

    def browse_dictionary_file(self):
        """Browse for MDRM dictionary file"""
        filename = filedialog.askopenfilename(
            title="Select MDRM Dictionary File",
            filetypes=[
                ("JSON files", "*.json"),
                ("All files", "*.*")
            ]
        )
        if filename:
            self.bulk_dict_path.set(filename)

    def load_mdrm_dictionary(self):
        """Load the MDRM dictionary"""
        dict_path = self.bulk_dict_path.get()
        if not dict_path or not os.path.exists(dict_path):
            messagebox.showwarning("Invalid Path", "Please select a valid dictionary file")
            return
        
        try:
            # Test loading the dictionary
            with open(dict_path, 'r') as f:
                data = json.load(f)
            
            # Check format
            if '_sample_codes' in data:
                count = len(data['_sample_codes'])
            else:
                count = len(data)
            
            self.dict_status_label.config(
                text=f"Dictionary Status: ✓ Loaded ({count:,} codes)",
                fg=self.success_color
            )
            
            # Store path for later use
            self.loaded_dict_path = dict_path
            
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to load dictionary: {str(e)}")
            self.dict_status_label.config(
                text="Dictionary Status: Load Failed",
                fg=self.error_color
            )

    def update_bulk_progress(self, message):
        """Update bulk processing progress display with enhanced formatting"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        
        # Parse message for special formatting
        if "Schedule" in message and ":" in message:
            # Extract schedule code for highlighting
            parts = message.split("Schedule", 1)
            if len(parts) > 1:
                schedule_part = parts[1].split(":", 1)[0].strip()
                remaining = parts[1].split(":", 1)[1] if ":" in parts[1] else ""
                
                # Format with schedule highlighting
                formatted_msg = f"{timestamp} - {parts[0]}Schedule "
                self.bulk_progress.insert('end', formatted_msg)
                
                # Add schedule code in bold/colored
                self.bulk_progress.insert('end', schedule_part, 'schedule_tag')
                
                # Add the rest of the message
                self.bulk_progress.insert('end', f":{remaining}\n")
            else:
                self.bulk_progress.insert('end', f"{timestamp} - {message}\n")
        elif "✓" in message:
            # Success messages in green
            self.bulk_progress.insert('end', f"{timestamp} - ", 'timestamp')
            self.bulk_progress.insert('end', f"{message}\n", 'success_tag')
        elif "✅" in message:
            # Completion messages in bold green
            self.bulk_progress.insert('end', f"{timestamp} - ", 'timestamp')
            self.bulk_progress.insert('end', f"{message}\n", 'complete_tag')
        elif "⚠️" in message:
            # Warning messages in yellow
            self.bulk_progress.insert('end', f"{timestamp} - ", 'timestamp')
            self.bulk_progress.insert('end', f"{message}\n", 'warning_tag')
        elif "File" in message and "/" in message:
            # File progress messages with progress bar visualization
            self.bulk_progress.insert('end', f"{timestamp} - ", 'timestamp')
            self.bulk_progress.insert('end', f"{message}\n", 'progress_tag')
        else:
            # Standard messages
            self.bulk_progress.insert('end', f"{timestamp} - {message}\n")
        
        # Auto-scroll to bottom
        self.bulk_progress.see('end')
        
        # Update GUI to show changes immediately
        self.bulk_progress.update_idletasks()
    
    def configure_bulk_progress_tags(self):
        """Configure text tags for bulk progress display formatting"""
        # Schedule code highlighting
        self.bulk_progress.tag_configure('schedule_tag', 
                                    foreground=self.primary_color, 
                                    font=('Courier', 9, 'bold'))
        
        # Success messages (green)
        self.bulk_progress.tag_configure('success_tag', 
                                    foreground=self.success_color,
                                    font=('Courier', 9))
        
        # Completion messages (bold green)
        self.bulk_progress.tag_configure('complete_tag', 
                                    foreground=self.success_color,
                                    font=('Courier', 9, 'bold'))
        
        # Warning messages (yellow/orange)
        self.bulk_progress.tag_configure('warning_tag', 
                                    foreground='#FFA500',
                                    font=('Courier', 9))
        
        # Progress messages (blue)
        self.bulk_progress.tag_configure('progress_tag', 
                                    foreground=self.secondary_color,
                                    font=('Courier', 9))
        
        # Timestamp (gray)
        self.bulk_progress.tag_configure('timestamp', 
                                    foreground='#808080',
                                    font=('Courier', 9))

    def process_bulk_data(self, resume_mode=False, retry_failed=False):
        """
        Process bulk data files with flexible RSSD/Name validation and resume capability
        
        Args:
            resume_mode (bool): Whether to resume processing from pending files
            retry_failed (bool): Whether to retry previously failed files
        
        This method validates inputs and starts the bulk processing
        in a background thread.
        """
        # ===== ENHANCED VALIDATION FOR FLEXIBILITY =====
        rssd_id = self.bulk_rssd_id.get().strip()
        inst_name = self.bulk_inst_name.get().strip()
        
        # Handle flexible input scenarios
        if self.institution_mode.get() == "single":
            # Single institution mode - need either RSSD or name
            if not rssd_id and not inst_name:
                messagebox.showwarning(
                    "Input Required", 
                    "Please enter either:\n" +
                    "• An RSSD ID, or\n" +
                    "• An institution name\n\n" +
                    "You can also enter both for verification."
                )
                return
            
            # If only name provided, try to find RSSD
            if inst_name and not rssd_id:
                # Load lookup if needed
                if not hasattr(self, 'institution_lookup'):
                    self.load_institution_lookup()
                
                # Try exact match first
                found_rssd = None
                for lookup_rssd, lookup_name in self.institution_lookup.items():
                    if lookup_name.lower() == inst_name.lower():
                        found_rssd = lookup_rssd
                        break
                
                if found_rssd:
                    self.bulk_rssd_id.set(found_rssd)
                    rssd_id = found_rssd
                    self.update_bulk_progress(f"Found RSSD {rssd_id} for {inst_name}")
                else:
                    # Try partial match
                    result = messagebox.askyesno(
                        "No Exact Match",
                        f"No exact match found for '{inst_name}'.\n\n" +
                        "Would you like to search for similar names?"
                    )
                    if result:
                        self.find_rssd_by_name()
                        return  # Let user complete selection first
                    else:
                        # Proceed without RSSD (will process all institutions)
                        messagebox.showinfo(
                            "Processing All Institutions",
                            "Processing will include all institutions.\n" +
                            "This may take longer and create a larger file."
                        )
                        rssd_id = None
            
            # If only RSSD provided, try to find name
            elif rssd_id and not inst_name:
                # Load lookup if needed
                if not hasattr(self, 'institution_lookup'):
                    self.load_institution_lookup()
                
                if rssd_id in self.institution_lookup:
                    found_name = self.institution_lookup[rssd_id]
                    self.bulk_inst_name.set(found_name)
                    inst_name = found_name
                    self.update_bulk_progress(f"Found name '{inst_name}' for RSSD {rssd_id}")
                else:
                    # RSSD not in lookup - continue with just RSSD
                    self.update_bulk_progress(f"RSSD {rssd_id} not found in lookup, continuing...")
        
        # ===== ENHANCED VALIDATION FOR MULTI-INSTITUTION MODE =====
        elif self.institution_mode.get() == "multi":
            # Perform comprehensive validation
            is_valid, warnings = self.validate_multi_institution_selection()
            
            # Show warnings if any exist
            if warnings or not is_valid:
                proceed = self.show_validation_warnings(warnings)
                if not proceed:
                    return
            
            # Show time estimate
            if not self.show_processing_time_estimate(len([self.primary_institution] + self.selected_peers)):
                return
            
            # Prepare institution list for processing
            self.multi_institutions = [self.primary_institution]
            self.multi_institutions.extend(self.selected_peers)
            
            # Log the comparison details
            self.update_bulk_progress("=" * 60)
            self.update_bulk_progress("🏦 MULTI-INSTITUTION COMPARISON MODE")
            self.update_bulk_progress("=" * 60)
            self.update_bulk_progress(f"Primary Institution: {self.primary_institution['name']}")
            for i, peer in enumerate(self.selected_peers, 1):
                self.update_bulk_progress(f"Peer {i}: {peer['name']}")
            self.update_bulk_progress("=" * 60)
        
        # ===== CHECK FOR PENDING/FAILED FILES IF RESUMING =====
        if resume_mode or retry_failed:
            # Initialize file manager if needed
            if not hasattr(self, 'file_manager'):
                from bulk_file_manager import BulkFileManager
                self.file_manager = BulkFileManager()
            
            # Get quarter if selected
            quarter = self.selected_quarter.get() if hasattr(self, 'selected_quarter') else None
            
            if retry_failed:
                # Get failed files
                failed_files = self.file_manager.get_failed_files(quarter)
                if not failed_files:
                    messagebox.showinfo("No Failed Files", "No failed files to retry.")
                    return
                
                # Reset failed files to pending
                self.file_manager.reset_failed_files(quarter)
                self.update_bulk_progress(f"Reset {len(failed_files)} failed files for retry")
            
            # Check for pending files
            pending_files = self.file_manager.get_pending_files(quarter)
            if not pending_files:
                messagebox.showinfo(
                    "All Files Processed", 
                    "All files have been processed successfully!\n" +
                    "No pending files to resume."
                )
                return
            
            self.update_bulk_progress(f"Found {len(pending_files)} pending files to process")
        
        # ===== ORIGINAL VALIDATION CONTINUES =====
        # Check if directory is selected
        if self.bulk_mode.get() == "directory":
            selected_dir = self.bulk_dir_path.get()
            
            if not selected_dir:
                # Try auto-detected folder
                if hasattr(self, 'found_data_var') and self.found_data_var.get():
                    selected_dir = self.found_data_var.get()
                    self.bulk_dir_path.set(selected_dir)
                else:
                    # Show helpful message
                    result = messagebox.askyesno(
                        "No Data Selected",
                        "No FFIEC data folder selected.\n\n"
                        "Would you like to see the download guide?"
                    )
                    if result:
                        self.show_download_guide()
                    return
            
            # Verify it's a valid FFIEC folder
            if os.path.exists(selected_dir):
                txt_files = [f for f in os.listdir(selected_dir) if f.endswith('.txt')]
                ffiec_files = [f for f in txt_files if 'FFIEC' in f and 'Schedule' in f]
                
                if len(ffiec_files) < 10:
                    messagebox.showwarning(
                        "Invalid Folder",
                        f"Selected folder contains only {len(ffiec_files)} FFIEC files.\n"
                        f"Expected ~47 schedule files.\n\n"
                        f"Please select the extracted FFIEC bulk data folder."
                    )
                    return
        
        # File mode validation
        if self.bulk_mode.get() == "single":
            if not self.bulk_file_path.get():
                messagebox.showwarning("Input Required", "Please select a file to process")
                return
        else:
            if not self.bulk_dir_path.get():
                messagebox.showwarning("Input Required", "Please select a directory to process")
                return
        
        # Check dictionary
        if not hasattr(self, 'loaded_dict_path'):
            messagebox.showwarning("Dictionary Required", "Please load the MDRM dictionary first")
            return
        
        # ===== UPDATE UI BUTTONS =====
        # Hide/show resume and retry buttons based on file status
        if hasattr(self, 'bulk_resume_button'):
            self.bulk_resume_button.pack_forget()
        if hasattr(self, 'bulk_retry_button'):
            self.bulk_retry_button.pack_forget()
        
        # Clear progress
        self.bulk_progress.delete('1.0', 'end')
        
        # Reset progress summary if it exists
        if hasattr(self, 'progress_summary_frame'):
            # Show the progress frame
            self.progress_summary_frame.pack(fill='x', padx=20, pady=10, before=self.bulk_process_button)
            
            # Reset all progress indicators
            self.current_schedule_label.config(text="Initializing...")
            self.files_progress_label.config(text="0 / 0")
            self.schedules_completed_label.config(text="0")
            self.overall_progress_label.config(text="0%")
            self.bulk_progress_bar['value'] = 0
        
        # Disable button
        self.bulk_process_button.config(state='disabled', text="Processing...")
        
        # Store final values for thread based on mode
        if self.institution_mode.get() == "multi":
            # Multi-institution mode
            self.processing_mode = "multi"
            self.processing_institutions = self.multi_institutions
        else:
            # Single institution mode
            self.processing_mode = "single"
            self.processing_rssd = rssd_id
            self.processing_inst_name = inst_name
        
        self.processing_resume_mode = resume_mode
        self.processing_retry_failed = retry_failed
        
        # Run processing in thread
        thread = threading.Thread(target=self._process_bulk_data_thread)
        thread.daemon = True
        thread.start()
      
    
    #process_bulk_data_multi
    def process_bulk_data_multi(self):
        """Process bulk data for multiple institutions comparison"""
        try:
            # Import the processor
            from bulk_data_processor import BulkDataProcessor
            from bulk_file_manager import BulkDataOrganizer
            
            self.queue.put(('progress', 'Initializing multi-institution processor...'))
            
            # Create processor instance
            processor = BulkDataProcessor(
                dictionary_path=self.loaded_dict_path
            )
            
            # Get the list of institutions
            institutions = self.multi_institutions
            
            # Store all results by institution
            all_institution_results = {}
            
            # Process each institution
            for idx, inst in enumerate(institutions):
                rssd_id = inst['rssd_id']
                inst_name = inst['name']
                
                self.queue.put(('progress', f'Processing institution {idx + 1} of {len(institutions)}: {inst_name}'))
                
                if self.bulk_mode.get() == "directory":
                    directory = self.bulk_dir_path.get()
                    
                    # Check if using quarter selection
                    if hasattr(self, 'selected_quarter') and self.selected_quarter.get():
                        # Use file manager for organized processing
                        if not hasattr(self, 'file_manager'):
                            from bulk_file_manager import BulkFileManager
                            self.file_manager = BulkFileManager()
                        
                        organizer = BulkDataOrganizer(self.file_manager, processor)
                        
                        quarter = self.selected_quarter.get()
                        self.queue.put(('progress', f'Processing {inst_name} for quarter {quarter}...'))
                        
                        # Prepare batch
                        files = organizer.prepare_quarter_batch(quarter, directory)
                        
                        if not files:
                            self.queue.put(('progress', f'No files found for {inst_name}'))
                            continue
                        
                        # Process with progress tracking
                        def multi_progress_callback(progress):
                            # Build enhanced message with institution context
                            institution_prefix = f"{inst_name} ({idx + 1}/{len(institutions)})"
                            
                            if 'message' in progress:
                                msg = f"{institution_prefix}: {progress['message']}"
                            elif 'current_schedule' in progress:
                                schedule = progress['current_schedule']
                                msg = f"{institution_prefix} - Processing Schedule {schedule}"
                                if 'schedule_name' in progress:
                                    msg += f": {progress['schedule_name']}"
                            else:
                                msg = f"{institution_prefix} - Processing..."
                            
                            self.queue.put(('progress', msg))
                            
                            # Update progress bar with overall progress
                            if 'percentage' in progress:
                                # Calculate overall progress considering all institutions
                                institution_progress = (idx / len(institutions)) * 100
                                file_progress = progress['percentage'] / len(institutions)
                                overall_progress = institution_progress + file_progress
                                
                                if hasattr(self, 'bulk_progress_bar'):
                                    self.bulk_progress_bar['value'] = overall_progress
                                    self.overall_progress_label.config(text=f"{overall_progress:.1f}%")

                        result = organizer.process_batch(
                            files, 
                            target_rssd_id=rssd_id,
                            progress_callback=multi_progress_callback
                        )
                        
                        # Extract data from result
                        if isinstance(result, dict) and 'data' in result:
                            institution_data = result['data']
                        else:
                            institution_data = result
                    else:
                        # Process directory without quarter selection
                        self.queue.put(('progress', f'Processing directory for {inst_name}...'))
                        institution_data = processor.process_directory(
                            directory, 
                            target_rssd_id=rssd_id
                        )
                else:
                    # Single file mode
                    filepath = self.bulk_file_path.get()
                    self.queue.put(('progress', f'Processing file for {inst_name}...'))
                    
                    df = processor.process_bulk_file(filepath, target_rssd_id=rssd_id)
                    if not df.empty:
                        schedule_code = processor.line_mapper.parse_schedule_code(os.path.basename(filepath))
                        institution_data = {schedule_code: df}
                    else:
                        institution_data = {}
                
                # Store results
                if institution_data:
                    all_institution_results[rssd_id] = {
                        'name': inst_name,
                        'data': institution_data
                    }
                    self.queue.put(('progress', f'✓ Completed {inst_name}: {len(institution_data)} schedules'))
                else:
                    self.queue.put(('progress', f'⚠️ No data found for {inst_name}'))
            
            # Create comparison Excel file
            if all_institution_results:
                self.queue.put(('progress', '💾 Creating comparison Excel file...'))
                
                # Generate output filename
                primary_name = institutions[0]['name'].replace(' ', '_').replace('/', '_').replace('\\', '_')
                filename = f"{primary_name}_Multi_Institution_Comparison_{datetime.now().strftime('%Y%m%d')}.xlsx"
                output_path = os.path.join(self.output_path_var.get(), filename)
                
                # Create the comparison Excel
                self.save_multi_institution_excel(all_institution_results, output_path, processor)
                
                # Success
                total_schedules = len(next(iter(all_institution_results.values()))['data'])
                total_institutions = len(all_institution_results)

                # Update progress to 100%
                if hasattr(self, 'bulk_progress_bar'):
                    self.bulk_progress_bar['value'] = 100
                    self.overall_progress_label.config(text="100%")

                self.queue.put(('bulk_complete', {
                    'output_path': output_path,
                    'schedules': total_schedules,
                    'total_rows': sum(
                        sum(len(df) for df in inst_data['data'].values())
                        for inst_data in all_institution_results.values()
                    ),
                    'institutions': total_institutions
                }))
            else:
                self.queue.put(('error', 'No data found for any institutions'))
                
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            
            error_log_path = os.path.join(
                os.path.dirname(__file__), 
                'logs', 
                f'multi_error_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
            )
            os.makedirs(os.path.dirname(error_log_path), exist_ok=True)
            
            with open(error_log_path, 'w') as f:
                f.write(f"Multi-Institution Processing Error: {str(e)}\n\n")
                f.write(f"Traceback:\n{error_details}\n\n")
            
            error_msg = (f'Multi-institution processing error: {str(e)}\n\n'
                        f'Detailed error log saved to:\n{error_log_path}')
            self.queue.put(('error', error_msg))
            
    def save_multi_institution_excel(self, all_institution_results, output_path, processor):
        """Save multi-institution comparison to Excel with side-by-side format"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create summary sheet
            ws_summary = wb.create_sheet("Summary")
            self._create_multi_summary_sheet(ws_summary, all_institution_results)
            
            
            # NEW: Create Executive Dashboard (Excel Prototype)
         
            try:
                self.queue.put(('progress', '📊 Creating Executive Dashboard...'))
                enhancer = ExcelEnhancementProcessor(self.logger if hasattr(self, 'logger') else None)
                
                # Executive Dashboard
                ws_dashboard = wb.create_sheet("Executive Dashboard")
                enhancer.create_executive_dashboard(ws_dashboard, all_institution_results, processor)
          
                # Key Metrics
                # Key Metrics
                self.queue.put(('progress', '📈 Creating Key Metrics sheet...'))
                ws_metrics = wb.create_sheet("Key Metrics")
                enhancer.create_key_metrics_sheet(ws_metrics, all_institution_results, processor)
                
                self.queue.put(('progress', '📊 Created Executive Dashboard and Key Metrics sheets'))
                
            except Exception as e:
                # If enhancement fails, log but continue with standard report
                self.queue.put(('progress', f'⚠️ Could not create enhanced sheets: {str(e)}'))
                # Remove any partially created sheets
                if "Executive Dashboard" in wb.sheetnames:
                    wb.remove(wb["Executive Dashboard"])
                if "Key Metrics" in wb.sheetnames:
                    wb.remove(wb["Key Metrics"])
            
            # Get all unique schedules across all institutions
            all_schedules = set()
            for inst_data in all_institution_results.values():
                all_schedules.update(inst_data['data'].keys())
            
            for schedule_code in sorted(all_schedules):
                self.queue.put(('progress', f'Creating comparison sheet for Schedule {schedule_code}'))
                
                # Calculate progress for Excel creation
                schedule_index = sorted(all_schedules).index(schedule_code)
                excel_progress = (schedule_index / len(all_schedules)) * 100
                
                if hasattr(self, 'bulk_progress_bar'):
                    # Show Excel creation progress (90-100% range)
                    final_progress = 90 + (excel_progress * 0.1)
                    self.bulk_progress_bar['value'] = final_progress
                    self.overall_progress_label.config(text=f"{final_progress:.1f}%")

                # Create sheet for this schedule
                sheet_name = f"Schedule {schedule_code}"[:31]  # Excel limit
                ws = wb.create_sheet(sheet_name)
                
                # Headers
                headers = ['Line Item', 'Description', 'MDRM Code']
                institutions = list(all_institution_results.values())
                
                # Add institution names as headers
                for inst in institutions:
                    headers.append(inst['name'])
                
                # Style definitions
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Write headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Collect all unique line items across institutions for this schedule
                all_line_items = {}
                
                for rssd_id, inst_data in all_institution_results.items():
                    if schedule_code in inst_data['data']:
                        schedule_df = inst_data['data'][schedule_code]
                        
                        # Group by line item
                        for _, row in schedule_df.iterrows():
                            line_item = row.get('Line Item', '')
                            mdrm_code = row.get('MDRM Code', '')
                            description = row.get('Description', '')
                            
                            # Create unique key
                            key = (line_item, mdrm_code)
                            
                            if key not in all_line_items:
                                all_line_items[key] = {
                                    'line_item': line_item,
                                    'description': description,
                                    'mdrm_code': mdrm_code,
                                    'values': {}
                                }
                            
                            # Store value for this institution
                            all_line_items[key]['values'][rssd_id] = row.get('Amount', '')
                
                # Sort line items by line item number/code
                sorted_items = sorted(all_line_items.items(), 
                                    key=lambda x: self._sort_line_item_key(x[1]['line_item']))
                
                # Write data rows
                row_num = 2
                
                # Alternating row colors
                light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                for (line_item, mdrm_code), item_data in sorted_items:
                    # Line Item
                    ws.cell(row=row_num, column=1, value=item_data['line_item'])
                    
                    # Description
                    ws.cell(row=row_num, column=2, value=item_data['description'])
                    
                    # MDRM Code
                    ws.cell(row=row_num, column=3, value=item_data['mdrm_code'])
                    
                    # Institution values
                    col_num = 4
                    for rssd_id, inst_data in all_institution_results.items():
                        value = item_data['values'].get(rssd_id, '—')
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        
                        # Format numbers
                        if value and value != '—':
                            try:
                                num_value = float(str(value).replace(',', ''))
                                cell.value = num_value
                                cell.number_format = '#,##0'
                            except:
                                pass
                        
                        # Center align
                        cell.alignment = Alignment(horizontal='center')
                        
                        col_num += 1
                    
                    # Apply alternating row color
                    if row_num % 2 == 0:
                        for col in range(1, len(headers) + 1):
                            ws.cell(row=row_num, column=col).fill = light_fill
                    
                    row_num += 1
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in ws.iter_rows(min_row=1, max_row=row_num-1, min_col=1, max_col=len(headers)):
                    for cell in row:
                        cell.border = thin_border
                
                # Auto-fit columns
                for column_cells in ws.columns:
                    length = max(len(str(cell.value or "")) for cell in column_cells)
                    col_letter = column_cells[0].column_letter
                    
                    if col_letter == 'A':  # Line Item
                        adjusted_width = min(max(length, 10), 20)
                    elif col_letter == 'B':  # Description
                        adjusted_width = min(max(length, 40), 70)
                    elif col_letter == 'C':  # MDRM Code
                        adjusted_width = min(max(length, 10), 15)
                    else:  # Institution columns
                        adjusted_width = min(max(length, 15), 30)
                    
                    ws.column_dimensions[col_letter].width = adjusted_width
                
                # Freeze panes (keep headers and first 3 columns visible)
                ws.freeze_panes = 'D2'
                
                # Add hyperlink back to summary
                back_cell = ws.cell(row=1, column=len(headers)+2, value="← Back to Summary")
                back_cell.hyperlink = "#Summary!A1"
                back_cell.font = Font(color="0563C1", underline="single", bold=True)
            
            # Save workbook
            wb.save(output_path)
            self.queue.put(('progress', f'✅ Comparison Excel saved: {output_path}'))
            
        except Exception as e:
            raise Exception(f"Error creating comparison Excel: {str(e)}")
 
    # Update _create_multi_summary_sheet in fire_analyzer_gui.py
    def _create_multi_summary_sheet(self, ws, all_institution_results):
        """Create summary sheet for multi-institution comparison"""
        # Import required styling classes
        from openpyxl.styles import Font
        
        # Title
        ws['A1'] = "Multi-Institution Comparison Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Report info
        ws['A3'] = "Report Date:"
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        ws['A4'] = "Institutions Compared:"
        ws['B4'] = len(all_institution_results)
        
        # NEW: Add quick links to enhanced reports
        ws['A6'] = "Quick Links:"
        ws['A6'].font = Font(bold=True)
        
        # Executive Dashboard link
        ws['A7'] = "📊 Executive Dashboard"
        ws['A7'].hyperlink = "#'Executive Dashboard'!A1"
        ws['A7'].font = Font(color="0563C1", underline="single", bold=True)
        
        # Key Metrics link
        ws['A8'] = "📈 Key Metrics & Peer Analysis"
        ws['A8'].hyperlink = "#'Key Metrics'!A1"
        ws['A8'].font = Font(color="0563C1", underline="single", bold=True)
        
        # Institution list
        ws['A10'] = "Institutions:"
        ws['A10'].font = Font(bold=True)
        
        row = 11
        for idx, (rssd_id, inst_data) in enumerate(all_institution_results.items(), 1):
            ws.cell(row=row, column=1, value=f"{idx}.")
            ws.cell(row=row, column=2, value=inst_data['name'])
            ws.cell(row=row, column=3, value=f"RSSD: {rssd_id}")
            ws.cell(row=row, column=4, value=f"{len(inst_data['data'])} schedules")
            
            if idx == 1:
                ws.cell(row=row, column=5, value="(Primary)").font = Font(italic=True, color="666666")
            
            row += 1
        
        # Schedule summary
        row += 1
        ws.cell(row=row, column=1, value="Schedules Included:").font = Font(bold=True)
        row += 1
    
        # Get all schedules
        all_schedules = set()
        for inst_data in all_institution_results.values():
            all_schedules.update(inst_data['data'].keys())
        
        for schedule_code in sorted(all_schedules):
            cell = ws.cell(row=row, column=1, value=schedule_code)
            # Add hyperlink to schedule sheet
            sheet_name = f"Schedule {schedule_code}"[:31]
            cell.hyperlink = f"#{sheet_name}!A1"
            cell.font = Font(color="0563C1", underline="single")
            
            # Add schedule name if available
            schedule_info = self.processor.dictionary.get_schedule_info(schedule_code) if hasattr(self, 'processor') else {}
            ws.cell(row=row, column=2, value=schedule_info.get('name', ''))
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15  
           
    def _sort_line_item_key(self, line_item):
        """Create a sortable key for line items"""
        if not line_item:
            return (999, 0, 0, "")
        
        # Try to parse line item format (e.g., "RC.1.a")
        parts = line_item.split('.')
        
        try:
            if len(parts) >= 2:
                # Extract schedule and main number
                schedule = parts[0]
                main_num = int(parts[1]) if parts[1].isdigit() else 999
                
                # Extract sub-parts
                sub_num = 0
                sub_letter = ""
                
                if len(parts) > 2:
                    sub_part = parts[2]
                    if sub_part.isdigit():
                        sub_num = int(sub_part)
                    else:
                        sub_letter = sub_part
                
                return (0, main_num, sub_num, sub_letter)
        except:
            pass
        
        # Fallback - just use the string
        return (999, 0, 0, line_item)

    def refresh_quarters(self):
        """Refresh available quarters from selected directory"""
        if self.bulk_mode.get() != "directory":
            messagebox.showinfo("Info", "Please select 'Directory' mode to scan for quarters")
            return
            
        directory = self.bulk_dir_path.get()
        if not directory:
            messagebox.showwarning("Input Required", "Please select a directory first")
            return
        
        try:
            # Initialize file manager if needed
            if not hasattr(self, 'file_manager'):
                self.file_manager = BulkFileManager()
            
            # Clear progress
            self.bulk_progress.delete('1.0', 'end')
            self.update_bulk_progress("Scanning directory for quarters...")
            
            # Scan directory
            quarters = self.file_manager.scan_directory(directory)
            
            # Update dropdown
            quarter_list = sorted(quarters.keys(), reverse=True)
            self.quarter_dropdown['values'] = quarter_list
            
            if quarter_list:
                self.selected_quarter.set(quarter_list[0])
                self.update_quarter_info()
                self.update_bulk_progress(f"Found {len(quarter_list)} quarters")
            else:
                self.update_bulk_progress("No FFIEC bulk files found in directory")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to scan directory: {str(e)}")

    def update_quarter_info(self):
        """Update display with selected quarter information"""
        if not self.selected_quarter.get() or not hasattr(self, 'file_manager'):
            return
        
        quarter = self.selected_quarter.get()
        quarters = self.file_manager.get_cached_metadata()
        
        if quarter in quarters:
            files = quarters[quarter]
            validation = self.file_manager.validate_quarter_completeness(files)
            
            info_text = f"""Quarter: {quarter}
    Files Found: {len(files)} of {validation['expected_count']} expected
    Completeness: {validation['completeness_percentage']:.1f}%
    Status: {'✅ Complete' if validation['is_complete'] else '⚠️ Incomplete'}"""
            
            if validation['missing_schedules']:
                info_text += f"\nMissing: {', '.join(validation['missing_schedules'][:5])}"
                if len(validation['missing_schedules']) > 5:
                    info_text += f" (+{len(validation['missing_schedules'])-5} more)"
            
            self.quarter_info_label.config(text=info_text, fg='black')
            
            # Add this line to check processing status when quarter changes
            self.check_bulk_processing_status()
    
    def check_bulk_processing_status(self):
        """Check for pending or failed files and show appropriate buttons"""
        try:
            # Initialize file manager if needed
            if not hasattr(self, 'file_manager'):
                from bulk_file_manager import BulkFileManager
                self.file_manager = BulkFileManager()
            
            # Get current quarter if selected
            quarter = self.selected_quarter.get() if hasattr(self, 'selected_quarter') and self.selected_quarter.get() else None
            
            # Check for pending files
            pending_files = self.file_manager.get_pending_files(quarter)
            
            # Check for failed files  
            failed_files = self.file_manager.get_failed_files(quarter)
            
            # Update button visibility based on status
            if hasattr(self, 'bulk_resume_button'):
                if pending_files:
                    self.bulk_resume_button.pack(side='left', padx=5)
                else:
                    self.bulk_resume_button.pack_forget()
                    
            if hasattr(self, 'bulk_retry_button'):
                if failed_files:
                    self.bulk_retry_button.pack(side='left', padx=5)
                else:
                    self.bulk_retry_button.pack_forget()
                    
            # Update status display if processing history exists
            if hasattr(self, 'processing_status_text'):
                status_lines = []
                
                if pending_files:
                    status_lines.append(f"📋 {len(pending_files)} files pending processing")
                    
                if failed_files:
                    status_lines.append(f"❌ {len(failed_files)} files failed processing")
                    
                if not pending_files and not failed_files:
                    status_lines.append("✅ All files processed successfully")
                    
                self.processing_status_text.config(state='normal')
                self.processing_status_text.delete('1.0', 'end')
                self.processing_status_text.insert('1.0', '\n'.join(status_lines))
                self.processing_status_text.config(state='disabled')
                
        except Exception as e:
            self.logger.warning(f"Could not check processing status: {e}")
            
    def clear_processing_history(self):
        """Clear the processing history database and reset status display"""
        try:
            result = messagebox.askyesno(
                "Clear Processing History",
                "This will reset all file processing records.\n\n" +
                "Files will appear as unprocessed and can be processed again.\n\n" +
                "Are you sure you want to clear the history?"
            )
            
            if result:
                # Initialize file manager if needed
                if not hasattr(self, 'file_manager'):
                    from bulk_file_manager import BulkFileManager
                    self.file_manager = BulkFileManager()
                
                # Clear the database
                with sqlite3.connect(self.file_manager.db_path) as conn:
                    conn.execute("DELETE FROM file_metadata")
                    conn.commit()
                
                # Clear the status display
                if hasattr(self, 'processing_status_text'):
                    self.processing_status_text.config(state='normal')
                    self.processing_status_text.delete('1.0', 'end')
                    self.processing_status_text.insert('1.0', "✨ Processing history cleared")
                    self.processing_status_text.config(state='disabled')
                
                # Hide resume/retry buttons
                if hasattr(self, 'bulk_resume_button'):
                    self.bulk_resume_button.pack_forget()
                if hasattr(self, 'bulk_retry_button'):
                    self.bulk_retry_button.pack_forget()
                    
                # Refresh quarter info if a quarter is selected
                if hasattr(self, 'selected_quarter') and self.selected_quarter.get():
                    self.refresh_quarters()
                    
                messagebox.showinfo("Success", "Processing history has been cleared.")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to clear history: {str(e)}")
       
    def on_tab_changed(self, event):
        """Handle tab change events"""
        try:
            # Get the currently selected tab
            current_tab = self.notebook.index(self.notebook.select())
            
            # Check if we're switching to the bulk data tab (index 3)
            if current_tab == 3:  # Bulk Data Processing tab
                # Check for pending/failed files when entering the tab
                self.check_bulk_processing_status()
                
        except Exception as e:
            # Silently handle any errors to avoid disrupting the UI
            pass

    #UPDATE       
    def _process_bulk_data_thread(self):
        """
        Thread function for bulk data processing
        
        This runs in the background to process FFIEC bulk data files
        and convert them to standardized Excel format.
        """
        try:
            # Check if we're in multi-institution mode
            if hasattr(self, 'processing_mode') and self.processing_mode == "multi":
                # Call the multi-institution processing method
                self.process_bulk_data_multi()
                return
            
            # Otherwise, continue with single institution processing
            # Import the processor
            from bulk_data_processor import BulkDataProcessor
            # Import file management
            from bulk_file_manager import BulkDataOrganizer 
            
            self.queue.put(('progress', 'Initializing bulk data processor...'))
            
            # Create processor instance
            processor = BulkDataProcessor(
                dictionary_path=self.loaded_dict_path
            )
            
            # Get filter parameters (use stored values from validation)
            rssd_id = getattr(self, 'processing_rssd', None)
            inst_name = getattr(self, 'processing_inst_name', None)
            resume_mode = getattr(self, 'processing_resume_mode', False)
            retry_failed = getattr(self, 'processing_retry_failed', False)
            
            # If no institution name was found/provided, use a default
            if not inst_name and rssd_id:
                inst_name = f"Institution_{rssd_id}"
            elif not inst_name:
                inst_name = "All_Institutions"
            
            if self.bulk_mode.get() == "single":
                # Process single file
                filepath = self.bulk_file_path.get()
                schedule_code = processor.line_mapper.parse_schedule_code(os.path.basename(filepath))
                self.queue.put(('progress', f'Processing file: {os.path.basename(filepath)} (Schedule {schedule_code})'))
                
                df = processor.process_bulk_file(filepath, target_rssd_id=rssd_id)
                
                if df.empty:
                    self.queue.put(('error', 'No data found in file'))
                    return
                
                # Create results dictionary
                results = {schedule_code: df}
            else:
                # Process directory with file management
                directory = self.bulk_dir_path.get()
                
                # Enhanced progress callback for detailed updates
                def enhanced_progress_callback(progress):
                    # Build detailed message
                    msg_parts = []
                    
                    # Add file progress
                    if 'current_file' in progress and 'total_files' in progress:
                        msg_parts.append(f"File {progress['current_file']}/{progress['total_files']}")
                    
                    # Add schedule info
                    if 'current_schedule' in progress:
                        schedule = progress['current_schedule']
                        if 'schedule_name' in progress:
                            msg_parts.append(f"Schedule {schedule}: {progress['schedule_name']}")
                        else:
                            msg_parts.append(f"Schedule {schedule}")

                    # Add multi-institution context if available
                    if hasattr(self, 'processing_mode') and self.processing_mode == "multi":
                        if hasattr(self, 'current_institution_info'):
                            msg_parts.insert(0, self.current_institution_info)
                    
                    # Add percentage
                    if 'percentage' in progress:
                        msg_parts.append(f"({progress['percentage']:.1f}%)")
                    
                    # Build full message
                    if 'message' in progress:
                        full_msg = f"{' - '.join(msg_parts)} - {progress['message']}"
                    else:
                        full_msg = ' - '.join(msg_parts)
                    
                    # Send to queue
                    self.queue.put(('progress', full_msg))
                    
                    # Send special updates for schedule completion
                    if progress.get('schedule_completed'):
                        self.queue.put(('schedule_complete', progress['current_schedule']))
                    
                    if progress.get('batch_complete'):
                        self.queue.put(('progress', f"✅ Completed {progress.get('schedules_processed', 0)} schedules"))
                
                # Check if using quarter selection with resume support
                if hasattr(self, 'selected_quarter') and self.selected_quarter.get():
                    # Use file manager for organized processing
                    if not hasattr(self, 'file_manager'):
                        from bulk_file_manager import BulkFileManager
                        self.file_manager = BulkFileManager()
                    
                    organizer = BulkDataOrganizer(self.file_manager, processor)
                    
                    quarter = self.selected_quarter.get()
                    
                    if resume_mode or retry_failed:
                        self.queue.put(('progress', f'🔄 Resuming quarter {quarter} processing...'))
                        # Prepare batch with resume capability
                        files = organizer.prepare_quarter_batch_with_resume(quarter, directory, retry_failed)
                    else:
                        self.queue.put(('progress', f'🗓️ Processing quarter {quarter}...'))
                        # Prepare regular batch
                        files = organizer.prepare_quarter_batch(quarter, directory)
                    
                    if not files:
                        self.queue.put(('progress', '✅ All files already processed successfully!'))
                        self.queue.put(('bulk_complete', {
                            'output_path': 'No new files to process',
                            'schedules': 0,
                            'total_rows': 0
                        }))
                        return
                    
                    self.queue.put(('progress', f'📁 Found {len(files)} files to process'))
                    
                    # Process with enhanced progress tracking and resume support
                    result = organizer.process_batch(files, target_rssd_id=rssd_id, 
                                                progress_callback=enhanced_progress_callback,
                                                resume_mode=resume_mode)
                    
                    # Handle both data and failed files
                    if isinstance(result, dict) and 'data' in result:
                        results = result['data']
                        failed_files = result.get('failed_files', [])
                        
                        # Show resume/retry buttons if there are failed files
                        if failed_files and hasattr(self, 'bulk_retry_button'):
                            self.queue.put(('show_retry_button', None))
                    else:
                        results = result
                else:
                    # Fallback to original directory processing with progress
                    self.queue.put(('progress', f'📁 Processing directory: {directory}'))
                    results = processor.process_directory(directory, target_rssd_id=rssd_id,
                                                        progress_callback=enhanced_progress_callback)
                        
                if not results:
                    self.queue.put(('error', 'No valid files found in directory'))
                    return
            
            
            # Generate output filename
            # Generate output filename
            if rssd_id and inst_name:
                filename = f"{inst_name.replace(' ', '_')}_CallReport_BulkData.xlsx"
            elif rssd_id:
                filename = f"Institution_{rssd_id}_CallReport_BulkData.xlsx"
            else:
                filename = f"FFIEC_CallReport_BulkData_{datetime.now().strftime('%Y%m%d')}.xlsx"

            output_path = os.path.join(self.output_path_var.get(), filename)

            # Save to Excel with enhancements enabled for single institution
            self.queue.put(('progress', f'💾 Saving to Excel: {filename}'))

            # Pass rssd_id and enable enhancements for single institution mode
            processor.save_to_excel(
                results, 
                output_path, 
                institution_name=inst_name,
                include_enhancements=True,  # Enable Executive Dashboard and Key Metrics
                rssd_id=rssd_id  # Pass RSSD ID for proper data lookup
            )

            # Success with enhanced flag
            self.queue.put(('bulk_complete', {
                'output_path': output_path,
                'schedules': len(results),
                'total_rows': sum(len(df) for df in results.values()),
                'enhanced': True  # Flag to indicate enhanced reports were created
            }))

        except Exception as e:
            import traceback
            
            # Get detailed error information
            error_details = traceback.format_exc()
            
            # Log to file
            if 'processor' in locals() and hasattr(processor, 'logger'):
                processor.logger.error(f"Processing failed: {str(e)}")
                processor.logger.error(f"Full traceback:\n{error_details}")
            
            # Create error log file
            error_log_path = os.path.join(
                os.path.dirname(__file__), 
                'logs', 
                f'error_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
            )
            os.makedirs(os.path.dirname(error_log_path), exist_ok=True)
            
            with open(error_log_path, 'w') as f:
                f.write(f"Error: {str(e)}\n\n")
                f.write(f"Traceback:\n{error_details}\n\n")
                f.write(f"Settings:\n")
                f.write(f"  Mode: {self.bulk_mode.get()}\n")
                f.write(f"  RSSD: {getattr(self, 'processing_rssd', 'None')}\n")
                f.write(f"  Directory: {self.bulk_dir_path.get()}\n")
            
            # Send error message with log location
            error_msg = (f'Processing error: {str(e)}\n\n'
                        f'Detailed error log saved to:\n{error_log_path}')
            self.queue.put(('error', error_msg))
            
            # Re-enable button
            self.bulk_process_button.config(state='normal', text="Process Bulk Data")

    def view_log_file(self):
        """Open the most recent log file"""
        log_dir = os.path.join(os.path.dirname(__file__), 'logs')
        if not os.path.exists(log_dir):
            messagebox.showinfo("No Logs", "No log files found yet.")
            return
        
        # Find most recent log file
        log_files = [f for f in os.listdir(log_dir) if f.endswith('.log')]
        if not log_files:
            messagebox.showinfo("No Logs", "No log files found yet.")
            return
        
        latest_log = max([os.path.join(log_dir, f) for f in log_files], 
                        key=os.path.getctime)
        
        # Open in default text editor
        if sys.platform == 'win32':
            os.startfile(latest_log)
        elif sys.platform == 'darwin':
            subprocess.run(['open', latest_log])
        else:
            subprocess.run(['xdg-open', latest_log])
    
    def open_logs_folder(self):
        """Open the logs folder"""
        log_dir = os.path.join(os.path.dirname(__file__), 'logs')
        os.makedirs(log_dir, exist_ok=True)
        
        if sys.platform == 'win32':
            os.startfile(log_dir)
        elif sys.platform == 'darwin':
            subprocess.run(['open', log_dir])
        else:
            subprocess.run(['xdg-open', log_dir])


    def validate_multi_institution_selection(self):
        """
        Comprehensive validation for multi-institution comparison
        Returns: (is_valid, warnings_list)
        """
        warnings = []
        is_valid = True
        
        # Check if we have a primary institution
        if not hasattr(self, 'primary_institution') or not self.primary_institution:
            warnings.append("❌ No primary institution selected")
            is_valid = False
            return is_valid, warnings
        
        # Get all institutions (primary + peers)
        all_institutions = [self.primary_institution] + self.selected_peers
        total_count = len(all_institutions)
        
        # Check institution count
        if total_count < 2:
            warnings.append(f"❌ Minimum 2 institutions required (currently: {total_count})")
            is_valid = False
        elif total_count > 4:
            warnings.append(f"❌ Maximum 4 institutions allowed (currently: {total_count})")
            is_valid = False
        
        # Load asset size data if available
        asset_sizes = self.get_institution_asset_sizes(all_institutions)
        
        if asset_sizes:
            # Check for significant size differences
            sizes = [size for size in asset_sizes.values() if size > 0]
            if len(sizes) >= 2:
                max_size = max(sizes)
                min_size = min(sizes)
                
                if max_size > min_size * 100:
                    # More than 100x difference
                    largest_inst = next(inst['name'] for inst in all_institutions 
                                    if asset_sizes.get(inst['rssd_id']) == max_size)
                    smallest_inst = next(inst['name'] for inst in all_institutions 
                                    if asset_sizes.get(inst['rssd_id']) == min_size)
                    warnings.append(
                        f"⚠️ Large asset size difference detected (>100x):\n"
                        f"   • {largest_inst}: ${max_size/1e9:.1f}B\n"
                        f"   • {smallest_inst}: ${min_size/1e9:.1f}B\n"
                        f"   This may make comparisons less meaningful."
                    )
                elif max_size > min_size * 50:
                    # 50-100x difference
                    warnings.append(
                        f"⚠️ Significant asset size difference (>50x):\n"
                        f"   Consider selecting more similarly-sized institutions."
                    )
        
        # Processing time warnings
        if total_count == 4:
            warnings.append(
                "⏱️ Processing 4 institutions may take 5-10 minutes depending on data size.\n"
                "   Consider starting with 2-3 institutions for faster results."
            )
        elif total_count == 3:
            warnings.append("⏱️ Processing 3 institutions typically takes 3-5 minutes.")
        
        # Data availability warning
        warnings.append(
            "📊 Note: Comparison will only include schedules present in all selected institutions."
        )
        
        return is_valid, warnings
   
    def get_institution_asset_sizes(self, institutions):
        """
        Get asset sizes for institutions (placeholder - would connect to actual data)
        In real implementation, this would query a database or API
        """
        # This is a placeholder - in production, you would:
        # 1. Query your bulk data for the most recent RC schedule
        # 2. Extract total assets (RCFD2170 or RCON2170)
        # 3. Cache the results
        
        # For now, return mock data for demonstration
        mock_sizes = {
            "112837": 434000000000,  # Capital One (~$434B)
            "451965": 1900000000000,  # Wells Fargo (~$1.9T)
            "852218": 3900000000000,  # JPMorgan Chase (~$3.9T)
            "480228": 3200000000000,  # Bank of America (~$3.2T)
            "476810": 2400000000000,  # Citibank (~$2.4T)
            "723112": 1800000000000,  # Fifth Third (~$1.8T)
            "504713": 1900000000000,  # U.S. Bank (~$1.9T)
            "817824": 560000000000,   # PNC (~$560B)
            "233031": 230000000000,   # Regions (~$230B)
        }
        
        asset_sizes = {}
        for inst in institutions:
            rssd_id = inst.get('rssd_id', '')
            if rssd_id in mock_sizes:
                asset_sizes[rssd_id] = mock_sizes[rssd_id]
            else:
                # For unknown institutions, generate a reasonable size for demo
                import random
                # Generate between $1B and $50B for smaller banks
                asset_sizes[rssd_id] = random.randint(1000000000, 50000000000)
        
        return asset_sizes
    

    def show_validation_warnings(self, warnings):
        """Display validation warnings in a dialog"""
        if not warnings:
            return True
        
        # Create warning dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Multi-Institution Comparison Validation")
        dialog.geometry("650x450")
        
        # Make it modal
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Header
        header_frame = tk.Frame(dialog, bg=self.primary_color, height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        tk.Label(
            header_frame,
            text="⚠️ Validation Warnings",
            font=('Arial', 16, 'bold'),
            bg=self.primary_color,
            fg='white'
        ).pack(expand=True)
        
        # Warning content
        content_frame = tk.Frame(dialog, padx=20, pady=10)
        content_frame.pack(fill='both', expand=True)
        
        # Scrollable text widget
        text_widget = tk.Text(
            content_frame,
            wrap='word',
            font=('Arial', 11),
            height=12,
            bg='#f5f5f5'
        )
        text_widget.pack(fill='both', expand=True)
        
        # Add warnings with formatting
        for warning in warnings:
            if warning.startswith("❌"):
                text_widget.insert('end', warning + '\n\n', 'error')
            elif warning.startswith("⚠️"):
                text_widget.insert('end', warning + '\n\n', 'warning')
            elif warning.startswith("⏱️"):
                text_widget.insert('end', warning + '\n\n', 'time')
            else:
                text_widget.insert('end', warning + '\n\n', 'info')
        
        # Configure tags
        text_widget.tag_configure('error', foreground='#d32f2f', font=('Arial', 11, 'bold'))
        text_widget.tag_configure('warning', foreground='#f57c00', font=('Arial', 11, 'bold'))
        text_widget.tag_configure('time', foreground='#1976d2', font=('Arial', 11))
        text_widget.tag_configure('info', foreground='#388e3c', font=('Arial', 11))
        
        text_widget.config(state='disabled')
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(content_frame, command=text_widget.yview)
        scrollbar.pack(side='right', fill='y')
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        # Button frame
        button_frame = tk.Frame(dialog, bg='#f0f0f0')
        button_frame.pack(fill='x', pady=10)
        
        user_choice = {'proceed': False}
        
        def on_proceed():
            user_choice['proceed'] = True
            dialog.destroy()
        
        def on_cancel():
            user_choice['proceed'] = False
            dialog.destroy()
        
        # Only show proceed button if validation passed (no errors)
        has_errors = any(w.startswith("❌") for w in warnings)
        
        if not has_errors:
            tk.Button(
                button_frame,
                text="Proceed with Comparison",
                command=on_proceed,
                bg=self.success_color,
                fg='white',
                font=('Arial', 12, 'bold'),
                padx=25,
                pady=10,
                cursor='hand2'
            ).pack(side='left', padx=(100, 10))
        
        tk.Button(
            button_frame,
            text="Go Back" if has_errors else "Cancel",
            command=on_cancel,
            bg='#757575' if has_errors else self.error_color,
            fg='white',
            font=('Arial', 12, 'bold'),
            padx=25,
            pady=10,
            cursor='hand2'
        ).pack(side='left', padx=10)
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - dialog.winfo_width()) // 2
        y = (dialog.winfo_screenheight() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        self.root.wait_window(dialog)
        
        return user_choice['proceed']
    
    def show_processing_time_estimate(self, institution_count):
        """Show estimated processing time before starting"""
        estimates = {
            2: ("2-3 minutes", "~94 files (2 × 47 schedules)"),
            3: ("3-5 minutes", "~141 files (3 × 47 schedules)"),
            4: ("5-10 minutes", "~188 files (4 × 47 schedules)")
        }
        
        time_est, file_est = estimates.get(institution_count, ("several minutes", "many files"))
        
        result = messagebox.askokcancel(
            "Processing Time Estimate",
            f"Processing {institution_count} institutions will take approximately {time_est}.\n\n" +
            f"📁 Files to process: {file_est}\n\n" +
            "The process includes:\n" +
            "• Loading and filtering data for each institution\n" +
            "• Matching line items across institutions\n" +
            "• Creating formatted comparison sheets\n" +
            "• Building summary and navigation\n\n" +
            "💡 Tip: Start with 2 institutions for faster initial results.\n\n" +
            "Would you like to continue?",
            icon='info'
        )
        
        return result
    
    # Update check_multi_institution_validation to use enhanced validation
    def check_multi_institution_validation(self):
        """Real-time validation feedback as user selects institutions"""
        warnings = []
        
        # Get total institutions
        total_institutions = 1 if hasattr(self, 'primary_institution') else 0
        total_institutions += len(self.selected_peers)
        
        # Basic count validation
        if total_institutions < 2:
            warnings.append(f"⚠️ Need at least 2 institutions (have {total_institutions})")
        elif total_institutions > 4:
            warnings.append(f"⚠️ Maximum 4 institutions (have {total_institutions})")
        
        # Check for size differences if we have enough institutions
        if hasattr(self, 'primary_institution') and len(self.selected_peers) >= 1:
            all_institutions = [self.primary_institution] + self.selected_peers
            asset_sizes = self.get_institution_asset_sizes(all_institutions)
            
            if asset_sizes:
                sizes = list(asset_sizes.values())
                if len(sizes) >= 2:
                    max_size = max(sizes)
                    min_size = min(sizes)
                    ratio = max_size / min_size if min_size > 0 else float('inf')
                    
                    if ratio > 100:
                        warnings.append("⚠️ Asset sizes differ by >100x")
                    elif ratio > 50:
                        warnings.append("⚠️ Asset sizes differ by >50x")
        
        # Processing time hint
        if total_institutions == 4:
            warnings.append("⏱️ 4 institutions: ~5-10 min processing")
        elif total_institutions == 3:
            warnings.append("⏱️ 3 institutions: ~3-5 min processing")
        
        # Update warning label
        if warnings:
            self.multi_warning_label.config(
                text="\n".join(warnings), 
                fg='orange'
            )
        else:
            if total_institutions >= 2:
                self.multi_warning_label.config(
                    text="✓ Valid selection - ready to process", 
                    fg=self.success_color
                )
            else:
                self.multi_warning_label.config(text="", fg='orange')

# ===== MAIN ENTRY POINT =====        
def main():
    """
    Main application entry point
    
    Creates the Tkinter root window and starts the application.
    """
    # Create root window
    root = tk.Tk()
    
    # Create application instance
    app = FIREAnalyzer(root)
    
    # Start the GUI event loop
    root.mainloop()


if __name__ == "__main__":
    main()